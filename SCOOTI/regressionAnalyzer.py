"""
regressionAnalyzer.py
=================================================================================
A collection of methods for analysis of metabolic objectives and flux predictions
"""


# packages
import pandas as pd
import pickle
import json
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import cobra
import openpyxl
import sys
#from .GeneralMethods.AnalysisKits import *
from SCOOTI.MatplotProp import CanvasStyle, PltProps, Significance
PltProps()
import warnings; warnings.simplefilter('ignore')
from statsmodels.stats.multitest import fdrcorrection
import scipy.stats as ss
import scipy.spatial.distance as ssd
import os
from tqdm.notebook import tqdm, trange
import networkx as nx
import matplotlib.patheffects as path_effects
import statsmodels.api as sm
from adjustText import adjust_text

# Regression models
#from SCOOTI.regressorCollection import *
from SCOOTI.regressorMetaLearner import *
# Set cobra solver to glpk in order to avoid err msg
config = cobra.Configuration()
config.solver = "glpk"

"""

███████ ██      ██    ██ ██   ██                      
██      ██      ██    ██  ██ ██                       
█████   ██      ██    ██   ███                        
██      ██      ██    ██  ██ ██                       
██      ███████  ██████  ██   ██                      
                                                      
                                                      
██████  ███████  █████  ██████  ██ ███    ██  ██████  
██   ██ ██      ██   ██ ██   ██ ██ ████   ██ ██       
██████  █████   ███████ ██   ██ ██ ██ ██  ██ ██   ███ 
██   ██ ██      ██   ██ ██   ██ ██ ██  ██ ██ ██    ██ 
██   ██ ███████ ██   ██ ██████  ██ ██   ████  ██████  
                                                      
---
Flux reading
- load_files_gen
- load_KOmat_gen
- unconstrained_models
- constrained_models
- load_multiObj_models
- load_geneKO_models
- fva_constrained_models
"""

# data loading function generating python generators
def load_files_gen(filenames):
    for filename in tqdm(filenames):
        yield (
                pd.read_csv(
                    filename
                    )
                )

# generater to load fluxes
def load_KOmat_gen(files, colnames, geneList_path='/home/daweilin/StemCell/unique_gene_list.mat'):


    # import scipy for reading .mat
    import scipy
    # load unique gene list
    geneList = scipy.io.loadmat(geneList_path)
    #tmp = scipy.io.loadmat('/nfs/turbo/umms-csriram/daweilin/fluxPrediction/Tzelepis2016/Recon3D_GeneKO/[Oct1720230322]CFR_ct1_obj1_data1_CFR-geneDel.mat')
    #geneList_path='/home/daweilin/StemCell/recon3d_unique_gene_list.mat'
    #print(geneList)
    # unique genes with unknown labels and wilde type columns
    key = 'geneList' if 'recon' in geneList_path else 'unique_gene_list'
    genes_columns = [
            g[0][0] if len(g[0])>0 else 'unknown' for g in geneList[key]
            ]
    # add wild type columns
    genes_columns = np.append('WT', genes_columns)

    def mat_to_df_process(path, genes_columns, colnames):
        # load knockout mat files
        mat = scipy.io.loadmat(path)
        # make reaction list
        rxns_index = [
                r[0][0] if len(r[0])>0 else 'unknown' for r in mat['rxn_labels']
                ]
        # convert the result into dataframe
        geneko_flux = pd.DataFrame(
                mat['geneko_flux'],
                index=rxns_index,
                columns=genes_columns
                )
        geneko_flux.columns = pd.Series(geneko_flux.columns).apply(lambda x: f'{colname}_{x}')
        return geneko_flux


    for file, colname in zip(files, colnames):
        yield mat_to_df_process(file, genes_columns, colname)




# ideally optimal fluxes with single objectives
def unconstrained_models(root_path, norm=False, return_variables=True, medium='DMEMF12', biomass_reaction_name='biomass_objective'):
    
    """

    Description
    -----------
    Unconstrained modeling of fluxes with different objectives

    Arguments
    ---------
    root_path (string): path to access all the files
    norm (bool): if normalize the data or not
    return_variables (bool): True for return fluxes or False for objective values
    medium (string): medium used for input metabolites e.g. DMEMF12 or KSOM

    Returns
    -------
    res_df (optional, pandas.DataFrame): reaction fluxes excluding objective values
    dm_df (optional, pandas.DataFrame): fluxes of demand reactions of candidate metabolites


    """

    print('Start processing the unconstrained models...')
    
    flux_files = [f for f in os.listdir(root_path) if os.path.isfile(os.path.join(root_path, f))]
    df_collect = []
    for file in flux_files:
        if '_metadata' in file:
            try:
                with open(root_path+file, 'rb') as J:
                    JSON = json.load(J) # load metadata
                    medium_from_model = JSON['medium'] # get medium
                    colname = np.array(JSON['obj'])[np.array(JSON['obj_c'])!=0][0] # make sure we got 
                if medium_from_model==medium and colname!='':
                    df = pd.read_csv(
                            root_path+'{0}_fluxes.csv.gz'.format(
                                file.split('_metadata')[0]
                                )
                            )
                    row = df['rxns']
                    df = pd.DataFrame({colname:df.iloc[:,-1].to_numpy()})
                    df_collect.append(df)
            except:
                print('No file', root_path+'{0}_fluxes.csv.gz'.format(file.split('_metadata')[0]))
    
    # stack into a large table
    res_df = pd.concat((df_collect),axis=1)
    res_df.index = row
    res_df = res_df.loc[:,~res_df.T.duplicated(keep='first')]
    
    # objective fluxes
    obj_df = res_df[res_df.index.str.contains('Obj')]
    # return reaction fluxes
    if return_variables==True:
        if norm==True:
            #obj_df = res_df[res_df.index.str.contains('Obj')]
            res_df = res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}')==0].div(obj_df.to_numpy(), axis=1)
        else:
            res_df = res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}')==0]

        res_df = res_df[np.array(res_df.columns)[np.array(res_df.columns)!='']]
        
        print('Returning reaction fluxes...')
        print('Finishing up the data loading of unconstrained models...')

        return res_df

    # return demand fluxes (objective fluxes)
    else:
        if norm==True:
            dm_df = res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}')==1].div(obj_df.to_numpy(), axis=1)
        else:
            dm_df = res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}')==1]

        dm_df = dm_df[np.array(dm_df.columns)[np.array(dm_df.columns)!='']]
        
        print('Returning demand fluxes...')
        print('Finishing up the data loading of unconstrained models...')
    
        return dm_df

# fluxes simulated with constraints
def constrained_models(
        root_path, CFR_paraScan=False, DFA_paraScan=False, randomScan=False, norm=True,
        file_suffix='_fluxes.csv.gz', input_path_pattern='',
        biomass_reaction_name='biomass_objective',
        CFR_k=[10, 1, 0.1, 0.01, 0.001],
        CFR_r=[10, 1, 0.1, 0.01, 0.001], 
        DFA_k=[10, 1, 0.1, 0.01, 0.001],
        ):


    """

    Description
    -----------
    Organize constrained modeling of fluxes without objectives

    Arguments
    ---------
    root_path (string): path to access all the files
    CFR_paraScan (list of float): CFR parameters of interest used to output fluxes
    DFA_paraScan (list of float): DFA parameters of interest used to output fluxes
    randomScan (bool): read all models without considering parameters
    norm (bool): if normalize the data by the summation of fluxes or not
    file_suffix (string): the suffix of filenames to read

    Returns
    -------
    res_df (pandas.DataFrame): reaction fluxes with parameters of interest


    """
    
    print('Start processing the data of constrained models...')
    flux_files = [
            f for f in os.listdir(root_path) if os.path.isfile(os.path.join(root_path, f))
            ]
    df_collect = []
    colnames = []

    print('Start collecting filenames...')
    for file in tqdm(flux_files):
        if '_metadata' in file:
            with open(root_path+file, 'rb') as J:
                JSON = json.load(J)

                # Checkpoint for parameters
                if randomScan==True:
                    switch = 1
                elif CFR_paraScan==True:
                    JSON_CFR_kappa, JSON_CFR_rho = JSON['CFR_kappa'], JSON['CFR_rho']
                    switch = JSON_CFR_kappa in CFR_k and JSON_CFR_rho in CFR_r
                elif DFA_paraScan==True:
                    JSON_DFA_kappa = JSON['DFA_kappa']
                    switch = JSON_DFA_kappa in DFA_k
                else:
                    switch = 1

                # Checkpoint for data input
                if input_path_pattern!='':
                    JSON_input_path = JSON['input_path']
                    sel_data = input_path_pattern in JSON_input_path
                    switch = sel_data*switch

                if switch:
                    objMet = np.array(JSON['obj'])[np.array(JSON['obj_c'])!=0][0]
                    cellType = JSON['input_path']
                    # get celltype
                    # split by .xlsx if there is .xlsx in the path string
                    cellType = cellType.split('/')[-1].split('.xlsx')[0]
                    # split by .csv if there is .csv in the string
                    cellType = cellType.split('/')[-1].split('.csv')[0]
                    # file names of flux data
                    fn = '{0}{1}'.format(
                            file.split('_metadata')[0], file_suffix
                                )
                    # skip this part if the file of flux data doesnt exist
                    if fn in flux_files:
                        
                        # make column names with corresponding parameters
                        if CFR_paraScan==True:
                            cellType = f"{cellType}_k{JSON['CFR_kappa']}_r{JSON['CFR_rho']}"
                        elif DFA_paraScan==True:
                            cellType = f"{cellType}_dk{JSON['DFA_kappa']}"
                        
                        # only get fluxes without objective
                        if objMet=='' or objMet=='gh':
                            try:
                                df_collect.append(
                                    root_path+'{0}_fluxes.csv.gz'.format(
                                        file.split('_metadata')[0]
                                        )
                                    )
                                colnames.append('rxns')
                                if cellType=='':
                                    cellType = file.split('_')[3]
                                colnames.append(cellType)
                            except:
                                print('[WARNING] No corresponding csv files...')
                    # skip reading the file
                    else:
                        pass
    print('Start merging tables')
    # using generator to read csv and concatenate into a large table
    res_df = pd.concat(load_files_gen(df_collect),axis=1)
    res_df.columns = colnames # replace column name with cell type name
    res_df = res_df.loc[:,~res_df.columns.duplicated(keep='first')] # drop duplicate columns
    res_df.index = res_df['rxns'] # move reaction list to index of the table
    res_df = res_df.iloc[:, 1:] # remove the rxn list from the table
    print('debug', res_df.sum().sum())
    if norm==True:
        # normalize with the sum of all the reactions
        res_df = res_df.div(res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}|WT')==0].abs().sum(axis=0), axis=1)
        res_df = res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}')==0]
    else:
        res_df = res_df[res_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}')==0] # remove objective and demand reactions

    print(res_df)
    print('Finishing up loading the constrained models...')

    return res_df
 

# function for reading fluxes modeled with multiobjective functions
def load_multiObj_models(
        root_path, medium='KSOM', return_variables=True, norm=False,
        CFR_paraScan=False, DFA_paraScan=False, randomScan=False,
        topology_use=False,
        file_suffix='_fluxes.csv.gz',
        ind_labels=False,
        biomass_reaction_name='biomass_objective',
        CFR_k=[10, 1, 0.1, 0.01, 0.001],
        CFR_r=[10, 1, 0.1, 0.01, 0.001], 
        DFA_k=[10, 1, 0.1, 0.01, 0.001],
        ):
 

    """

    Description
    -----------
    Organize any types of models

    Arguments
    ---------
    root_path (string): path to access all the files
    medium (string): medium used for input metabolites e.g. DMEMF12 or KSOM
    return_variables (bool): True for return fluxes or False for objective values
    norm (bool): if normalize the data by the summation of fluxes or not
    CFR_paraScan (bool): if filtering models with CFR parameters of interest
    DFA_paraScan (bool): if filtering models with DFA parameters of interest
    randomScan (bool): read all models without considering parameters
    topology_use (bool): remove fluxes of candidate metabolites but biomass objective
    CFR_k (list of floats): CFR kappa of interest used to output fluxes
    CFR_r (list of floats): CFR rho of interest used to output fluxes
    CFR_r (list of floats): DFA parameters of interest used to output fluxes
    file_suffix (string): the suffix of filenames to read

    Returns
    -------
    res_df (pandas.DataFrame): reaction fluxes with parameters of interest


    """



    print('Start processing the data of constrained models...')
    # get file names
    flux_files = [f for f in os.listdir(root_path) if os.path.isfile(os.path.join(root_path, f))]
    df_collect = []
    colnames = []
    print('Start collecting results...')

    for i, file in tqdm(enumerate(flux_files)):
        if '_metadata' in file:
            with open(root_path+file, 'rb') as J:
                JSON = json.load(J)
                if randomScan==True:
                    switch = 1
                elif CFR_paraScan==True:
                    JSON_CFR_kappa, JSON_CFR_rho = JSON['CFR_kappa'], JSON['CFR_rho']
                    switch = JSON_CFR_kappa in CFR_k and JSON_CFR_rho in CFR_r
                elif DFA_paraScan==True:
                    JSON_DFA_kappa = JSON['DFA_kappa']
                    switch = JSON_DFA_kappa in DFA_k
                else:
                    switch = 1
                if switch:
                    objMet = np.array(JSON['obj'])[np.array(JSON['obj_c'])!=0][0]
                    #cellType = JSON['upStage']
                    cellType = JSON['input_path']
                    cellType = cellType.split('/')[-1].split('.xlsx')[0]
                    cellType = cellType.split('.csv')[0]

                    # get models without objectives
                    if JSON['medium']==medium:
                        fn = '{0}{1}'.format(
                                    file.split('_metadata')[0], file_suffix
                                    )
                        if fn in flux_files:
                            df_collect.append(root_path+fn)
                            colnames.append('rxns')
                            if cellType=='':
                                cellType = file.split('_')[3]
                            # make column names with corresponding parameters
                            if CFR_paraScan==True:
                                cellType = f"{cellType}_k{JSON['CFR_kappa']}_r{JSON['CFR_rho']}"
                            elif DFA_paraScan==True:
                                cellType = f"{cellType}_dk{JSON['DFA_kappa']}"
                            # make index labels for cells
                            if ind_labels==True:
                                cellType = f'{cellType}_{i}'
                            colnames.append(cellType)
                        else:
                            pass

    print('Start merging tables')
    # using generator to read csv and concatenate into a large table
    cell_df = pd.concat(load_files_gen(df_collect),axis=1)
    print(cell_df)
    cell_df.columns = colnames # replace column name with cell type name
    cell_df = cell_df.loc[:,~cell_df.columns.duplicated(keep='first')] # drop duplicate columns
    cell_df.index = cell_df['rxns'].fillna('') # move reaction list to index of the table
    cell_df = cell_df.iloc[:, 1:] # remove the rxn list from the table
    print('debug', cell_df.sum().sum())

    if norm==True:
        cell_df = cell_df.div(cell_df[cell_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}|WT')==0].abs().sum(axis=0), axis=1)
    if return_variables==True:
        cell_df = cell_df[cell_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}|WT')==0] # remove objective and demand reactions
    elif topology_use==True:
        cell_df = cell_df[cell_df.index.str.contains('Obj|_demand|WT')==0] # remove objective and demand reactions
    else: # return objective values
        cell_df = cell_df[cell_df.index.str.contains('_demand|biomass_objective|WT')] # remove objective and demand reactions
    #print(cell_df)
    print('Finishing up loading the constrained models based on omics datasets...')
    
    return cell_df


# function for reading fluxes modeled with multiobjective functions
def load_geneKO_models(
        root_path, medium='KSOM', return_variables=True, norm=False,
        CFR_paraScan=False, DFA_paraScan=False, randomScan=False,
        topology_use=False, geneList_path='',
        file_suffix='_CFR-geneDel.mat',
        ind_labels=False,
        biomass_reaction_name='biomass_objective',
        CFR_k=[10, 1, 0.1, 0.01, 0.001],
        CFR_r=[10, 1, 0.1, 0.01, 0.001], 
        DFA_k=[10, 1, 0.1, 0.01, 0.001],
        ):
    
    print('Start processing the data of constrained models...')
    # get file names
    flux_files = [f for f in os.listdir(root_path) if os.path.isfile(os.path.join(root_path, f))]
    df_collect = []
    colnames = []
    print('Start collecting results...')

    for i, file in tqdm(enumerate(flux_files)):
        if '_metadata' in file:
            with open(root_path+file, 'rb') as J:
                JSON = json.load(J)
                if randomScan==True:
                    switch = 1
                elif CFR_paraScan==True:
                    JSON_CFR_kappa, JSON_CFR_rho = JSON['CFR_kappa'], JSON['CFR_rho']
                    switch = JSON_CFR_kappa in CFR_k and JSON_CFR_rho in CFR_r
                elif DFA_paraScan==True:
                    JSON_DFA_kappa = JSON['DFA_kappa']
                    switch = JSON_DFA_kappa in DFA_k
                else:
                    switch = 1

                if switch:
                    print('in')
                    objMet = np.array(JSON['obj'])[np.array(JSON['obj_c'])!=0][0]
                    cellType = JSON['input_path']
                    cellType = cellType.split('/')[-1].split('.xlsx')[0]
                    cellType = cellType.split('.csv')[0]
                    
                    # make file names
                    fn = '{0}{1}'.format(
                                file.split('_metadata')[0], file_suffix
                                )
                    # get models without objectives
                    if JSON['medium']==medium and fn in flux_files:
                        print('debug', fn)
                        if fn in flux_files:
                            df_collect.append(root_path+fn)
                            #colnames.append('rxns')
                            if cellType=='':
                                cellType = file.split('_')[3]
                            # make column names with corresponding parameters
                            if CFR_paraScan==True:
                                cellType = f"{cellType}_k{JSON['CFR_kappa']}_r{JSON['CFR_rho']}"
                            elif DFA_paraScan==True:
                                cellType = f"{cellType}_dk{JSON['DFA_kappa']}"
                            # make index labels for cells
                            if ind_labels==True:
                                cellType = f'{cellType}_{i}'
                            colnames.append(cellType)
                        else:
                            pass

    print('Start merging tables')
    print(df_collect[0])
    # using generator to read csv and concatenate into a large table
    cell_df = pd.concat(load_KOmat_gen(df_collect, colnames, geneList_path=geneList_path),axis=1)
    print(cell_df)
    #cell_df.columns = pd.Series(cell_df).apply(lambda x: f'{colnames}_{x}') # replace column name with cell type name
    cell_df = cell_df.loc[:,~cell_df.columns.duplicated(keep='first')] # drop duplicate columns
    #cell_df.index = cell_df['rxns'].fillna('') # move reaction list to index of the table
    #cell_df = cell_df.iloc[:, 1:] # remove the rxn list from the table
    print('debug', cell_df.sum().sum())

    if norm==True:
        cell_df = cell_df.div(cell_df[cell_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}|WT')==0].abs().sum(axis=0), axis=1)
    if return_variables==True:
        cell_df = cell_df[cell_df.index.str.contains(f'Obj|_demand|{biomass_reaction_name}|WT')==0] # remove objective and demand reactions
    elif topology_use==True:
        cell_df = cell_df[cell_df.index.str.contains('Obj|_demand|WT')==0] # remove objective and demand reactions
    else: # return objective values
        cell_df = cell_df[cell_df.index.str.contains(f'_demand|{biomass_reaction_name}|WT')] # remove objective and demand reactions
    #print(cell_df)
    print('Finishing up loading the constrained models based on omics datasets...')
    
    return cell_df



# reading fluxes of FVA ranges
def fva_constrained_models(root_path):
    """
    Data loading: constrained modeling of HumanAtlas without objectives
    """
    
    print('Start processing the data of constrained models...')
    flux_files = [f for f in os.listdir(root_path) if os.path.isfile(os.path.join(root_path, f))]
    df_collect = []
    colnames = []

    print('Start collecting filenames...')
    for file in tqdm(flux_files):
        if '_metadata' in file:
            with open(root_path+file, 'rb') as J:
                JSON = json.load(J)
                objMet = np.array(JSON['obj'])[np.array(JSON['obj_c'])!=0][0]
                cellType = JSON['input_path']
                # get celltype
                # split by .xlsx if there is .xlsx in the path string
                cellType = cellType.split('/')[-1].split('.xlsx')[0]
                # split by .csv if there is .csv in the string
                cellType = cellType.split('/')[-1].split('.csv')[0]
            
            # only get fluxes without objective
            if objMet=='':
                df_collect.append(
                    root_path+'{0}_FVA.csv.gz'.format(
                        file.split('_metadata')[0]
                        )
                    )
                colnames.append('rxns')
                if cellType=='':
                    cellType=file.split('_')[3]
                # append twice
                colnames.append(f'{cellType}_max')
                colnames.append(f'{cellType}_min')
    
    print('Start merging tables')
    # using generator to read csv and concatenate into a large table
    res_df = pd.concat(load_files_gen(df_collect),axis=1)
    res_df.columns = colnames # replace column name with cell type name
    res_df = res_df.loc[:,~res_df.columns.duplicated(keep='first')] # drop duplicate columns
    res_df.index = res_df['rxns'] # move reaction list to index of the table

    res_df = res_df.iloc[:, 1:] # remove the rxn list from the table
    #print('debug', res_df.sum().sum())
    res_df = res_df[res_df.index.str.contains('Obj|_demand|WT')==0] # remove objective and demand reactions
    print(res_df)
    print('Finishing up loading the constrained models...')
    return res_df

"""

███    ███  ██████  ██████  ███████ ██                                         
████  ████ ██    ██ ██   ██ ██      ██                                         
██ ████ ██ ██    ██ ██   ██ █████   ██                                         
██  ██  ██ ██    ██ ██   ██ ██      ██                                         
██      ██  ██████  ██████  ███████ ███████                                    
                                                                               
                                                                               
██████  ██████   ██████   ██████ ███████ ███████ ███████ ██ ███    ██  ██████  
██   ██ ██   ██ ██    ██ ██      ██      ██      ██      ██ ████   ██ ██       
██████  ██████  ██    ██ ██      █████   ███████ ███████ ██ ██ ██  ██ ██   ███ 
██      ██   ██ ██    ██ ██      ██           ██      ██ ██ ██  ██ ██ ██    ██ 
██      ██   ██  ██████   ██████ ███████ ███████ ███████ ██ ██   ████  ██████  
                                                                               
                                                                               
---
Model processing
- remove_all_zeros
- load_pageRankData
- model_training
- matrixProcessing
- topology
- spearmanCI
- metabolite_converter

"""



# remove the zeros from the datatables
def remove_all_zeros(unconstrained_models, constrained_models):
    
    # rearrange datasets
    ideal_model = unconstrained_models[unconstrained_models.index.isin(constrained_models.index)]
    con1 = (ideal_model.sum(axis=1)==0)
    con2 = (constrained_models.sum(axis=1)==0)
    unconstrained_clean = ideal_model[(con1 & con2)==0].replace([np.inf, -np.inf, np.nan], [0,0,0])
    constrained_clean = constrained_models[(con1 & con2)==0].replace([np.inf, -np.inf, np.nan], [0,0,0])
    
    print(
            'Check the size of the models:',
            unconstrained_clean.shape,
            constrained_clean.shape
            )
    
    return unconstrained_clean, constrained_clean


# load data
def load_pageRankData(suffix, cell_name):
    
    # read pagerank data
    pg_uncon_models = pd.read_csv('/home/daweilin/StemCell/Project_mESC_JinZhang/pageRankAnalysis/unconstraint_pageRank_{0}.csv'.format(suffix),
            index_col=0)
    pg_con_models = pd.read_csv('/home/daweilin/StemCell/Project_mESC_JinZhang/pageRankAnalysis/{1}_pageRank_{0}.csv'.format(suffix, cell_name),
            index_col=0)

    return pg_uncon_models, pg_con_models

# train regression models
def model_training(uncon_models, con_models, cell_name, suffix, pg_uncon_models=[], pg_con_models=[], input_type='flux', cluster_path=''):
    
    # Part1: regression with fluxes
    if input_type=='flux':
        if cluster_path=='':
            # make regression with constrained CFR results
            reg = regression_methods(uncon_models, con_models)
            sl_res = reg.SuperLearner()
        else:
            # cluster labels should be integers
            # make regression with constrained CFR results
            reg = regression_methods(
                    uncon_models,
                    con_models,
                    cluster_label_path=cluster_path,
                    )
            sl_res = reg.clusterSuperLearner()
        for sl_ in sl_res:
            sl_.columns = con_models.columns

        # integrate models together
        #integrate_res = sl_res[1].copy()
        #for col in sl_res[1].columns:
        #    integrate_res[col] = sl_res[1][col].mul(sl_res[0][col][0])+sl_res[2][col].mul(sl_res[0][col][1])
        #    #if isinstance(integrate_res, list):
        #    #    integrate_res = integrate_res[0]
        #    #    integrate_res = pd.DataFrame(integrate_res)
        #
        #integrate_res.columns = con_models.columns
        #RFElm_res = sl_res[1]
        #RFElm_res.columns = con_models.columns
        #lasso_res = sl_res[2]
        #lasso_res.columns = con_models.columns
    
    
        return sl_res#RFElm_res, lasso_res, integrate_res
    
    # Part 2:
    # PageRank as input
    elif input_type=='pageRank':
        
        # read PageRank of unconstrained models and NCI60 datasets
        ideal_model = pg_uncon_models[pg_uncon_models.index.isin(pg_con_models.index)]
        con1 = (ideal_model.sum(axis=1)==0)
        con2 = (pg_con_models.sum(axis=1)==0)
        pg_uncon_models = ideal_model[(con1 & con2)==0].replace([np.inf, -np.inf, np.nan], [0,0,0])
        pg_con_models = pg_con_models[(con1 & con2)==0].replace([np.inf, -np.inf, np.nan], [0,0,0])
        
        print('Check the size of ccle:', pg_uncon_models.shape, pg_con_models.shape)
        
        
        # make regression with ccle/NCI-60 CFR
        reg = regression_methods(pg_uncon_models, pg_con_models)
        pg_sl_res = reg.SuperLearner()
        # pg_sl_res = reg.clusterSuperLearner()
        
        pg_integrate_res = pg_sl_res[1].copy()
        for col in pg_sl_res[1].columns:
            pg_integrate_res[col] = pg_sl_res[1][col].mul(pg_sl_res[0][col][0])+pg_sl_res[2][col].mul(pg_sl_res[0][col][1])
            #if isinstance(integrate_res, list):
            #    pg_integrate_res = pg_integrate_res[0]
            #    pg_integrate_res = pd.DataFrame(pg_integrate_res)
        pg_integrate_res.columns = pg_con_models.columns
        pg_RFElm_res = pg_sl_res[1]
        pg_RFElm_res.columns = pg_con_models.columns
        pg_lasso_res = pg_sl_res[2]
        pg_lasso_res.columns = pg_con_models.columns
        
     
        return pg_RFElm_res, pg_lasso_res, pg_integrate_res
    
    else:# input_type=='mixture': # mixture of pageRank and fluxes
        
        # Part 3.Stack pagerank and fluxes input together
        print(pg_con_models.shape)
        print(con_models.shape)


        st_uncon_models = pd.concat((
            pg_uncon_models,
            uncon_models
            ), axis=0)
        st_con_models = pd.concat((
            pg_con_models,
            con_models
            ), axis=0)

        st_uncon_models.index = np.arange(len(st_uncon_models.index))
        st_con_models.index = np.arange(len(st_con_models.index))

        
        # make regression with ccle/NCI-60 CFR
        reg = regression_methods(st_uncon_models, st_con_models)
        st_sl_res, st_RFElm_res, st_lasso_res = reg.SuperLearner()
        
        st_integrate_res = st_RFElm_res.copy()
        for col in st_RFElm_res.columns:
            st_integrate_res[col] = st_RFElm_res[col].mul(st_sl_res[col][0])+st_lasso_res[col].mul(st_sl_res[col][1])
        

        st_integrate_res.columns = st_con_models.columns
        st_RFElm_res.columns = st_con_models.columns
        st_lasso_res.columns = st_con_models.columns

        return st_RFElm_res, st_lasso_res, st_integrate_res

    


# normalize data
def matrixProcessing(input_dict, norm='max_norm'):

    
    # empty list for collecting labels and dataframes
    labels = np.array([])
    dfcollect = []
    # iterate thru all dictionaries of different cell types
    for k in input_dict.keys():
        labels = np.append(labels, list([k]*len(input_dict[k].columns)))
        dfcollect.append(input_dict[k])

    # concat dataframes
    pca_df = pd.concat((dfcollect), axis=1)
    pca_df = pca_df.fillna(0) # replace nan with 0
    if norm=='max_norm':
        pca_df = pca_df.div(pca_df.sum(axis=0)) # normalize by the summation of each column
    elif norm=='binary':
        pca_df = pca_df>0
    elif norm=='rank':
        pca_df = pca_df[pca_df.any(axis=1)].rank(axis=0)
    else:
        pca_df = pca_df

    return pca_df, labels

#def uniform_flux_sampling(df):
#    
#    def sampling(arr):
#        s = np(arr[0], arr[-1], 100)
#        random.sample(s, 1)
#
#    df.apply(lambda x: np.random.uniform(x[0], x[1]), axis=1)   df.apply()

# 
def topology(df, model_path='/home/daweilin/StemCell/cancer_model.mat'):
    dfcopy = df.copy()
    model = cobra.io.load_matlab_model(model_path)
    reversibility = np.array([rxn.reversibility for rxn in model.reactions])
    S = cobra.util.array.create_stoichiometric_matrix(model)
    print(S.shape)
    print(df.shape)
    S2m = np.concatenate((np.matmul(S, np.identity(S.shape[1])), np.matmul(-S, np.diag(reversibility))), axis=1)
    splus2m = 0.5*(np.abs(S2m)+S2m)
    sminus2m = 0.5*(np.abs(S2m)-S2m)

    def massFlowGraph(fba):
        v2m = 0.5*np.append((np.abs(fba)+fba).T, (np.abs(fba)-fba).T)

        jvplus = np.matmul(splus2m, v2m)
        jvminus = np.matmul(sminus2m, v2m)

        mfg = np.matmul(np.matmul(np.matmul(splus2m, np.diag(v2m)).T,
                              np.linalg.pinv(np.diag(jvplus))),
                    np.matmul(sminus2m, np.diag(v2m)))
        
        Adj = pd.DataFrame(mfg)

        # Create graph, A.astype(bool).tolist() or (A / A).tolist() can also be used.
        gr = nx.from_numpy_matrix(mfg, create_using=nx.DiGraph())
        
        pagerank = nx.pagerank(gr, alpha=0.85)
        
        return pagerank
    
    def No_to_label(labels, input_dict):
        labels = np.append(labels, [f'{r}_reverse' for r in labels])
        graph_rxn_map = {num:labels[num] for num in input_dict.keys()}
        res_dict = {graph_rxn_map[k]:input_dict[k] for k in input_dict.keys()}
        return res_dict
    
    def process(token):
        return token['text']
    
    nx_feature = pd.DataFrame({
        i:massFlowGraph(df.iloc[:,i]) for i in trange(df.shape[1], desc='PageRank calculation')
        })
     
    return nx_feature


# confidence interval of spearman correlation
def spearmanCI(r, num):
    import math
    stderr = 1.0 / math.sqrt(num - 3)
    delta = 1.96 * stderr
    lower = math.tanh(math.atanh(r) - delta)
    upper = math.tanh(math.atanh(r) + delta)
    
    return (lower, upper)

# function for mapping metabolites to common IDs (like PubChem)
def metabolite_converter(met_list):
    # metabolite converter
    import requests
    url = 'http://api.xialab.ca/mapcompounds'
    
    req_str = ";".join(met_list)
    
    payload = '{\n\t"queryList\": \"'+req_str+';\",\n\t\"inputType\": \"name\"\n}'
    
    headers = {
            "Content-Type": "application/json",
            "cache-control": "no-cache",
            }
    
    response = requests.request('POST', url, data=payload, headers=headers)

    import json
    model_match = pd.DataFrame(json.loads(response.text))
    return model_match

# predict fluxes based on regression coefficients
def flux_reconstruction(
        coef_sel,
        root_path='/nfs/turbo/umms-csriram/daweilin/fluxPrediction/unconstrained_models/pfba/KSOM/',
        medium='KSOM',
        ):

    # Confirm metabolic functions selected by the objectives
    # sum(coef*metabolic graph)--> score of reaction activity
    # import unconstrained models
    uncon_res = unconstrained_models(root_path, norm=True, medium=medium)
    obj_res = unconstrained_models(root_path, norm=True, return_variables=False, medium=medium)
    uncon_res = uncon_res[uncon_res.columns[uncon_res.any(axis=0)]]
    # get selected objectives
    rxn_sel = uncon_res[coef_sel.index]#[coef_sel.index.isin(['gthrd', 'accoa', 'gly', 'gln-L'])].index]
    obj_sel = obj_res[coef_sel.index]
    # calculate fluxes
    weight_rxn = rxn_sel @ coef_sel#.div(coef_sel.sum(axis=0), axis=1)
    weight_obj = obj_sel @ coef_sel

    return weight_rxn, weight_obj


"""

██    ██ ██ ███████ ██    ██  █████  ██      ██ ███████  █████  ████████ ██  ██████  ███    ██ 
██    ██ ██ ██      ██    ██ ██   ██ ██      ██    ███  ██   ██    ██    ██ ██    ██ ████   ██ 
██    ██ ██ ███████ ██    ██ ███████ ██      ██   ███   ███████    ██    ██ ██    ██ ██ ██  ██ 
 ██  ██  ██      ██ ██    ██ ██   ██ ██      ██  ███    ██   ██    ██    ██ ██    ██ ██  ██ ██ 
  ████   ██ ███████  ██████  ██   ██ ███████ ██ ███████ ██   ██    ██    ██  ██████  ██   ████ 
                                                                                               

██████   █████  ███████ ██  ██████     ███████ ████████  █████  ████████ ███████ 
██   ██ ██   ██ ██      ██ ██          ██         ██    ██   ██    ██    ██      
██████  ███████ ███████ ██ ██          ███████    ██    ███████    ██    ███████ 
██   ██ ██   ██      ██ ██ ██               ██    ██    ██   ██    ██         ██ 
██████  ██   ██ ███████ ██  ██████     ███████    ██    ██   ██    ██    ███████ 
                                                                                 
                                                                                 


---
Visualization with basic statistical methods
- stars
- mets_category
- add_median_labels
- heatmap_pvalues
- coef_distance_to_biomassObj
- sankey_plot
- parallel_plots
- lollipop_fig
- multiGrids_fig
- boxplot_fig

"""

# labels for pvalues
def stars(p):
    if p < 0.0001:
        return "****"
    elif (p < 0.001):
        return "***"
    elif (p < 0.01):
        return "**"
    elif (p < 0.05):
        return "*"
    else:
        return "n.s."



# metabolites as single objective functions and corresponding group of metabolites
def mets_category():

    mets_category =  {'gh':'other','atp':'nucleotide','nadh':'nucleotide','nadph':'nucleotide','amet':'other','gthox':'other','gthrd':'other','nmn':'nucleotide','accoa':'other','ala-L':'amino_acid','amp':'nucleotide','arg-L':'amino_acid','asn-L':'amino_acid','asp-L':'amino_acid','chsterol':'lipid','clpn_hs':'phospholipid','cmp':'nucleotide','cys-L':'amino_acid','dag_hs':'lipid','damp':'nucleotide','dcmp':'nucleotide','dgmp':'nucleotide','dtmp':'nucleotide','gln-L':'amino_acid','glu-L':'amino_acid','gly':'amino_acid','glygn1':'other','gmp':'nucleotide','h2o':'other','his-L':'amino_acid','ile-L':'amino_acid','leu-L':'amino_acid','lpchol_hs':'phospholipid','lys-L':'amino_acid','mag_hs':'lipid','met-L':'amino_acid','pa_hs':'phospholipid','pail_hs':'phospholipid','pchol_hs':'phospholipid','pe_hs':'phospholipid','phe-L':'amino_acid','pro-L':'amino_acid','ps_hs':'phospholipid','ser-L':'amino_acid','sphmyln_hs':'phospholipid','tag_hs':'lipid','thr-L':'amino_acid','trp-L':'amino_acid','tyr-L':'amino_acid','ump':'nucleotide','val-L':'amino_acid','xolest_hs':'phospholipid'}
    
    return mets_category


# calculate median values
def add_median_labels(ax, fmt='.2e'):
    lines = ax.get_lines()
    boxes = [c for c in ax.get_children() if type(c).__name__ == 'PathPatch']
    lines_per_box = int(len(lines) / len(boxes))
    for median in lines[4:len(lines):lines_per_box]:
        x, y = (data.mean() for data in median.get_data())
        # choose value depending on horizontal or vertical plot orientation
        value = x if (median.get_xdata()[1] - median.get_xdata()[0]) == 0 else y
        text = ax.text(x, y, f'{value:{fmt}}', ha='center', va='center',  #f'{value:{fmt}}'
                       fontweight='bold', color='white', fontsize=10)
        # create median-colored border around white text for contrast
        text.set_path_effects([
            path_effects.Stroke(linewidth=3, foreground=median.get_color()),
            path_effects.Normal(),
        ])


# visualization of p-values (multiple comparisons) with heatmap
def heatmap_pvalues(pca_df, mets_category, labels, col_str, prefix, sel_samples=[]):

    # prepare dataframe for boxplots
    dfs = pca_df[pca_df.any(axis=1)].copy()
    dfs.columns = labels
    dfs['Mets_type'] = dfs.index.map(mets_category)
    dfs['tmp'] = dfs[dfs.columns[dfs.columns==col_str]].median(axis=1)
    dfs = dfs.sort_values(by=['Mets_type','tmp'])
    dfs = dfs.drop(columns=['Mets_type', 'tmp'])
    print(dfs)
    pvalues = {}
    discard = []
    if len(sel_samples)>0:
        cols = sel_samples
    else:
        cols = dfs.columns.unique()
    for col1 in cols:
        # col that has been scanned
        discard.append(col1)
        for col2 in dfs.columns.unique():
            if col2 not in discard:
                # compare pvalues
                ps = []
                for i in range(len(dfs)):
                   _, p = ss.mannwhitneyu(dfs[col1].iloc[i,:], dfs[col2].iloc[i,:])
                   ps.append(p)
                
                # save pvalues in a comparison
                pvalues[f'{col1} vs {col2}'] = ps

    plot_df = pd.DataFrame(pvalues)
    plot_df.index = dfs.index
    plot_df = plot_df[(plot_df<0.05).any(axis=1)]
    print(plot_df)
    #plot_df = plot_df>0.05
    #fig, ax = plt.subplots(1,1,figsize=(len(plot_df.columns.unique())//2, len(plot_df.index)))
    cmap = sns.color_palette("vlag", as_cmap=True)
    #g = sns.heatmap(plot_df, cmap=cmap, square=True, cbar_kws={'shrink': 0.6}, center=0)
    kws = dict(cbar_kws=dict(ticks=[0, 0.05], orientation='horizontal'), figsize=(1+len(plot_df.columns.unique())//2, len(plot_df)//2+1))
    #g = sns.clustermap(plot_df, cmap='Greys', **kws)#cbar_kws={'pad': 0.02, 'aspect':100}, figsize=(12,12))#, center=0)
    g = sns.clustermap(plot_df, vmin=0, vmax=0.05, cmap='mako_r', **kws)
    x0, _y0, _w, _h = g.cbar_pos
    g.ax_cbar.set_position([x0, 0.9, g.ax_row_dendrogram.get_position().width/2, 0.02])
    #plt.tight_layout()
    #CanvasStyle(g, square=True)
    plt.setp(g.ax_heatmap.yaxis.get_majorticklabels(), rotation=0)
    plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/{prefix}_heatmap_coef_compare.png')

    return plot_df

# calculate the Eucleadian distance from the biomass objective to the inferred objectives
def coef_distance_diffObj(
        df1, df2, labels, prefix, norm=False, func='euclidean', zscore=False,
        model_path='/home/daweilin/StemCell/cancer_model.mat',
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
        histplot=True, boxplot_cols=[], boxplot_order=[],
        ):
    
    # load human metabolic model with acetylation-related reactions
    model = cobra.io.load_matlab_model(model_path)

    # get biomass reaction
    r = model.reactions.get_by_id('biomass_objective')
    biomass_coef_df = pd.DataFrame({
        'Objective':[k.id.split('[')[0] for k in r.metabolites.keys()],
        'coefficient':[-v for v in r.metabolites.values()]
        })
    biomass_coef_df = biomass_coef_df[biomass_coef_df['coefficient']>=0]
    biomass_coef_df.index = biomass_coef_df['Objective']
    biomass_coef_df = biomass_coef_df.drop(columns=['Objective'])
    df_collect = []
    for dftmp in [df1, df2]:
        # outer join
        dfjoin = pd.concat((biomass_coef_df, dftmp), axis=1, join='outer').fillna(0)
        # breakdown coefficients in biomass if selected
        if (dfjoin.index=='gh').sum():
            # melt the coefficient of biomass objectives and add back to each metabolite coef
            distT = dfjoin.copy().T
            distT.loc[distT.index!='coefficient', biomass_coef_df.T.columns] = distT.loc[
                    distT.index!='coefficient', biomass_coef_df.T.columns
                    ].add(
                            np.dot(
                                dfjoin.T['gh'].iloc[1:].to_numpy().reshape(len(dfjoin.T['gh'].iloc[1:]), 1),
                                [biomass_coef_df['coefficient'].to_numpy()],
                                )
                            )
            # overwrite
            dfjoin = distT.T
            dfjoin = dfjoin[dfjoin.index!='gh']
            dfjoin = dfjoin[dfjoin.index!='h2o']

        # remove biomass coefficients
        dfjoin = dfjoin[dfjoin.index!='coefficient']
        # normalize
        dfjoin = dfjoin.div(dfjoin.sum(axis=0), axis=1)
        df_collect.append(dfjoin)

    df1c, df2c = df_collect[0].astype(float), df_collect[1].astype(float)
    vs = []
    cells = []
    for col in df1.columns:
        df = pd.concat((df1c[col], df2c[col]), axis=1, join='outer').fillna(0)
        v = ssd.pdist(
            df.T.values, func
        )[0]
        vs.append(v)
        cells.append(col)
    # distance dataframe
    dist_df = pd.DataFrame({'cellNames':cells, 'Distances':vs})
    dist_df['cellTypes'] = labels
    out_copy = dist_df.copy()
    from sklearn.preprocessing import normalize, Normalizer, MinMaxScaler
    if zscore:
        #min_max_scaler = MinMaxScaler()
        #dist_df['Distances'] = min_max_scaler.fit_transform(dist_df['Distances'].to_numpy().reshape(len(dist_df), 1)).flatten()
        dist_df['Distances'] = dist_df['Distances'].sub(
                dist_df['Distances'].mean()
                ).div(dist_df['Distances'].std())#.div(corrs['Distances'].max())
    print(dist_df)


    if histplot==True:
        # creating a figure composed of two matplotlib.Axes objects (ax_box and ax_hist)
        f, (ax_hist, ax_box) = plt.subplots(2, sharex=True, gridspec_kw={"height_ratios": (.75, .25)}, figsize=(6+len(dist_df['cellTypes'].unique()), 6))
        # assigning a graph to each ax
        sns.histplot(
                data=dist_df, x='Distances', hue='cellTypes', stat='density', kde=True,
                palette='Pastel2', common_norm=False, bins=100, log_scale=False, element='step', ax=ax_hist,
                )
        # plot zero lines
        ax_box.axvline(x=0, linestyle='--', color='grey')
        #ax_hist.axvline(x=0, linestyle='--', color='grey')
        # plot each data point
        g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                data=dist_df, hue='cellTypes', alpha=.5, ax=ax_box, zorder=1)
        gp = sns.pointplot(y="cellTypes", x='Distances',
                      data=dist_df, dodge=False,#.8 - .8 / 3,
                      join=False, hue='cellTypes',palette="Dark2",
                      markers="x", scale=1.8, ci=None, ax=ax_box)
        ax_box.get_legend().remove()

        # Remove x axis name for the boxplot
        ylabel_colors = []
        for ct in dist_df['cellTypes'].unique():
            print(dist_df[dist_df['cellTypes']==ct]['Distances'])
            _, p = ss.mannwhitneyu(dist_df[dist_df['cellTypes']==ct]['Distances'], dist_df[dist_df['cellTypes']!=ct]['Distances'])
            c = 'tomato' if p<0.05 else 'k'
            ylabel_colors.append(c)

        ax_hist.set(xlabel='')
        ax_box.set(ylabel='')
        ax_box.set_xlim([float(dist_df['Distances'].min()), float(dist_df['Distances'].max())])
        ax_box.set_xlabel('Normalized distance')

        ticklabels = ax_box.axes.get_yticklabels()
        print(ticklabels)
        #for i, tick_label in enumerate(ticklabels):
        #    print(tick_label)
        #    tick_label.set_color(ylabel_colors[i])
        CanvasStyle(ax_box, square=True, lw=8, ticks_lw=3)
        CanvasStyle(ax_hist, square=True, lw=8, ticks_lw=3)
        plt.savefig(f'{save_root_path}/{prefix}_histogram_box_coefDist_{func}_{norm}.png')


    if len(boxplot_cols)>0:
        # function for fixing boxplot
        def fixed_boxplot(*args, label=None, **kwargs):
            sns.boxplot(*args, **kwargs, labels=[label])
        def coef_boxplot(dfs):
            # compare pvalues
            fig, ax = plt.subplots(1,1,figsize=(4+len(dfs['cellTypes'].unique()), 8))
            g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                    data=dist_df, hue='cellTypes', alpha=.5, ax=ax, zorder=1)
            g = sns.pointplot(y="cellTypes", x='Distances',
                          data=dist_df, dodge=False,#.8 - .8 / 3,
                          join=False, hue='cellTypes',palette="Dark2",
                          markers="x", scale=1.8, ci=None, ax=ax)

            # plot zero lines
            ax.axvline(x=0, linestyle='--', color='grey')
            handles, labels = ax.get_legend_handles_labels()
            # enlarge dot size
            for dot in handles:
                dot.set_sizes(dot.get_sizes() * 10)
            l = plt.legend(
                    handles[0:2],
                    labels[0:2],
                    #bbox_to_anchor=(1.05, 1),
                    loc='best',#2,
                    borderaxespad=0.,
                    frameon=False
                    )

            ax.set_xlim([float(dist_df['Distances'].min()), float(dist_df['Distances'].max())])
            ax.set_xlabel('Normalized distance')
            CanvasStyle(g, lw=8, ticks_lw=3)
            plt.savefig(f'{save_root_path}/{prefix}_coefDist_compare_{func}_{norm}.png')
        
        coef_boxplot(dist_df)

    return out_copy

# get coefficients of biomass objective
def getRxnCoefficients(
        model_path='/home/daweilin/StemCell/cancer_model.mat',
        model_name='Recon1',
):
    """Breakdown coefficients of biomass objectives from a given metabolic model

    A biomass objective is a linear combination of multiple metabolites as reactants with non-negative coefficients.
    Although there would be products produced by a biomass objective reaction, the function will ignore them.
    The biomass objective could be from different metabolic models such as Recon1, Recon2.2, and Recon3D.
    
    Parameters
    ----------
    model_path : {string}, default='/home/daweilin/StemCell/cancer_model.mat'
        path to access a metabolic model
    model_name : {string}, default='Recon1'
        name of the metabolic model

    Returns
    -------
    biomass_coef_df : {pandas.DataFrame} of shape (n_metabolites, 1)
        table that saves metabolites in rows and corresponding coefficients in the first column
    """
    print('Loading the model...')
    import scipy
    if model_name=='Recon3D':
        # load model
        biomass_raw_eq = scipy.io.loadmat(
                '/nfs/turbo/umms-csriram/daweilin/data/BiGG/recon3d_biomass_coefficients.mat'
                )['rf'][0]
        reactants = biomass_raw_eq.split('  ->')[0].split(' ')
        reactants = np.array(reactants)[np.array(reactants)!='+']
        coef_ind = np.array([int(i%2)==0 for i in np.arange(len(reactants))])
        # make equation into a dataframe
        biomass_coef_df = pd.DataFrame(dict(
                Objective=[ele.split('_c')[0] for ele in reactants[coef_ind==False]],
                coefficient=reactants[coef_ind]
                ))
        biomass_coef_df = biomass_coef_df[biomass_coef_df['Objective']!='h2o']
        biomass_coef_df.index = biomass_coef_df['Objective']
        biomass_coef_df = biomass_coef_df.drop(columns=['Objective'])
        
        # remove duplicates
        res = biomass_coef_df.astype(float).copy()
        res['index'] = res.index
        res = res.groupby('index').sum()
        biomass_coef_df = res
        
    elif model_name=='Recon2.2':
        # load model
        biomass_raw_eq = scipy.io.loadmat(
                '/nfs/turbo/umms-csriram/daweilin/data/BiGG/recon2.2_biomass_coefficients.mat'
                )['rf']
        biomass_coef_df_collect = []
        for arr in biomass_raw_eq:
            print(arr[0][0])
            arr = arr[0][0]
            reactants = arr.split('  ->')[0].split(' ')
            print(reactants)
            reactants = np.array(reactants)[np.array(reactants)!='+']
            coef_ind = np.array([int(i%2)==0 for i in np.arange(len(reactants))])
            # make equation into a dataframe
            biomass_coef_df = pd.DataFrame(dict(
                    Objective=[ele.split('_c')[0] for ele in reactants[coef_ind==False]],
                    coefficient=reactants[coef_ind]
                    ))
            biomass_coef_df = biomass_coef_df[biomass_coef_df['Objective']!='h2o']
            biomass_coef_df.index = biomass_coef_df['Objective']
            biomass_coef_df = biomass_coef_df.drop(columns=['Objective'])
            biomass_coef_df_collect.append(biomass_coef_df)
            
        # merge tables
        biomass_coef_df = pd.concat(biomass_coef_df_collect, axis=0)

        # remove duplicates
        res = biomass_coef_df.astype(float).copy()
        res['index'] = res.index
        res = res.groupby('index').sum()
        biomass_coef_df = res


    else: # recon1
        # load human metabolic model with acetylation-related reactions
        model = cobra.io.load_matlab_model(model_path)
        # get biomass reaction
        biomass_name = 'biomass_objective'
        r = model.reactions.get_by_id(biomass_name)
        biomass_coef_df = pd.DataFrame({
            'Objective':[k.id.split('[')[0] for k in r.metabolites.keys()],
            'coefficient':[-v for v in r.metabolites.values()]
            })
        biomass_coef_df = biomass_coef_df[biomass_coef_df['coefficient']>=0]
        biomass_coef_df = biomass_coef_df[biomass_coef_df['Objective']!='h2o']
        biomass_coef_df.index = biomass_coef_df['Objective']
        biomass_coef_df = biomass_coef_df.drop(columns=['Objective'])
    
    print('Successfully record the coefficients of biomass objectives...')
    return biomass_coef_df




# calculate the Eucleadian distance from the biomass objective to the inferred objectives
def coef_distance_to_biomassObj(
        df, labels, biomass_coef_df, prefix, func='euclidean', norm=False, rank=False, zscore=False, ref_input_type='model',
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
        histplot=True, boxplot_cols=[], boxplot_order=[],
        ):
    """Visualize Eucleadian distances from inferred objectives to biomass objectives
    
    Objective coefficients of up to three different samples will be compared to the coefficients
    in the biomass objectives metabolites with Eucleadian distances and statistically tested with T-test.
    Note that the coefficients of h2o is ignored by this function.

    Parameters
    ----------
    df : {pandas.DataFrame} of shape (n_metabolites, n_samples)
        a table with metabolites as rows and cells as columns
    labels : {list/pandas.DataFrame/numpy.array},
        desired labels for columns
    biomass_coef_df : {pandas.DataFrame} of shape (n_metabolites, 1)
        table that saves coefficients of a biomass objective that could be from Recon1, Recon2.2, or Recon3D
    prefix : {str}
        name of the experiment for saving figures
    func : {string}, default='euclidean'
        name of the method used for the calculation of distances between coefficients
    norm : {bool}, default=False
        normalize the coefficient of each sample by the summation of coefficients of metabolites
    save_root_path : {str}, default='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/'
        path to save the figure
    histplot : {bool}, default=True
        show histogram next to the scatterplot
    boxplot_cols : {array-like}, default=[]
        [WARNING] this input has been deprecated. will be removed in the future
    boxplot_order : {array-like}, default=[]
        names of samples corresponding to labels that order the scatterplot from top to bottom


    Returns
    -------
    None
    """
    # outer join
    dist_df = pd.concat((biomass_coef_df, df), axis=1, join='outer').fillna(0)
    print(dist_df.columns)
    # breakdown coefficients in biomass if selected
    if (dist_df.index=='gh').sum():
        distT = dist_df.copy().T
        distT.loc[distT.index!='coefficient', biomass_coef_df.T.columns] = distT.loc[
                distT.index!='coefficient', biomass_coef_df.T.columns
                ].add(
                        np.dot(
                            dist_df.T['gh'].iloc[1:].to_numpy().astype(float).reshape(len(dist_df.T['gh'].iloc[1:]), 1),
                            [biomass_coef_df['coefficient'].to_numpy().astype(float)],
                            )
                        )
        
        dist_df = distT.T
        dist_df = dist_df[dist_df.index!='gh']
        dist_df = dist_df[dist_df.index!='h2o']
        dist_df = dist_df.astype(float)
    # only focus on biomass component
    #dist_df = dist_df[dist_df.index.isin(biomass_coef_df.index)]
    # normalized to the proportion of the weights
    if norm==True:
        dist_df = dist_df.div(dist_df.sum(axis=0), axis=1)
    if rank==True:
        dist_df = dist_df.rank()
    #dist_df = dist_df.sub(dist_df.mean(axis=1), axis=0).div(dist_df.std(axis=1), axis=0).fillna(0)
    vs = []
    cells = []
    for col in df.columns:
        v = ssd.pdist(
                dist_df[['coefficient', col]].T.astype(float).values, func
            )[0]
        vs.append(v)
        cells.append(col)
    # distance dataframe
    corrs = pd.DataFrame({'cellNames':cells, 'Distances':vs})
    corrs['cellTypes'] = labels
    out_copy = corrs.copy()
    if zscore:
        corrs['Distances'] = corrs['Distances'].sub(
                corrs['Distances'].mean()
                ).div(corrs['Distances'].std())#.div(corrs['Distances'].max())
    print(corrs)

    # significance
    if len(np.unique(labels))==2:
        _, p = ss.f_oneway(
                corrs[corrs['cellTypes']==np.unique(labels)[0]]['Distances'],
                corrs[corrs['cellTypes']==np.unique(labels)[1]]['Distances'],
                )
    else: # >2
        _, p = ss.f_oneway(
                corrs[corrs['cellTypes']==np.unique(labels)[0]]['Distances'],
                corrs[corrs['cellTypes']==np.unique(labels)[1]]['Distances'],
                corrs[corrs['cellTypes']==np.unique(labels)[2]]['Distances'],
                )
    p = np.round(p, 5) #if np.round(p, 5)>0 else ''

    if histplot==True:
        # creating a figure composed of two matplotlib.Axes objects (ax_box and ax_hist)
        f, (ax_hist, ax_box) = plt.subplots(2, sharex=True, gridspec_kw={"height_ratios": (.75, .25)},
                #figsize=(6+len(corrs['cellTypes'].unique()), 6))
            figsize=(8, 4+len(corrs['cellTypes'].unique())))
        # assigning a graph to each ax
        sns.histplot(
                data=corrs, x='Distances', hue='cellTypes', stat='density', kde=True,
                palette='Pastel2', common_norm=False, bins=100, hue_order=boxplot_order,
                log_scale=False, element='step', ax=ax_hist,
                )
        # plot zero lines
        ax_box.axvline(x=0, linestyle='--', color='grey')
        #ax_hist.axvline(x=0, linestyle='--', color='grey')
        # plot each data point
        g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                data=corrs, hue='cellTypes', alpha=.5, ax=ax_box, zorder=1,
                hue_order=boxplot_order, order=boxplot_order)
        gp = sns.pointplot(y="cellTypes", x='Distances',
                      data=corrs, dodge=False,#.8 - .8 / 3,
                      join=False, hue='cellTypes',palette="Dark2",
                      markers="x", scale=1.8, ci=None, ax=ax_box,
                      hue_order=boxplot_order, order=boxplot_order)
        ax_box.get_legend().remove()
        # Remove x axis name for the boxplot
        ax_hist.set_title(f'P-value:{np.round(p, 5)}', fontsize=24)
        ax_hist.set(xlabel='')
        ax_box.set(ylabel='')
        ax_box.set_xlim([float(corrs['Distances'].min()), float(corrs['Distances'].max())])
        #ax_box.set_xlim([1.23, float(corrs['Distances'].max())])
        ax_box.set_xlabel('Distance')

        ticklabels = ax_box.axes.get_yticklabels()
        print(ticklabels)
        CanvasStyle(ax_box, square=True, lw=8, ticks_lw=3)
        CanvasStyle(ax_hist, square=True, lw=8, ticks_lw=3)
        plt.savefig(f'{save_root_path}/{prefix}_histogram_box_dist_{func}_{norm}.png')


    if len(boxplot_cols)>0:
        # function for fixing boxplot
        def fixed_boxplot(*args, label=None, **kwargs):
            sns.boxplot(*args, **kwargs, labels=[label])
        def coef_boxplot(dfs):
            # compare pvalues
            fig, ax = plt.subplots(1,1,figsize=(4+len(dfs['cellTypes'].unique()), 8))
            g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                    data=corrs, hue='cellTypes', alpha=.5, ax=ax, zorder=1)
            g = sns.pointplot(y="cellTypes", x='Distances',
                          data=corrs, dodge=False,#.8 - .8 / 3,
                          join=False, hue='cellTypes',palette="Dark2",
                          markers="x", scale=1.8, ci=None, ax=ax)

            # plot zero lines
            ax.axvline(x=0, linestyle='--', color='grey')
            handles, labels = ax.get_legend_handles_labels()
            # enlarge dot size
            for dot in handles:
                dot.set_sizes(dot.get_sizes() * 10)
            l = plt.legend(
                    handles[0:2],
                    labels[0:2],
                    #bbox_to_anchor=(1.05, 1),
                    loc='best',#2,
                    borderaxespad=0.,
                    frameon=False
                    )
            ax.set_title(f'P-value:{np.round(p, 5)}', fontsize=24)
            ax.set_xlim([1.23, float(corrs['Distances'].max())])
            #ax.set_xlim([float(corrs['Distances'].min()), float(corrs['Distances'].max())])
            ax.set_xlabel('Distance')
            CanvasStyle(g, lw=8, ticks_lw=3)
            plt.savefig(f'{save_root_path}/{prefix}_coef_compare_{func}_{norm}.png')
        
        coef_boxplot(corrs)

    return out_copy

# calculate the Eucleadian distance from the biomass objective to the inferred objectives
def metabolites_dist_to_biomassObj(
        df, labels, prefix, func='euclidean', norm=False, rank=False, zscore=False,
        model_path='/home/daweilin/StemCell/cancer_model.mat',
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
        histplot=True, boxplot_cols=[], boxplot_order=[],
        ):
    # load human metabolic model with acetylation-related reactions
    model = cobra.io.load_matlab_model(model_path)

    # get biomass reaction
    r = model.reactions.get_by_id('biomass_objective')
    biomass_coef_df = pd.DataFrame({
        'Objective':[k.id.split('[')[0] for k in r.metabolites.keys()],
        'coefficient':[-v for v in r.metabolites.values()]
        })
    biomass_coef_df = biomass_coef_df[biomass_coef_df['coefficient']>=0]
    biomass_coef_df.index = biomass_coef_df['Objective']
    biomass_coef_df = biomass_coef_df.drop(columns=['Objective'])
    # outer join
    dist_df = pd.concat((biomass_coef_df, df), axis=1, join='outer').fillna(0)
    # breakdown coefficients in biomass if selected
    if (dist_df.index=='gh').sum():
        distT = dist_df.copy().T
        distT.loc[distT.index!='coefficient', biomass_coef_df.T.columns] = distT.loc[
                distT.index!='coefficient', biomass_coef_df.T.columns
                ].add(
                        np.dot(
                            dist_df.T['gh'].iloc[1:].to_numpy().reshape(len(dist_df.T['gh'].iloc[1:]), 1),
                            [biomass_coef_df['coefficient'].to_numpy()],
                            )
                        )

        dist_df = distT.T
        dist_df = dist_df[dist_df.index!='gh']
    # only focus on biomass component
    dist_df = dist_df[dist_df.index.isin(biomass_coef_df.index)]
    # normalized to the proportion of the weights
    if norm==True:
        dist_df = dist_df.div(dist_df.sum(axis=0), axis=1)
    if rank==True:
        dist_df = dist_df.rank()
    print(dist_df)
    vs = {}
    for col in df.columns:
        col_vs = []
        print(col)
        for i in range(len(dist_df)):
            v = ssd.pdist(
                    [[dist_df[['coefficient', col]].T.iloc[:, i].values[0]],
                        [dist_df[['coefficient', col]].T.iloc[:, i].values[1]]], func
                )[0]
            col_vs.append(v)
        vs[col] = col_vs
    # distance dataframe
    corrs = pd.DataFrame(vs)
    corrs.index = dist_df.index
    out_copy = corrs.copy()
    if zscore:
        corrs['Distances'] = corrs['Distances'].sub(
                corrs['Distances'].mean()
                ).div(corrs['Distances'].std())#.div(corrs['Distances'].max())
    print(corrs)


    if histplot==True:
        # creating a figure composed of two matplotlib.Axes objects (ax_box and ax_hist)
        f, (ax_hist, ax_box) = plt.subplots(2, sharex=True, gridspec_kw={"height_ratios": (.75, .25)}, figsize=(6+len(corrs['cellTypes'].unique()), 6))
        # assigning a graph to each ax
        sns.histplot(
                data=corrs, x='Distances', hue='cellTypes', stat='density', kde=True,
                palette='Pastel2', common_norm=False, bins=100, log_scale=False, element='step', ax=ax_hist,
                )
        # plot zero lines
        ax_box.axvline(x=0, linestyle='--', color='grey')
        #ax_hist.axvline(x=0, linestyle='--', color='grey')
        # plot each data point
        g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                data=corrs, hue='cellTypes', alpha=.5, ax=ax_box, zorder=1)
        gp = sns.pointplot(y="cellTypes", x='Distances',
                      data=corrs, dodge=False,#.8 - .8 / 3,
                      join=False, hue='cellTypes',palette="Dark2",
                      markers="x", scale=1.8, ci=None, ax=ax_box)
        ax_box.get_legend().remove()
        # Remove x axis name for the boxplot
        ylabel_colors = []
        for ct in corrs['cellTypes'].unique():
            print(corrs[corrs['cellTypes']==ct]['Distances'])
            _, p = ss.mannwhitneyu(corrs[corrs['cellTypes']==ct]['Distances'], corrs[corrs['cellTypes']!=ct]['Distances'])
            c = 'tomato' if p<0.05 else 'k'
            ylabel_colors.append(c)

        ax_hist.set(xlabel='')
        ax_box.set(ylabel='')
        ax_box.set_xlim([float(corrs['Distances'].min()), float(corrs['Distances'].max())])
        ax_box.set_xlabel('Normalized distance')

        ticklabels = ax_box.axes.get_yticklabels()
        print(ticklabels)
        CanvasStyle(ax_box, square=True, lw=8, ticks_lw=3)
        CanvasStyle(ax_hist, square=True, lw=8, ticks_lw=3)
        plt.savefig(f'{save_root_path}/{prefix}_histogram_box_dist_{func}_{norm}.png')


    if len(boxplot_cols)>0:
        # function for fixing boxplot
        def fixed_boxplot(*args, label=None, **kwargs):
            sns.boxplot(*args, **kwargs, labels=[label])
        def coef_boxplot(dfs):
            # compare pvalues
            fig, ax = plt.subplots(1,1,figsize=(4+len(dfs['cellTypes'].unique()), 8))
            g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                    data=corrs, hue='cellTypes', alpha=.5, ax=ax, zorder=1)
            g = sns.pointplot(y="cellTypes", x='Distances',
                          data=corrs, dodge=False,#.8 - .8 / 3,
                          join=False, hue='cellTypes',palette="Dark2",
                          markers="x", scale=1.8, ci=None, ax=ax)

            # plot zero lines
            ax.axvline(x=0, linestyle='--', color='grey')
            handles, labels = ax.get_legend_handles_labels()
            # enlarge dot size
            for dot in handles:
                dot.set_sizes(dot.get_sizes() * 10)
            l = plt.legend(
                    handles[0:2],
                    labels[0:2],
                    #bbox_to_anchor=(1.05, 1),
                    loc='best',#2,
                    borderaxespad=0.,
                    frameon=False
                    )

            ax.set_xlim([float(corrs['Distances'].min()), float(corrs['Distances'].max())])
            ax.set_xlabel('Normalized distance')
            CanvasStyle(g, lw=8, ticks_lw=3)
            plt.savefig(f'{save_root_path}/{prefix}_coef_compare_{func}_{norm}.png')
        
        coef_boxplot(corrs)

    return out_copy


# calculate the Eucleadian distance between the fluxes optimizing
# biomass objective and optimizing the inferred objectives
def coef_distance_to_biomassFlux(
        df, labels, prefix,
        medium='KSOM', norm=False, func='euclidean', rank=False, zscore=False,
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
        histplot=True, boxplot_cols=[], boxplot_order=[],
        ):

        # the unconstrained models optimizing biomass
        root_path = f'/nfs/turbo/umms-csriram/daweilin/fluxPrediction/unconstrained_models/pfba/{medium}/'
        uncon_res = unconstrained_models(root_path, norm=True, medium=medium)
        biomass_flux = uncon_res['gh']
        # outer join
        dist_df = pd.concat((biomass_flux, df), axis=1, join='outer').fillna(0)
        # normalization based on the sum of all the fluxes
        if norm==True:
            dist_df = dist_df.div(dist_df.sum(axis=0), axis=1)
        if rank==True:
            dist_df = dist_df.rank()
        print(dist_df)
        # get distance or correlation
        vs = []
        cells = []
        for col in df.columns:
            #print(dist_df[['coefficient', col]].values)
            print(col)
            v = ssd.pdist(
               dist_df[['gh', col]].T.values, func
            )[0]
            vs.append(v)
            cells.append(col)
        
        # distance dataframe
        corrs = pd.DataFrame({'cellNames':cells, 'Distances':vs})
        corrs['cellTypes'] = labels
        out_copy = corrs.copy()
        if zscore:
            corrs['Distances'] = corrs['Distances'].sub(
                    corrs['Distances'].mean()
                    ).div(corrs['Distances'].std())#.div(corrs['Distances'].max())
        print(corrs)


        if histplot==True:
            # creating a figure composed of two matplotlib.Axes objects (ax_box and ax_hist)
            f, (ax_hist, ax_box) = plt.subplots(2, sharex=True, gridspec_kw={"height_ratios": (.75, .25)}, figsize=(6+len(corrs['cellTypes'].unique()), 6))
            # assigning a graph to each ax
            sns.histplot(
                    data=corrs, x='Distances', hue='cellTypes', stat='density', kde=True,
                    palette='Pastel2', common_norm=False, bins=100, log_scale=False, element='step', ax=ax_hist,
                    )
            # plot zero lines
            ax_box.axvline(x=0, linestyle='--', color='grey')
            #ax_hist.axvline(x=0, linestyle='--', color='grey')
            # plot each data point
            g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                    data=corrs, hue='cellTypes', alpha=.5, ax=ax_box, zorder=1, hue_order=boxplot_order)
            gp = sns.pointplot(y="cellTypes", x='Distances',
                          data=corrs, dodge=False,#.8 - .8 / 3,
                          join=False, hue='cellTypes',palette="Dark2",
                          markers="x", scale=1.8, ci=None, ax=ax_box, hue_order=boxplot_order)
            ax_box.get_legend().remove()
            # Remove x axis name for the boxplot
            ylabel_colors = []
            for ct in corrs['cellTypes'].unique():
                print(corrs[corrs['cellTypes']==ct]['Distances'])
                _, p = ss.mannwhitneyu(corrs[corrs['cellTypes']==ct]['Distances'], corrs[corrs['cellTypes']!=ct]['Distances'])
                c = 'tomato' if p<0.05 else 'k'
                ylabel_colors.append(c)

            ax_hist.set(xlabel='')
            ax_box.set(ylabel='')
            ax_box.set_xlim([float(corrs['Distances'].min()), float(corrs['Distances'].max())])
            ax_box.set_xlabel('Normalized distance')

            ticklabels = ax_box.axes.get_yticklabels()
            print(ticklabels)
            CanvasStyle(ax_box, square=True, lw=8, ticks_lw=3)
            CanvasStyle(ax_hist, square=True, lw=8, ticks_lw=3)
            plt.savefig(f'{save_root_path}/{prefix}_flux_histogram_box_{func}_{norm}.png')


        if len(boxplot_cols)>0:
            # function for fixing boxplot
            def fixed_boxplot(*args, label=None, **kwargs):
                sns.boxplot(*args, **kwargs, labels=[label])
            def coef_boxplot(dfs):
                # compare pvalues
                fig, ax = plt.subplots(1,1,figsize=(4+len(dfs['cellTypes'].unique()), 8))
                g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
                        data=corrs, hue='cellTypes', alpha=.5, ax=ax, zorder=1, hue_order=boxplot_order)
                g = sns.pointplot(y="cellTypes", x='Distances',
                              data=corrs, dodge=False,#.8 - .8 / 3,
                              join=False, hue='cellTypes',palette="Dark2",
                              markers="x", scale=1.8, ci=None, ax=ax, hue_order=boxplot_order)

                # plot zero lines
                ax.axvline(x=0, linestyle='--', color='grey')
                handles, labels = ax.get_legend_handles_labels()
                # enlarge dot size
                for dot in handles:
                    dot.set_sizes(dot.get_sizes() * 10)
                l = plt.legend(
                        handles[0:2],
                        labels[0:2],
                        #bbox_to_anchor=(1.05, 1),
                        loc='best',#2,
                        borderaxespad=0.,
                        frameon=False
                        )

                ax.set_xlim([float(corrs['Distances'].min()), float(corrs['Distances'].max())])
                ax.set_xlabel('Normalized distance')
                CanvasStyle(g, lw=8, ticks_lw=3)
                plt.savefig(f'{save_root_path}/{prefix}_flux_compare_{func}_{norm}.png')
            
            coef_boxplot(corrs)

        return out_copy


# significant reactions for each stage
def sankey_plot(
            pca_df, category, labels, col_str, col1, col2,
            show_sig=True, compare_means=True, pv_th=0.05, fc_th=2,
):

    # visualization of comparisons (single comparison)
    def sigRxns(
            pca_df, category, labels, col_str, col1, col2,
            show_sig=True, compare_means=True, pv_th=0.05, fc_th=2,
            ):
    
        # prepare dataframe for boxplots
        dfs = pca_df[pca_df.any(axis=1)].copy()
        dfs.columns = labels
        if category!='':
            dfs['subsystems'] = dfs.index.map(category)
        else:
            dfs['subsystems'] = dfs.index
        dfs['reactions'] = dfs.index
        dfs['tmp'] = dfs[dfs.columns[dfs.columns.str.contains(col_str)]].median(axis=1)
        dfs = dfs.sort_values(by=['subsystems','tmp'])
        dfs = dfs.drop(columns=['tmp'])
        
        # calculate pvalues and foldchanges
        def pvalues_calculate(dfs):
            pvalues = []
            means = []
            # compare mean values
            for i in range(len(dfs)):
                means.append(np.mean(dfs[col2].iloc[i,:])/np.mean(dfs[col1].iloc[i,:]))
            # compare pvalues
            for i in range(len(dfs)):
                _, p = ss.ttest_ind(dfs[col1].iloc[i,:], dfs[col2].iloc[i,:])
                pvalues.append(p)
            
            # inserts two columns
            dfs['pvalues'] = pvalues#fdrcorrection(pvalues)[1]
            dfs['foldChanges'] = means
            # remove lower changes rows
            if compare_means==True:
                dfs = dfs[(dfs['foldChanges'].abs()>fc_th) | (dfs['foldChanges'].abs()<1/fc_th)]
            # remove not significant rows
            if show_sig==True:
                dfs = dfs[dfs['pvalues']<pv_th]
            print(dfs.shape)
            return dfs
        
        return pvalues_calculate(dfs)
    
    # import model
    model = cobra.io.load_matlab_model('/home/daweilin/StemCell/cancer_model.mat')
    # get subsystems
    subsys_map = {}
    for rxn in pca_df.index:
        r = model.reactions.get_by_id(rxn)
        subsys_map[rxn] = r.subsystem
    
    # sample names
    #col1 = 'sc1C2C'
    #col2 = 'sc2CBC'
    # get significant genes
    sig_df = sigRxns(
            pca_df, subsys_map, labels,
            col_str, col1, col2, fc_th=fc_th, pv_th=pv_th,
            show_sig=True, compare_means=True
            )
    
    
    # significance
    # subsystems
    # Tumor/normal
    import matplotlib.pyplot as plt
    
    def get_cmap(n, name='Spectral'):
        '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
        RGB color; the keyword argument name must be a standard mpl colormap name.'''
        return plt.cm.get_cmap(name, n)
    
    
    plot_df = sig_df.iloc[:, -4:]
    plot_df['Sig'] = plot_df.pvalues<0.05
    plot_df['CellType'] = [
            col2 if fc>fc_th else col1 if fc<1/fc_th else 'None' for fc in plot_df.foldChanges
            ]
    node_labels = np.append(
            plot_df['subsystems'].unique(),
            plot_df['CellType'].unique()
            )
    
    cmap = get_cmap(len(node_labels))
    node_colors = [''.join(f'rgba{cmap(i)}'.split(' ')) for i in range(len(node_labels))]
    volumns = []
    targets = []
    sources = []
    for i, subsys in enumerate(plot_df['subsystems'].unique()):
        subsys_chop = plot_df[plot_df['subsystems']==subsys]
        for j, ct in enumerate([col1, col2, 'None']):
            volumns.append(len(subsys_chop[subsys_chop['CellType']==ct]))
            targets.append(j+len(plot_df['subsystems'].unique()))
            sources.append(i)
    
    line_colors = [node_colors[s] for s in sources]
    
    def sankey_diag(data):
     
        import plotly.graph_objects as go
        fig = go.Figure(data=[go.Sankey(
            node = dict(
              pad = 15,
              thickness = 20,
              line = dict(color = "black", width = 0.5),
              label = node_labels,
              color = node_colors
              ),
            link = dict(
              source = sources,
              target = targets,
              value = volumns,
              #label = [f'{s}' for s in sources],
              color = line_colors
          ))])
        
        fig.update_layout(title_text="Significant reactions in embryogenesis", font_size=24)
        import plotly.io as pio
        fig.write_image("/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/sankey_diag.png", width=1000, height=2000)

    sankey_diag(sig_df)
### end the sankey plot function


# parallel plots
def parallel_plots(
        pca_df, mets_category, labels, prefix, col_str,
        stat_test=True, col1='', col2='', pv=0.05,
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/',
        ):

    
    # Import the library
    import plotly.express as px
    
    # prepare dataframe for boxplots
    dfs = pca_df[pca_df.any(axis=1)].copy()
    dfs.columns = labels
    dfs['Mets_type'] = dfs.index.map(mets_category)
    dfs['Objective'] = dfs.index
    dfs['tmp'] = dfs[dfs.columns[dfs.columns.str.contains(col_str)]].median(axis=1)
    dfs = dfs.sort_values(by=['Mets_type','tmp'])
    dfs = dfs.drop(columns=['tmp'])

    if stat_test==True:
        # calculate pvalues
        pvalues = []
        sig = []
        means = []

        for i in range(len(dfs)):
            # compare mean values
            means.append(np.mean(dfs[col2].iloc[i,:])/np.mean(dfs[col1].iloc[i,:]))
            # compare pvalues
            _, p = ss.mannwhitneyu(dfs[col1].iloc[i,:], dfs[col2].iloc[i,:])
            pvalues.append(p)

        # add info into the dataframe
        dfs['pvalues'] = pvalues
        dfs['foldChanges'] = means
        print(dfs)
        dfs = dfs[dfs['pvalues']<pv]
        dfs = dfs.sort_values(by=['pvalues', 'foldChanges'])
        space = '<br>'*37
        dfs['Objective'] = dfs[['Objective', 'pvalues', 'foldChanges']].apply(
                lambda x: ' <b>{0}{1} {3} fc={2} </b>'.format(
                    x.iloc[0],
                    stars(x.iloc[1]),
                    np.round(x.iloc[2], 3),
                    space
                        ), axis=1
                )
    
    plot_df = dfs.iloc[:, :-4].T
    plot_df.columns = dfs['Objective']
    plot_df['Stage'] = plot_df.index
    # get color sequences
    plot_df['Stage'] = plot_df['Stage'].map(
            {k:i+1 for i, k in enumerate(plot_df['Stage'].unique())}
            )
    plot_df.index = np.arange(len(plot_df))

    print(plot_df)
    #import matplotlib.pyplot as plt
    #def get_cmap(n, name='Spectral'):
    #    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    #    return plt.cm.get_cmap(name, n)
    #
    #cmap = get_cmap(len(node_labels))
    #node_colors = [''.join(f'rgba{cmap(i)}'.split(' ')) for i in range(len(node_labels))]
    # Create the chart:
    fig = px.parallel_coordinates(
            plot_df, 
            color=plot_df['Stage'], 
            #labels=plot_df.columns,
            color_continuous_scale=px.colors.diverging.Tealrose,
            #color_continuous_midpoint=2
        )
    
    # Hide the color scale that is useless in this case
    fig.update_layout(coloraxis_showscale=False, font=dict(size=18))

    # save plots
    import plotly.io as pio
    fig.write_image(
            f"{save_root_path}/{prefix}_parallel_plot.png", width=int(
                300+100*len(plot_df.columns)), height=800
            )


# proportion plot comparisons for two stages
def lollipop_fig(
        input_df, mets_category, labels, prefix, cols, cutoff=0.1, value_ordering=True,
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/'
        ):
    """Visualize metabolic objectives with lollipop plots
    
    1) Visualize proportion of metabolites selected by up to three different samples
    2) Generates figures

    Parameters
    ----------
    input_df : {pandas.DataFrame}
        a table with metabolites as rows and cells as columns
    mets_category : {dictionary}
        a dict of metabolites with corresponding types
    labels : {list/pandas.DataFrame/numpy.array},
        desired labels for columns
    cols : {list/numpy.array}
        names of the samples (e.g. control, exp1, exp2)
    prefix : {str}
        name of the experiment for saving figures
    value_ordering : {bool}, default=True
        sort values based on fold changes if set true
    cutoff : {float}, default=0.1
        the portion of cells selecting a metabolite
    save_root_path : {str}, default='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/'
        path to save the figure


    Returns
    -------
    plot_df : {pandas.DataFrame}
        generate a figure and dataframe of calculations

    """
    # color mapping to different samples
    colors = sns.color_palette("Dark2")[:len(cols)]
    
    # prepare dataframe for boxplots
    dfs = input_df[input_df.any(axis=1)].copy() # remove rows with all-zero entries
    dfs.columns = labels # cell types as columns

    # celltype-specific tables
    dfs_arr = [dfs[dfs.columns[labels==cellType]] for cellType in cols]

    # remove metabolites selected only a few portion of cells
    ms = np.concatenate([df.index[(df>0).mean(axis=1)>=cutoff].to_numpy() for df in dfs_arr])
    sel_obj = np.unique(ms)
    dfs = dfs[dfs.index.isin(sel_obj)]

    # binarize coefficients
    dfs_arr = [dfs[dfs.columns[labels==cellType]] for cellType in cols]
    dfs_binary = pd.concat([(df>0).mean(axis=1) for df in dfs_arr], axis=1)
    dfs_binary.columns = cols
    
    # organize tables
    dfs_binary['Mets_type'] = dfs_binary.index.map(mets_category)
    dfs_binary['Objective'] = dfs_binary.index
    dfs_binary['tmp'] = dfs_binary[dfs_binary.columns[dfs_binary.columns.str.contains(cols[0])]].median(axis=1)
    dfs_binary = dfs_binary.sort_values(by=['Mets_type','tmp'])
    dfs_binary = dfs_binary.drop(columns=['tmp'])
    dfs_binary.index = np.arange(len(dfs_binary))

    # make the dataframe for plots
    plot_df = pd.melt(
        dfs_binary,
        id_vars=['Objective'],
        value_vars=[ele for ele in dfs_binary.columns[:-2]]
        )
    plot_df.columns = ['Objective', 'cellType', 'Proportion']
    plot_df['Mets_type'] = plot_df['Objective'].apply(lambda x: x.split(' ')[0]).map(mets_category)

    if len(plot_df.Objective.unique())<3:
        width = len(plot_df.Objective.unique())+4
    else:
        width = len(plot_df.Objective.unique())
    
    # for vertical plot
    tmp_df = plot_df.copy()
    tmp_df.index = np.arange(len(tmp_df))
    tmp_max = tmp_df.groupby('Objective').max()
    tmp_min = tmp_df.groupby('Objective').min()
    # magnitude ordering
    if value_ordering==True:
        reind = pd.DataFrame({
            'max':plot_df.groupby('Objective').max()['Proportion'].to_numpy(),
            'var':plot_df.groupby('Objective').var().to_numpy().flatten(),
            'obj':plot_df.groupby('Objective').var().index}).sort_values(
                    by=['max', 'var'])['obj'].to_numpy()
        #reind = plot_df.groupby('Objective').var().sort_values(by=['Proportion']).index
        tmp_max = tmp_max.T[reind].T
        tmp_min = tmp_min.T[reind].T
        reind = np.repeat(reind, len(cols))
        plot_df.index = plot_df['Objective']
        plot_df = plot_df.T[reind].T
        
    # create an empty plot
    hue_order = cols
    sns.set_context("notebook", font_scale=2.)
    fig, ax = plt.subplots(1,1,figsize=(width//2, 6))
    palette = sns.color_palette(['grey'], len(plot_df['Objective'].unique()))
    #sns.lineplot(
    #        y='Proportion', x='Objective', hue='cellType', linewidth=3,
    #        palette="Pastel2", ax=ax, data=plot_df, legend=False
    #        )
    ax.vlines(
            x=tmp_max.index.to_numpy(),
            ymin=tmp_min['Proportion'].to_numpy().astype(float),
            ymax=tmp_max['Proportion'].to_numpy().astype(float),
            linewidth=3,
            color='grey'
            )
    g = sns.scatterplot(
            y='Proportion', x='Objective', hue_order=hue_order,
            data=plot_df, hue='cellType', palette="Dark2",
            markers="o", s=200, zorder=7, ax=ax
                  )
    
    handles, labels = ax.get_legend_handles_labels()
    # enlarge dot size
    for dot in handles:
        dot.set_sizes(dot.get_sizes() * 10)
    l = plt.legend(
            borderaxespad=0.,
            frameon=False
            )
    plt.xticks(rotation=90)
    ax.set_ylim([-0.05, 1.05])
    ax.set_xlabel('')
    CanvasStyle(ax, lw=12, ticks_lw=4)
    plt.tight_layout()
    plt.savefig(f'{save_root_path}/{prefix}_lollipop_feature_compare.png')
    
    return plot_df

# feature selection plots of multiple samples
def multiGrids_fig(
        pca_df, mets_category, labels, prefix, col_str,
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/'
        ):
    
    # prepare dataframe for boxplots
    dfs = pca_df[pca_df.any(axis=1)].copy() # remove rows with all-zero entries
    dfs.columns = labels # cell types as columns

    # save percentage of feature selection in a dataframe
    plot_dict = {}
    print(len(dfs.columns.unique()))
    for df_col in dfs.columns.unique():
        print(df_col)
        print(sum(dfs.columns==df_col))
        plot_dict[df_col] = (dfs[df_col]>0).sum(axis=1)/sum(dfs.columns==df_col)
    print(plot_dict)
    plot_df = pd.DataFrame(plot_dict)
    # additional information
    plot_df['Mets_type'] = plot_df.index.map(mets_category)
    plot_df['Objective'] = plot_df.index
    plot_df['tmp'] = plot_df[plot_df.columns[plot_df.columns.str.contains(col_str)]].median(axis=1)
    plot_df = plot_df.sort_values(by=['Mets_type','tmp'])
    plot_df = plot_df.drop(columns=['tmp'])
    plot_df.index = np.arange(len(plot_df))
    # make the plot
    sns.set_theme(style="whitegrid")
    
    if len(plot_df)<10:
        height = 5
    else:
        height = 10

    # Make the PairGrid
    g = sns.PairGrid(
            plot_df,
            x_vars=plot_df.columns[:-2],
            y_vars=['Objective'],
            height=height, aspect=.25,
            hue='Mets_type'
                     )
    print(plot_df)
    # Draw a dot plot using the stripplot function
    g.map(sns.stripplot, size=10, orient="h", jitter=False,
          palette="Set2", linewidth=1, edgecolor="w")#, hue=plot_df['Mets_type'])
    
    # Use the same x axis limits on all columns and add better labels
    g.set(xlim=(-0.1, 1.1), xlabel="Proportion", ylabel="")
    
    # Use semantically meaningful titles for the columns
    titles = plot_df.columns[:-2]

    for ax, title in zip(g.axes.flat, titles):
    
        # Set a different title for each axes
        ax.set(title=title)
    
        # Make the grid horizontal instead of vertical
        ax.xaxis.grid(False)
        ax.yaxis.grid(True)
    
    sns.despine(left=True, bottom=True)

    plt.savefig(f'{save_root_path}/{prefix}_feature_compare.png')



# importance plot
def allocation_plot(coef_df, labels, cellType, prefix='', norm=True, cutoff=0.0):
    """The allocation of each metabolite in inferred metabolic objectives.

    The allocation of coefficients will be represented by
    c/Sigma(c) if norm is set true

    Parameters
    ----------
    coef_df : {pandas.DataFrame} of shape (n_metabolites, n_samples),
        the coefficients of metabolic objectives for each samples

    labels : {array-like} of shape (n_samples,),
        the labels of samples/columns

    cellType : {string},
        the label selected to show allocations

    prefix : {string}, default=''
        experiment names for saving files

    norm : {bool}, default=True
        normalize coefficients by the sum of all the coefficients

    cutoff : {float}, default=0.0
        thresholds to show the allocation/coefficients of metabolites
        

    Returns
    -------
    None
    """

    # rank of coefficients
    sns.set_context("notebook", font_scale=2.)
    fig, ax = plt.subplots(1,1,figsize=(15,6))
    bp_df = coef_df[coef_df.columns[labels==cellType]]
    bp_df = bp_df[bp_df.any(axis=1)]
    # normalization
    if norm==True:
        bp_df = bp_df.div(bp_df.sum(axis=0), axis=1)
    mets_sort = bp_df.mean(axis=1).sort_values(ascending=False).index
    bp_df = bp_df.reindex(mets_sort)
    # remove rows with lower values
    bp_df = bp_df[bp_df.mean(axis=1)>cutoff]
    bp_df['Objectives'] = bp_df.index
    bp_df = bp_df.melt(id_vars=['Objectives'])
    sns.barplot(
            data=bp_df,
            x='Objectives',
            y='value',
            color='k',
            ax=ax
            )
    ax.set_xlabel('')
    ax.set_ylabel('Allocation')
    plt.xticks(rotation=90)
    CanvasStyle(ax, lw=8, ticks_lw=3)
    plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/{prefix}_{cellType}_allocations.png')


# importance plot
def overlap_allocation_plot(coef_df, labels, ref_col, prefix='', norm=True, cutoff=0.0):
    """The allocation of each metabolite in objectives that are shown with overlapped barplots.

    The allocation of coefficients will be represented by
    c/Sigma(c) if norm is set true

    Parameters
    ----------
    coef_df : {pandas.DataFrame} of shape (n_metabolites, n_samples),
        the coefficients of metabolic objectives for each samples

    labels : {array-like} of shape (n_samples,),
        the labels of samples/columns

    cellType : {string},
        the label used as a reference

    prefix : {string}, default=''
        experiment names for saving files

    norm : {bool}, default=True
        normalize coefficients by the sum of all the coefficients

    cutoff : {float}, default=0.0
        thresholds to show the allocation/coefficients of metabolites
        

    Returns
    -------
    None
    """
    import matplotlib.patches as mpatches
    # coefficients
    bp_df = coef_df.copy()
    bp_df = bp_df[bp_df.any(axis=1)]
    # normalization
    if norm==True:
        bp_df = bp_df.div(bp_df.sum(axis=0), axis=1)

    mets_sort = bp_df[bp_df.columns[labels==ref_col]].mean(axis=1).sort_values(ascending=False).index
    bp_df = bp_df.reindex(mets_sort)
    # remove rows with lower values
    bp_df = bp_df[bp_df.mean(axis=1)>cutoff]

    # rank of coefficients
    sns.set_context("notebook", font_scale=2.)
    fig, ax = plt.subplots(1,1,figsize=(15,6))

    #plotting columns
    i = 0
    colors = ['teal', 'tomato', 'slateblue', 'lightgrey',]
    patches = []
    for celltype in np.unique(labels):
        plot_df = bp_df[bp_df.columns[labels==celltype]].copy()
        plot_df['Objectives'] = plot_df.index
        plot_df = plot_df.melt(id_vars=['Objectives'])
        ax = sns.barplot(x=plot_df["Objectives"], y=plot_df["value"], color=colors[i], alpha=0.5)
        patches.append(mpatches.Patch(color=colors[i], label=celltype))
        i += 1
     
    #renaming the axes
    #ax.set(xlabel="x-axis", ylabel="y-axis")
    ax.legend(handles=patches)
    ax.set_xlabel('')
    ax.set_ylabel('Allocation')
    plt.xticks(rotation=90)
    CanvasStyle(ax, lw=8, ticks_lw=3)
    plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/{prefix}_overlapped_allocations.png')

# visualization of comparisons (single comparison)
def boxplot_fig(
        input_df, mets_category, labels, col_str, col1, col2,
        prefix, value_ordering=True, pv=0.05, fc=1, col3='', portion=0.0,
        norm=False, plottype='boxplot', xlabel='Normalized coefficients',
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/'
        ):

    """
    
    Description
    -----------
    - Identifies significant metabolites among up to three different samples
    - Calculates based on p-values and fold changes
    - Generates figures

    Arguments
    ---------
    input_df (pandas.DataFrame): a table with metabolites as rows and cells as columns
    mets_category (dictionary): a dict of metabolites with corresponding types
    labels (list/pandas.DataFrame/numpy.array): desired labels for columns
    col_str (str): name of the reference column used to sort values
    col1 (str): name of the first sample (e.g. control)
    col2 (str): name of the second sample (e.g. experiment)
    prefix (str): name of the experiment for saving figures
    
    Optional Arguments
    ------------------
    value_ordering (bool): sort values based on fold changes if set true
    pv (float): p-value thresholds
    fc (float): fold change thresholds
    col3 (str): name of the third sample
    portion (float): the portion of cells selecting a metabolite
    norm (bool): scale values to the range between [0, 1] if set True
    plottype (str): boxplot or stripplot with mean values
    xlabels (str): name of the label on the x-axis
    save_root_path (str): path to save the figure


    Returns
    -------
    coef_boxplot(dfs) (function): generate a figure and dataframe of calculations

    """
    
    # prepare dataframe for boxplots
    dfs = input_df[input_df.any(axis=1)].copy()
    dfs.columns = labels
    
    # remove metabolites selected only a few portion of cells
    if portion>0:
        dfs_arr = [dfs[dfs.columns[labels==cellType]] for cellType in [col1, col2, col3]]
        ms = np.concatenate([df.index[(df>0).mean(axis=1)>=portion].to_numpy() for df in dfs_arr])
        sel_obj = np.unique(ms)
        dfs = dfs[dfs.index.isin(sel_obj)]
    # rescale
    if norm==True:
        dfs = dfs.div(dfs.max(axis=1), axis=0)
    dfs['Mets_type'] = dfs.index.map(mets_category)
    dfs['Objective'] = dfs.index
    dfs['tmp'] = dfs[dfs.columns[dfs.columns.str.contains(col_str)]].median(axis=1)
    dfs = dfs.sort_values(by=['Mets_type','tmp'])
    dfs = dfs.drop(columns=['tmp'])
    
    # color for labels
    #colors = ['tomato', 'teal', 'slateblue'] if col3 else ['tomato', 'teal']
    colors = sns.color_palette("Dark2")[:3] if col3 else sns.color_palette("Dark2")[:2]
    # function for fixing boxplot
    def fixed_boxplot(*args, label=None, **kwargs):
        sns.boxplot(*args, **kwargs, labels=[label])
    def coef_boxplot(dfs, metType=''):
        pvalues, foldChanges, sig, means = [], [], [], []
        # compare mean values
        for i in range(len(dfs)):
            # three samples
            if len(col3)>0:
                tmp_fc, tmp_p, c_sel = 1, 1, 'k'
                for col, color in zip([col1, col2, col3], colors):
                    # foldchange
                    high = np.mean(dfs[col].iloc[i,:])/np.mean(dfs.T[dfs.columns!=col].T.iloc[i,:-2])
                    high = 10 if high==np.inf else high
                    # pvalues
                    _, p = ss.mannwhitneyu(
                            dfs[col].iloc[i,:].astype(float),
                            dfs.T[dfs.columns!=col].T.iloc[i,:-2].astype(float)
                            )
                    if high>tmp_fc and p<tmp_p:
                        tmp_fc, tmp_p, c_sel = high, p, color
                if tmp_fc<fc or tmp_p>p:
                    c_sel = 'k'

                # save the highest value
                foldChanges.append(tmp_fc)
                pvalues.append(tmp_p)
                means.append(c_sel)
            
            # two samples
            else:
                # pvalues
                _, p = ss.mannwhitneyu(dfs[col1].iloc[i,:], dfs[col2].iloc[i,:])
                pvalues.append(p)
                # fold changes
                high = np.mean(dfs[col2].iloc[i,:])/np.mean(dfs[col1].iloc[i,:])
                high = 10 if high==np.inf else high
                foldChanges.append(high)
                if high>fc and p<pv:
                    means.append(colors[1])
                elif high<1/fc and p<pv:
                    means.append(colors[0])
                else:
                    means.append('k')


        # add pvalues, foldchanges, and labels to a column
        dfs['pvalues'] = pvalues
        dfs['foldChanges'] = foldChanges
        dfs['means'] = means
        print(dfs.columns)
        # apply thresholds
        # remove not significant rows and lower foldchanges
        condition1 = (dfs['pvalues']<pv)
        condition2 = ((dfs['foldChanges']>fc) | (dfs['foldChanges']<1/fc))
        dfs = dfs[condition1 & condition2]
        # order the table with the foldChanges/pvalues
        if value_ordering==True:
            dfs = dfs.sort_values(by=['means', 'foldChanges'])
            means = dfs['means'].to_numpy()
        else:
            dfs = dfs.sort_values(by=['Mets_type'])
            means = dfs['means'].to_numpy()
        # checking
        print('obj', len(dfs['Objective']))
        print('pvalue', len(dfs['pvalues']))
        
        # edit the yticklabels with more information
        dfs['Objective'] = dfs[['Objective', 'pvalues']].apply(lambda x: '{0} \n {1}'.format(x.iloc[0], stars(x.iloc[1])), axis=1)
        plot_df = pd.melt(
            dfs,
            id_vars=['Objective'],
            value_vars=[ele for ele in dfs.columns[:-6]]
            )
        plot_df.columns = ['Objective', 'Stage', 'Coefficient']
        plot_df['Coefficient'] = plot_df['Coefficient'].astype(float)
        print('Debug', plot_df[['Objective', 'Stage', 'Coefficient']])
        
        # make fontsize bigger
        sns.set_context("notebook", font_scale=2.)
        # calculating width of the figure
        if len(plot_df.Objective.unique())<4:
            width = len(plot_df.Objective.unique())+3
        else:
            width = len(plot_df.Objective.unique())

        # setup
        if len(col3)>0:
            alpha = 0.2
            palette1 = {k:c for k,c in zip([col1, col2, col3], sns.color_palette("Pastel2")[:3])}
            palette2 = {k:c for k,c in zip([col1, col2, col3], sns.color_palette("Dark2")[:3])}
            dodge = .7 - .8 /4
            width = width*1.05
            HANDLES_ind = 3
            hue_order = [col1, col2, col3]
        else:
            alpha = 0.5
            palette1 = {k:c for k,c in zip([col1, col2], sns.color_palette("Pastel2")[:2])}
            palette2 = {k:c for k,c in zip([col1, col2], sns.color_palette("Dark2")[:2])}
            dodge = .7 - .8 /3
            width = width
            HANDLES_ind = 2
            hue_order = [col1, col2]
            
        # making plots
        if plottype=='boxplot':
            # create an empty plot
            fig, ax = plt.subplots(1,1,figsize=(width+2, 8))
            # boxplot
            g = sns.boxplot(
                    x='Objective', y='Coefficient', hue='Stage', showmeans=False,
                    data=plot_df, palette=palette1,
                    hue_order=hue_order
                    )

            plt.legend(loc='upper center', ncol=2, bbox_to_anchor=(0.5, 1.2))# bbox_to_anchor=(-0.5/len(plot_df.Objective.unique()), 1), col=2)
        else:
            # create an empty plot
            fig, ax = plt.subplots(1,1,figsize=(width+2, 8))
            # stripplot
            g = sns.stripplot(
                    y='Coefficient', x='Objective',
                    s=10, palette=palette1,
                    data=plot_df, hue='Stage',
                    alpha=alpha, dodge=True, zorder=0,
                    hue_order=hue_order
                    )
            g = sns.pointplot(
                    y='Coefficient', x='Objective',
                    data=plot_df, dodge=dodge,
                    join=False, hue='Stage', palette=palette2,
                    markers="X", scale=1, ci=None,
                    hue_order=hue_order
                    )
            ax.set_ylabel(xlabel)
            ax.set_xlabel('')
            handles, labels = ax.get_legend_handles_labels()
            # enlarge dot size
            for dot in handles:
                dot.set_sizes(dot.get_sizes() * 10)
            l = plt.legend(
                    handles[-HANDLES_ind:],
                    labels[-HANDLES_ind:],
                    bbox_to_anchor=(1.00, 1),
                    loc=2,
                    borderaxespad=0.,
                    frameon=False
                    )

        # mark yticklabels with colors by foldchange thresholds
        ticklabels = g.axes.get_yticklabels() if plottype=='boxplot' else g.axes.get_xticklabels()
        for i, tick_label in enumerate(ticklabels):
            print(tick_label)
            tick_label.set_color(means[i])
        # output figures
        plt.xticks(rotation=90)
        plt.tight_layout()
        CanvasStyle(g, lw=12, ticks_lw=4)
        plt.savefig(f'{save_root_path}/{prefix}_coef_compare.png')

        return plot_df, means, g.axes.get_yticklabels()
    
    return coef_boxplot(dfs)


# Correlation between ideally optimal solutions and regression-predicted fluxes
def evaluation_score(coef_sel, labels, flux_df, plot=False, uncon_path='/nfs/turbo/umms-csriram/daweilin/fluxPrediction/unconstrained_models/pfba/KSOM/',medium='KSOM'):
    
    # get weighted reaction reconstruction
    weight_rxn, _ = flux_reconstruction(
        coef_sel,
        root_path=uncon_path,
        medium=medium,
        )

    # Part 1:
    # evaluation
    from sklearn.metrics import r2_score
    reg_pearsonr, reg_spearmanr, reg_r2 = [], [], []
    for label, col in zip(labels, coef_sel.columns):
        a = weight_rxn[col]
        b = flux_df[col]
        print(ss.spearmanr(a,b))
        print(r2_score(a,b))
        reg_spearmanr.append(ss.spearmanr(a, b)[0])
        reg_r2.append(r2_score(a, b))
        reg_pearsonr.append(ss.pearsonr(a, b)[0])

    # convert into a df
    evaluation_df = pd.DataFrame({'R2':reg_r2, 'Pearson':reg_pearsonr, 'Spearman':reg_spearmanr})
    #evaluation_df = evaluation_df.pivot(id_vars=evaluation_df, value_vars=['R2', 'Pearson', 'Spearman'])

    if plot:
 
        # matplots
        fig, ax = plt.subplots(1,1,figsize=(12,8))
        sns.boxplot(data=evaluation_df, ax=ax)
        plt.savefig('/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/regression_evaluation.png')

    return evaluation_df

    
def subSystem_score(
        coef_sel, labels, sample_name='', fc_th=2,
        uncon_path='/nfs/turbo/umms-csriram/daweilin/fluxPrediction/unconstrained_models/pfba/KSOM/',
        model_path='/home/daweilin/StemCell/cancer_model.mat',
        save_root_path='home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
        medium='KSOM',
    ):


    # get weighted reaction reconstruction
    weight_rxn, _ = flux_reconstruction(
        coef_sel,
        root_path=uncon_path,
        medium=medium,
        )

    # Part 3:
    # activity score of subsystems
    gem = cobra.io.load_matlab_model(model_path)
    subsystems = []
    for ID in weight_rxn.index.to_numpy():
        subsys = gem.reactions.get_by_id(ID).subsystem
        if subsys == '':
            subsys = 'Other'
        subsystems.append(subsys)
    
    # add a column for subsystems
    weight_rxn['subsystems'] = subsystems
    wdf = weight_rxn.copy()
    # organize data
    wdf = wdf.replace([np.inf, -np.inf, np.nan], [0, 0, 0])
    wdf = wdf.sort_values(by=['subsystems'])
    systems = wdf['subsystems']
    wdf = wdf.drop(columns=['subsystems'])
    wdf.columns = labels
    
    # count pf subsystems
    uq_sys = np.unique(subsystems, return_counts=True)
    rxn_counts = pd.DataFrame({'counts':uq_sys[1]})
    rxn_counts.index = uq_sys[0]

    #def stat_func(x):
    #    foldChanges = []
    #    pvalues = []
    #    for n in lables.unique():
    #        others = labels.unique()[labels.unique()!=n]
    #        f = np.mean(x[n])/np.mean(x[others])
    #        p = ss.ttest_ind(x[n], x[others])[1]
    #        if 
    if len(labels.unique())==2:
        # calculate pvalues and fold changes
        pv_series = wdf.apply(lambda x: ss.ttest_ind(
            x[labels.unique()[0]],
            x[labels.unique()[1]]
            )[1],
            axis=1).fillna(1)
        fc_series = wdf.apply(
                lambda x: np.mean(x[labels.unique()[0]])/np.mean(x[labels.unique()[1]]), axis=1
                ).fillna(1).replace([np.inf, -np.inf], [fc_th+1, 1/(1+fc_th)])
        # select significant reactions
        sigSample1_rxns = pv_series[(pv_series<0.05) & (fc_series>fc_th)].index
        sigSample2_rxns = pv_series[(pv_series<0.05) & (fc_series<1/fc_th)].index
        sigSample1_df = pd.DataFrame({
            'score':np.zeros(len(sigSample1_rxns)),
            'subsystems':systems[sigSample1_rxns].to_numpy(),
            }).groupby(['subsystems']).count()
        sigSample1_df['labels'] = [labels.unique()[0]]*len(sigSample1_df)
        sigSample2_df = pd.DataFrame({
            'score':np.zeros(len(sigSample2_rxns)),
            'subsystems':systems[sigSample2_rxns].to_numpy(),
            }).groupby(['subsystems']).count()
        sigSample2_df['labels'] = [labels.unique()[1]]*len(sigSample2_df)
        bar_df = pd.concat((sigSample1_df, sigSample2_df), axis=0)
        bar_df['subsystems'] = bar_df.index
        bar_df.index = np.arange(len(bar_df))
        bar_df['score'] = bar_df.apply(lambda x: x['score']/rxn_counts.T[x['subsystems']], axis=1)
        bar_df = bar_df[bar_df['subsystems'].str.contains('Transport|Other|Demand')==0]

    
    if len(labels.unique())>2:
        sigSample_dfs = []
        for label in labels.unique():
            other = labels.unique()[labels.unique()!=label]
            # calculate pvalues and fold changes
            pv_series = wdf.apply(lambda x: ss.f_oneway(
                x[label],
                x[other],
                )[1],
                axis=1).fillna(1)
            # fold change
            fc_series = wdf.apply(
                    lambda x: np.mean(x[label])/np.mean(x[other]), axis=1
                    ).fillna(1).replace([np.inf, -np.inf], [fc_th+1, 1/(1+fc_th)])
             
            # select significant reactions
            sigSample1_rxns = pv_series[(pv_series<0.05) & (fc_series>fc_th)].index
            sigSample1_df = pd.DataFrame({
                'score':np.zeros(len(sigSample1_rxns)),
                'subsystems':systems[sigSample1_rxns].to_numpy(),
                }).groupby(['subsystems']).count()
            sigSample1_df['labels'] = [label]*len(sigSample1_df)
            sigSample_dfs.append(sigSample1_df)

        # merge and plot
        bar_df = pd.concat((sigSample_dfs), axis=0)
        bar_df['subsystems'] = bar_df.index
        bar_df.index = np.arange(len(bar_df))
        bar_df['score'] = bar_df.apply(lambda x: x['score']/rxn_counts.T[x['subsystems']], axis=1)
        bar_df = bar_df[bar_df['subsystems'].str.contains('Transport|Other|Demand')==0]

    # make plot
    sns.set_context("notebook", font_scale=2.)
    fig, ax = plt.subplots(1,1,figsize=(15,20))
    sns.barplot(data=bar_df.sort_values(by=['score', 'labels']), x="score", y="subsystems", hue="labels", ax=ax)
    ax.legend(frameon=False)
    CanvasStyle(ax, lw=12, ticks_lw=4)
    # save the figures
    plt.savefig(f'{save_root_path}/{sample_name}_{labels.unique()[0]}vs{labels.unique()[1]}_metabolicFunc_scores.png')
    
    return bar_df

"""

 ██████ ██      ██    ██ ███████ ████████ ███████ ██████  ██ ███    ██  ██████  
██      ██      ██    ██ ██         ██    ██      ██   ██ ██ ████   ██ ██       
██      ██      ██    ██ ███████    ██    █████   ██████  ██ ██ ██  ██ ██   ███ 
██      ██      ██    ██      ██    ██    ██      ██   ██ ██ ██  ██ ██ ██    ██ 
 ██████ ███████  ██████  ███████    ██    ███████ ██   ██ ██ ██   ████  ██████  
                                                                                
                                                                                
███    ███ ███████ ████████ ██   ██  ██████  ██████  ███████                    
████  ████ ██         ██    ██   ██ ██    ██ ██   ██ ██                         
██ ████ ██ █████      ██    ███████ ██    ██ ██   ██ ███████                    
██  ██  ██ ██         ██    ██   ██ ██    ██ ██   ██      ██                    
██      ██ ███████    ██    ██   ██  ██████  ██████  ███████                    
                                                                                
                                                                                
---
Clustering methods


"""

# clustering methods including PCA, UMAP, tSNE and PHATE
class clustering_func:
    
    """
    Parameter
    ---------
    pca_df (pandas dataframe): genes/reactions/objectives by cells/tissues IDs
    
    Return
    ------
    res_df (pandas dataframe): pca_df with new columns that records entropy and clusters
    """

    from sklearn.preprocessing import StandardScaler


    def __init__(self, pca_df, save_path, prefix, mets_category):
        
        # define class attributes
        self.pca_df = pca_df
        self.save_path = save_path
        self.prefix = prefix
        self.cluster_res = pd.DataFrame()
        self.mets_category = mets_category

    def clustering_evaluation(self, cluster_res, true_col, pred_col, method='MI'):

        if method=='MI': # adjusted mutual information
            from sklearn.metrics.cluster import adjusted_mutual_info_score
            score = adjusted_mutual_info_score(cluster_res[true_col], cluster_res[pred_col])
        elif method=='RI': # adjusted rand index
            from sklearn.metrics.cluster import adjusted_rand_score
            score = adjusted_rand_score(cluster_res[true_col], cluster_res[pred_col])

        else: # rand score
            from sklearn.metrics import silhouette_score
            X = np.array([
                cluster_res['UMAP1'],
                cluster_res['UMAP2']
                ]).reshape(len(cluster_res), 2)
            labels = cluster_res[true_col].to_numpy()
            score = silhouette_score(X, labels, metric="sqeuclidean")

        return score


    def corrmap(self, method='spearman'):

        # get correlations
        corr_map = self.pca_df.corr(method)
        fig, ax = plt.subplots(1,1,figsize=(20,20))
        cp = sns.heatmap(corr_map, cmap='viridis', ax=ax)
        plt.savefig(self.save_path+self.prefix+'corrmap.png')

    def corr_clustermap(self, labels=[], xlabel=False, show_cluster=True, method='average'):

        # clustermap
        pca_df = self.pca_df.copy()
        
        # get sample names without IDs
        if len(labels)==0:
            cols = pca_df.columns#pd.Series(pca_df.columns).apply(lambda x: x.split('_')[0])
        else:
            cols = labels
        # get colors for type of cells/samples
        cellType_pal = sns.color_palette(
                'Set3',
                n_colors=pd.Series(cols).unique().size,
                )
        # map the colors to cell types
        #cellType_pal = sns.color_palette("Set2")
        cellType_lut = dict(zip(map(str, pd.Series(cols).unique()), cellType_pal))
        cellType_colors = pd.Series(cols).map(cellType_lut)
        cellType_colors.index = pca_df.columns
        # Create a custom colormap for the heatmap values
        cmap = sns.diverging_palette(h_neg=210, h_pos=350, s=90, l=30, as_cmap=True)
        # make plots
        #sns.set(font_scale=1.2)
        #, col_linkage, col_lin, col_linkage, col_linkagekage calculate correlaiton
        corr_df = pca_df[pca_df.any(axis=1)].corr().replace([np.nan, np.inf, -np.inf], [0,0,0])
        print('Correlation size')
        print(corr_df.shape)

        if show_cluster:
            # import scipy functions for trimming dndrogram
            from scipy.spatial import distance
            from scipy.cluster import hierarchy
            # calculate linkages
            correlations_array = np.asarray(corr_df)
            row_linkage = hierarchy.linkage(
                distance.pdist(correlations_array), method=method)
            col_linkage = hierarchy.linkage(
                distance.pdist(correlations_array.T), method=method)
            fig, axtmp = plt.subplots(1, 1, figsize=(8, 18))
            dn = hierarchy.dendrogram(col_linkage, ax=axtmp, orientation='right')
            # save plot of dendrogram
            plt.savefig(f'{self.save_path}/{self.prefix}_corr_dendrogram.png')
            
            # label colors for clusters
            leaves_clusters = dn['leaves_color_list']
            network_pal = sns.cubehelix_palette(n_colors=len(np.unique(leaves_clusters)),
                                                light=.9, dark=.1, reverse=True,
                                                start=1, rot=-2)
            network_lut = dict(zip(map(str, np.unique(leaves_clusters)), network_pal))
            dn_mets_map = {met:network_lut[lc] for lc, met in zip(leaves_clusters, corr_df.index[dn['leaves']])}
            network_colors = pd.Series(corr_df.index.to_numpy()).map(dn_mets_map)
            network_colors.index = corr_df.index
            
            # clustergram
            g = sns.clustermap(
                    corr_df,
                    row_linkage=row_linkage,
                    col_linkage=col_linkage,
                    cmap='viridis',
                    col_colors=network_colors,
                    row_colors=network_colors,
                    center=0,
                    yticklabels=False,
                    xticklabels=xlabel,
                    figsize=(20, 20)
                    )
            # label dendrogram clusters with unique numbers
            network_colors = pd.DataFrame(network_colors)
            network_colors.columns = ['c']
            network_map = dict(
                    zip(
                        np.arange(len(np.unique(network_colors['c']))),
                        np.unique(network_colors['c'])
                        )
                    )
            network_colors['cluster'] = network_colors['c'].apply(
                    lambda x: np.nonzero([x==network_map[k] for k in network_map.keys()])[0][0]
                    )
            # put labels onto the figures
            for label, c in network_map.items():
                g.ax_row_dendrogram.bar(0, 0, color=c, label=str(label), linewidth=0)
            
            # edit legend boxes for the additioanl colors
            from matplotlib.pyplot import gcf
            # add legend with the corresponding clusters
            l2 = g.ax_row_dendrogram.legend(title='Cell Type', loc="upper right", bbox_to_anchor=(1., 1.),
                    ncol=len(np.unique(network_colors['c'])), bbox_transform=gcf().transFigure,frameon=False, fontsize=24)

            # save the figures
            plt.savefig(f'{self.save_path}/{self.prefix}_corr_clustermap.png')

            return network_colors

        else:
            # hierarchical clustering without labeling clusters
            g = sns.clustermap(
                    corr_df,
                     cmap='viridis',
                     #z_score=0,
                     row_colors=cellType_colors,
                     col_colors=cellType_colors,
                     row_cluster=True,
                     col_cluster=True,
                     yticklabels=False,
                     xticklabels=xlabel,
                     figsize=(20,20)
                    )
            
            # edit legend boxes for the additioanl colors
            from matplotlib.pyplot import gcf
            
            for label in pd.Series(cols).unique():
                g.ax_row_dendrogram.bar(0, 0, color=cellType_lut[label], label=label, linewidth=0)
            
            l2 = g.ax_row_dendrogram.legend(title='Cell Type', loc="upper right", bbox_to_anchor=(1., 1.),
                    ncol=4, bbox_transform=gcf().transFigure,frameon=False, fontsize=14)
            
            # save the figures
            plt.savefig(self.save_path+self.prefix+'corr_clustermap.png')

    def clustermap(self, labels=[], input_data=None):

        # sort cell types
        pca_dfT = self.pca_df.copy().T
        pca_dfT['cellType'] = labels
        pca_dfT = pca_dfT.sort_values(by=['cellType'])
        labels = pca_dfT['cellType'].to_numpy()
        pca_dfT = pca_dfT.drop(columns=['cellType'])
        pca_df = pca_dfT.T
        # organize data
        pca_df = pca_df.replace([np.inf, -np.inf, np.nan], [0, 0, 0])
        pca_df['Mets_type'] = pca_df.index.map(self.mets_category)
        pca_df['Objective'] = pca_df.index
        pca_df = pca_df.sort_values(by=['Mets_type'])
        #pca_df = pca_df.T[sorted(pca_df.T.columns)].T
        
        # Create a custom palette to identify the networks
        metsType_pal = sns.cubehelix_palette(
                pca_df['Mets_type'].unique().size,
                light=.9, dark=.1, reverse=True,
                start=1, rot=-2
                )

        # map the colors to metabolite category
        metsType_lut = dict(zip(map(str, pca_df['Mets_type'].unique()), metsType_pal))
        metsType_colors = pca_df['Mets_type'].map(metsType_lut)
        
        # Create a custom palette to identify the networks
        mets_labels = pca_df['Mets_type']
        pca_df = pca_df.iloc[:,:-2] # remove the additional columns
        # get sample names without IDs
        if len(labels)==0:
            cols = pca_df.columns#pd.Series(pca_df.columns).apply(lambda x: x.split('_')[0])
        else:
            cols = labels
        #    pca_df.columns = cols

        
        # get colors for type of cells/samples
        cellType_pal = sns.cubehelix_palette(
                pd.Series(cols).unique().size,
                )
        # map the colors to cell types
        #cellType_pal = sns.color_palette("Set2")
        cellType_lut = dict(zip(map(str, pd.Series(cols).unique()), cellType_pal))
        cellType_colors = pd.Series(cols).map(cellType_lut)
        cellType_colors.index = pca_df.columns
        # Create a custom colormap for the heatmap values
        cmap = sns.diverging_palette(h_neg=210, h_pos=350, s=90, l=30, as_cmap=True)
        # make plots
        sns.set_context("notebook", font_scale=1.2)
        g = sns.clustermap(
                    pca_df[pca_df.any(axis=1)].T,#.reset_index(drop=True),
                    cmap='viridis',
                    #z_score=0,
                    #row_colors=cellType_colors,
                    col_colors=metsType_colors,
                    row_cluster=True,
                    col_cluster=False,
                    yticklabels=False,
                    figsize=(len(pca_df[pca_df.any(axis=1)].index),20)
                )
        
        # edit legend boxes for the additioanl colors
        from matplotlib.pyplot import gcf
        
        for label in mets_labels.unique():
            g.ax_col_dendrogram.bar(0, 0, color=metsType_lut[label], label=label, linewidth=0)
        
        l1 = g.ax_col_dendrogram.legend(title='Metabolite Type', loc="lower right", bbox_to_anchor=(0.6, 0.8), bbox_transform=gcf().transFigure, frameon=False, fontsize=24)
        
        #for label in pd.Series(cols).unique():
        #    g.ax_row_dendrogram.bar(0, 0, color=cellType_lut[label], label=label, linewidth=0)
        
        #l2 = g.ax_row_dendrogram.legend(title='Cell Type', loc="lower left", bbox_to_anchor=(0.8, 0.8), bbox_transform=gcf().transFigure,frameon=False, fontsize=24)
        
        # save the figures
        plt.savefig(self.save_path+self.prefix+'clustermap.png')

    
    # functions for dimension reduction visualization
    @staticmethod
    def dimension_reduction_func(pca_df, para, func='PCA', seed=8, transparent=0.8):
        
        if func=='UMAP':
            # UMAP
            import umap
            pc = umap.UMAP(#metric='mahalanobis',
                    n_neighbors=para[0], n_components=para[1], random_state=seed,
                    #n_neighbors=30, n_components=100, random_state=seed,
            ).fit_transform(pca_df.T)

            return pc
        
        elif func=='tSNE':
            # tSNE
            from sklearn.manifold import TSNE
            pc = TSNE(
                    n_components=2,
                    learning_rate='auto',
                    init='random',
                    perplexity=30
                    ).fit_transform(pca_df.T)
            
            return pc
        
        elif func=='PHATE':
            # tSNE
            import phate
            phate_operator = phate.PHATE(k=15, t=100)
            pc = phate_operator.fit_transform(pca_df.T)
            
            return pc
        
        elif func=='PCA_scree':
            # PCA variance
            from sklearn.decomposition import PCA
            X = pca_df.T
            pc = PCA(n_components=8).fit(pca_df.T)
            
            return pc.explained_variance_ratio_
        
        #elif func=='PCA_contribution':

        else:
            # PCA
            from sklearn.decomposition import PCA
            X = pca_df.T
            pc = PCA(n_components=para[0]).fit_transform(pca_df.T)
            
            return pc

    # dimension reduction
    def reduction_scatter3D(
            self, labels, continuous=False, func='UMAP', hue_col='Cell type',
            high_dimension=True,
            title_suffix='', seed=8, text_by_dots=False, plot_order=[],
            para=[30, 50], rot=[30, 60+180], alpha=[0.4],
            save_animate=False, projection=True
            ):

    
        from mpl_toolkits import mplot3d
        PltProps()
        #sns.set_style({'legend.frameon':True, 'figure.facecolor': 'white'})
    
        # set up colormap for visualization
        multiColors = 'gist_rainbow' if len(np.unique(labels))>5 else 'Set2'
        color = multiColors if continuous==False else 'Blues'#sns.color_palette("dark:#69d", as_cmap=True)#"mako"#'viridis'
        pca_df = self.pca_df.copy()
        axis_label = func if func!='PCA' else 'PC'

        # get pcs
        pc = self.dimension_reduction_func(pca_df, para, func=func)

        # save the umap clustering results
        pc_plt = pd.DataFrame(pc[:,:3])
        pc_plt.columns = [f'{axis_label}1', f'{axis_label}2', f'{axis_label}3']
        pc_plt[hue_col] = labels
        pc_plt = pc_plt.sort_values(by=[hue_col])

        # save the clustering results as a class attribute
        self.cluster_res = pc_plt
            
        if high_dimension:
            # Set Axis labels and Title
            fig = plt.figure(figsize=(12,12), facecolor='w')
            ax = plt.axes(projection='3d')
            from matplotlib import cm
            if len(plot_order)>0:
                cellTypes = plot_order
            else:
                cellTypes = pc_plt[hue_col].unique()
            if len(alpha)==1:
                alpha = alpah*len(cellTypes)
            colors = ['lightgrey', 'tomato', 'teal', 'slateblue']
            for i, label in enumerate(cellTypes):
                color = colors[i]
                df = pc_plt[pc_plt[hue_col]==label]
                ax.scatter(df[f'{axis_label}1'], df[f'{axis_label}2'], df[f'{axis_label}3'],
                        marker='s', s=30, c=color, alpha=alpha[i], label=label)

            # 3D figure settings
            ax.legend(facecolor='w')#(bbox_to_anchor=(0.7, 0.8))
            ax.set_xlabel(f'{axis_label}1', linespacing=3.)
            ax.set_ylabel(f'{axis_label}2', linespacing=3.)
            ax.set_zlabel(f'{axis_label}3', linespacing=3.)
            ax.grid(linewidth=3, color='k')
            ax.w_xaxis.line.set_color("k")
            ax.w_yaxis.line.set_color("k")
            ax.w_zaxis.line.set_color("k")
            ax.set_facecolor('w')
            ax.view_init(rot[0], rot[1])
            ax.xaxis._axinfo["grid"].update({'color':'k'})
            ax.yaxis._axinfo["grid"].update({'color':'k'})
            ax.zaxis._axinfo["grid"].update({'color':'k'})

            # legend
            if len(pc_plt[hue_col].unique())>5:
                plt.legend(frameon=False, bbox_to_anchor=(1, 1.1), loc='upper left', fontsize=20)
            else:
                plt.legend(frameon=False, fontsize=20, loc='best')
            plt.tight_layout()
            plt.savefig(self.save_path+self.prefix+hue_col+f'_{title_suffix}_'+'{0}_reduction3D.png'.format(func))

            return pc_plt
        
        if projection==True:
            def pca_projection(pc, cols, labels, continuous):           
                # save the umap clustering results
                pc_plt = pd.DataFrame(pc[:,cols])
                pc_plt.columns = [f'{axis_label}{cols[0]+1}', f'{axis_label}{cols[1]+1}']
                pc_plt[hue_col] = labels
                
               
                # colorbar
                if continuous==True:

                    # create figures
                    fig = plt.figure(figsize = (6,6))
                    ax = fig.add_subplot(1,1,1) 
                    ax = sns.scatterplot(
                            x=f'{axis_label}{cols[0]+1}',
                            y=f'{axis_label}{cols[1]+1}',
                            data=pc_plt,
                            hue=hue_col,
                            alpha=0.5,
                            palette=color,
                            s=100,
                            legend=False,
                            linewidth=1,
                            edgecolor='black',
                            )

                    #plt.legend(bbox_to_anchor=(1.1, 1))
                    for pos in ['right', 'top', 'bottom', 'left']:
                        plt.gca().spines[pos].set_visible(False)
                    plt.xticks([])
                    plt.yticks([])

                    norm = plt.Normalize(0, 1)
                    sm = plt.cm.ScalarMappable(cmap=color, norm=norm)
                    sm.set_array([])
                    ax.figure.colorbar(sm)

                else:

                    colors = ['lightgrey', 'tomato', 'teal', 'slateblue']
                    # create figures
                    fig = plt.figure(figsize = (6,6))
                    ax = fig.add_subplot(1,1,1) 
                    ax = sns.scatterplot(
                            x=f'{axis_label}{cols[0]+1}',
                            y=f'{axis_label}{cols[1]+1}',
                            data=pc_plt,
                            hue=hue_col,
                            alpha=alpha,
                            palette=colors,
                            style=hue_col,
                            s=100,
                            legend=True
                            )
                    
                    for pos in ['right', 'top', 'bottom', 'left']:
                        plt.gca().spines[pos].set_visible(False)
                    plt.xticks([])
                    plt.yticks([])

                ax.set_facecolor('w')
                #ax.set_ylim([pc_plt[f'{axis_label}2'][pc_plt[f'{axis_label}2']>3].min(), pc_plt[f'{axis_label}2'].max()])
                # enlarge the legend dots
                handles, labels = ax.get_legend_handles_labels()
                # enlarge dot size
                for dot in handles:
                    dot.set_sizes(dot.get_sizes() * 10)
                if len(pc_plt[hue_col].unique())>5:
                    plt.legend(frameon=False, bbox_to_anchor=(1, 1.1), loc='upper left', fontsize=20, handles=handles)
                else:
                    plt.legend(frameon=False, fontsize=20, loc='upper left', handles=handles)

                plt.tight_layout()
                plt.savefig(self.save_path+self.prefix+hue_col+f'_{title_suffix}_'+f'{0}_{cols[0]+1}_{cols[1]+1}_reduction.png'.format(func))
        
            # execute
            pca_projection(pc, [0,1], labels, continuous)
            pca_projection(pc, [0,2], labels, continuous)
            pca_projection(pc, [1,2], labels, continuous)

        if save_animate==True:
            from matplotlib import animation
            from mpl_toolkits.mplot3d import Axes3D
            # new figure for animation
            fig = plt.figure(figsize=(10,10), facecolor='w')
            ax = plt.axes(projection='3d')
            def init():
                for i, label in enumerate(cellTypes):
                    color = colors[i]
                    df = pc_plt[pc_plt[hue_col]==label]
                    ax.scatter(df[f'{axis_label}1'], df[f'{axis_label}2'], df[f'{axis_label}3'],
                            marker='s', s=30, c=color, alpha=alpha[i], label=label)

                # 3D figure settings
                ax.legend(facecolor='w')#(bbox_to_anchor=(0.7, 0.8))
                ax.set_xlabel(f'{axis_label}1', linespacing=3.)
                ax.set_ylabel(f'{axis_label}2', linespacing=3.)
                ax.set_zlabel(f'{axis_label}3', linespacing=3.)
                ax.grid(linewidth=3, color='k')
                ax.w_xaxis.line.set_color("k")
                ax.w_yaxis.line.set_color("k")
                ax.w_zaxis.line.set_color("k")
                ax.set_facecolor('w')
                ax.view_init(rot[0], rot[1])
                ax.xaxis._axinfo["grid"].update({'color':'k'})
                ax.yaxis._axinfo["grid"].update({'color':'k'})
                ax.zaxis._axinfo["grid"].update({'color':'k'})

                # legend
                if len(pc_plt[hue_col].unique())>5:
                    plt.legend(frameon=False, bbox_to_anchor=(1, 1.1), loc='upper left', fontsize=20)
                else:
                    plt.legend(frameon=False, fontsize=20, loc='best')
                plt.tight_layout()

                return fig,

            def animate(i):
                ax.view_init(elev=rot[0], azim=i)
                return fig,

            # Animate
            anim = animation.FuncAnimation(fig, animate, init_func=init,
                                       frames=360, interval=200, blit=True)

            ## Save
            print('saving initiation...')
            writergif = animation.PillowWriter(fps=30) 
            anim.save(
                    self.save_path+self.prefix+hue_col+f'_{title_suffix}_'+'PCA_reduction3D_animation.gif',
                    writer=writergif
                    )
            print('saved the animation')



    # dimension reduction
    def reduction_scatter(
            self, labels, continuous=False, func='UMAP', hue_col='Cell type',
            title_suffix='', seed=8, text_by_dots=False, para=[30, 50], alpha=0.5
            ):
        
        PltProps()
        # set up colormap for visualization
        multiColors = 'gist_rainbow' if len(np.unique(labels))>5 else 'Set2'
        color = multiColors if continuous==False else 'Blues'#sns.color_palette("dark:#69d", as_cmap=True)#'viridis'
        pca_df = self.pca_df.copy()
        axis_label = func if func!='PCA' else 'PC'
        
        # get pcs
        pc = self.dimension_reduction_func(pca_df, para, func=func)

        # save the umap clustering results
        pc_plt = pd.DataFrame(pc[:,:2])
        pc_plt.columns = [f'{axis_label}1', f'{axis_label}2']
        pc_plt[hue_col] = labels
        
        # save the clustering results as a class attribute
        self.cluster_res = pc_plt
        
        # create figures
        fig = plt.figure(figsize = (6,6))
        ax = fig.add_subplot(1,1,1) 
        kwargs = {'edgecolor':"black", # for edge color
                     'linewidth':1, # line width of spot
                    }
        ax = sns.scatterplot(
                x=f'{axis_label}1',
                y=f'{axis_label}2',
                data=pc_plt,
                hue=hue_col,
                alpha=alpha,
                palette=color,
                style=hue_col,
                s=100,
                legend=text_by_dots!=1,
                linewidth=1,
                edgecolor='black',
                #**kwargs
                )
        
        for pos in ['right', 'top', 'bottom', 'left']:
            plt.gca().spines[pos].set_visible(False)
        plt.xticks([])
        plt.yticks([])
        
        # colorbar
        if continuous==True:

            # create figures
            fig = plt.figure(figsize = (6,6))
            ax = fig.add_subplot(1,1,1) 
            ax = sns.scatterplot(
                    x=f'{axis_label}1',
                    y=f'{axis_label}2',
                    data=pc_plt,
                    hue=hue_col,
                    alpha=alpha,
                    palette=color,
                    s=100,
                    legend=False,
                    #linewidth=1,
                    #edgecolor='black',
                    )
            
            #plt.legend(bbox_to_anchor=(1.1, 1))
            for pos in ['right', 'top', 'bottom', 'left']:
                plt.gca().spines[pos].set_visible(False)
            plt.xticks([])
            plt.yticks([])

            norm = plt.Normalize(0, 1)
            sm = plt.cm.ScalarMappable(cmap=color, norm=norm)
            sm.set_array([])
            ax.figure.colorbar(sm)
        
        if text_by_dots==True:
            texts = [
                    plt.text(
                        pc_plt[f'{axis_label}1'].iloc[i],
                        pc_plt[f'{axis_label}2'].iloc[i],
                        pc_plt[hue_col].iloc[i],
                        ) for i in range(len(pc_plt))
                    ]
            adjust_text(
                    texts,
                    arrowprops=dict(
                        arrowstyle="-", 
                        lw=2
                        ),
                    )

        ax.set_facecolor('w')
        #ax.set_ylim([pc_plt[f'{axis_label}2'][pc_plt[f'{axis_label}2']>3].min(), pc_plt[f'{axis_label}2'].max()])
        # enlarge dots in legend
        handles, labels = ax.get_legend_handles_labels()
        # enlarge dot size
        for dot in handles:
            dot.set_sizes(dot.get_sizes() * 10)
        if len(pc_plt[hue_col].unique())>5:
            plt.legend(frameon=False, bbox_to_anchor=(1, 1.1), loc='best', fontsize=20, handles=handles)
        else:
            plt.legend(frameon=False, fontsize=20, loc='best', handles=handles)
        plt.tight_layout()
        plt.savefig(self.save_path+self.prefix+hue_col+f'_{title_suffix}_'+'{0}_reduction.png'.format(func))
        
        # scree plot
        if func=='PCA':
            # scree plot
            var_ratio = self.dimension_reduction_func(pca_df, para, func='PCA_scree')
            
            # create figures
            sns.set_context("notebook", font_scale=2.)
            fig = plt.figure(figsize = (6,6))
            ax = fig.add_subplot(1,1,1) 
            ax = sns.lineplot(
                    y=var_ratio,
                    x=np.arange(len(var_ratio)),
                    color='grey',
                    marker='o',
                    markersize=10,
                    linewidth=5,
                    )

            ax.set_facecolor('w')
            ax.set_xlabel('Principle components')
            ax.set_ylabel('Explained variance')
            CanvasStyle(ax, lw=12, ticks_lw=4)
            plt.tight_layout()
            plt.savefig(self.save_path+self.prefix+hue_col+f'_{title_suffix}_'+'{0}_Screeplot.png'.format(func))
            
            

        return pc_plt

    
    def reclustering(self, cluster_res, cols, min_size=10, func='UMAP', title_suffix=''):
        
        """
        
        Parameters
        ----------
        cluster_res (pandas.DataFrame): clustering results. It could be umap, pca, and so on.
        cols (list): the column names used for reclustering.
        
        Returns
        -------
        recluster_res (pandas.DataFrame): data reclustering coming with the true/predicted labels.

        """


        # import hdbscan package
        import hdbscan
        clusterer = hdbscan.HDBSCAN(min_cluster_size=min_size, gen_min_span_tree=True)
        clusterer.fit(cluster_res[cols])
        
        # predicted clusters
        cluster_res['cluster'] = clusterer.labels_
        self.cluster_res = cluster_res

        # set up colormap for visualization
        color = 'Set2'
       
        # create figures
        fig = plt.figure(figsize = (6,6))
        ax = fig.add_subplot(1,1,1) 
        ax = sns.scatterplot(
                x=f'{func}1',
                y=f'{func}2',
                data=cluster_res,
                hue='cluster',
                alpha=0.5,
                palette=['k']+sns.color_palette('Set2')[:len(cluster_res['cluster'])-1],
                style='cluster',
                s=100,
                legend=False
                )
        
        for pos in ['right', 'top', 'bottom', 'left']:
            plt.gca().spines[pos].set_visible(False)
        plt.xticks([])
        plt.yticks([])
        
        ax.set_facecolor('w')
        plt.legend(frameon=False)
        plt.tight_layout()
        plt.savefig(self.save_path+self.prefix+f'_{title_suffix}_{func}_recluster.png')
        
        return cluster_res





"""
Allocation of reaction fluxes and objectives
"""
        
        

"""
Figure 5. 
Part A. allocation of reaction fluxes and objectives
"""
def logVariance_allocation(df, labels, sampleName, replace_dict,
        dataType='Objective', model_path='/home/daweilin/StemCell/cancer_model.mat'):

    """
    Parameter
    ---------
    flux_df (pandas.DataFrame): table of reaction fluxes
    obj_df (pandas.DataFrame): table of objective coefficients
    sampleName (str): name of the sample which will be used for the prefix of filenames (e.g. scEmbryo)
    replace_dict (dictionary): e.g. {'1Cto2cell':'sc1C2C', '2Cto32cell':'sc2CBC'}

    Return
    ------
    None
    """
    #def gamma_func_Enz(x):
    #    
    #    fit_alpha, fit_loc, fit_beta=ss.gamma.fit(x)
    #    for v in x:
    #        v1 = ss.gamma.cdf(v,
    #            fit_alpha, loc=fit_loc, scale=fit_beta,
    #            )
    #        v2 = ss.gamma.cdf(np.max(x[x<v]),
    #            fit_alpha, loc=fit_loc, scale=fit_beta,
    #            )
    #    return x
    ##

    #partition = df.apply(lambda x: gamma_func_Enz(x), axis=1).fillna(0)
    partition = df.apply(lambda x: (x/52)*np.log(x/52)).fillna(0)
    partition = -partition.sum(axis=0)
    scEnz = pd.DataFrame(partition)
    scEnz = scEnz.replace([-np.inf, np.inf], [0, 0])
    scEnz.columns = [f'Entropy of {dataType}']
    scEnz['cellType'] = labels.to_numpy()
    scEnz['cellType'] = scEnz['cellType'].replace(replace_dict)
    out_df = scEnz.copy().replace([0], [1])#[scEnz[f'Entropy of {dataType}'].mean()])
    scEnz = scEnz.replace([0], [1])#[scEnz[f'Entropy of {dataType}']!=0]
    print(scEnz)
    # 1D log var
    stripdf = scEnz.copy()
    for col in stripdf.columns[:-1]:
        eq = lambda x:f'ss.f_oneway({x})'
        input_str = ''
        for i in range(len(scEnz['cellType'].unique())):
            input_str += f'stripdf[col][stripdf.cellType==scEnz["cellType"].unique()[{i}]],'
        _, p = eval(eq(input_str))
        # make plot
        #stripdf['cellType'] = stripdf['cellType'].apply(lambda x: '{0} \n {1}'.format(x, stars(pv[x])), axis=1)
        sns.set_context("notebook", font_scale=2.)
        fig, ax = plt.subplots(1,1,figsize=(10, 4))
        g = sns.stripplot(y='cellType', x=col, s=10, palette="Pastel2",
                data=stripdf, hue='cellType', alpha=.5, zorder=0, legend=False)
        g = sns.pointplot(y="cellType", x=col,
                      data=stripdf, dodge=False,#.8 - .8 / 3,
                      join=False, hue='cellType',palette="Dark2",
                      markers="x", scale=3, ci=None,) #zorder=1)
        ax.set_title(f'P-value={p:.2g}',fontsize=24)
        ax.set_xlabel('Entropy')
        CanvasStyle(ax, lw=8, ticks_lw=3)
        leg = plt.legend()
        ax.get_legend().remove()
        plt.tight_layout()
        # save the figures
        plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/{sampleName}_ShannonEntropy_{col}.png')

    return out_df

"""
Figure 5.
Part B. graph entropy
"""

def graph_entropy():
    # function to calculate graph entropy
    def graph_calc(graph_path):
     
        import math
        
        def degree_distribution(G):
            vk = dict(G.degree())
            vk = list(vk.values()) # we get only the degree values
            maxk = np.max(vk)
            mink = np.min(min)
            kvalues= np.arange(0,maxk+1) # possible values of k
            Pk = np.zeros(maxk+1) # P(k)
            for k in vk:
                Pk[k] = Pk[k] + 1
            Pk = Pk/sum(Pk) # the sum of the elements of P(k) must to be equal to one
            
            return kvalues,Pk
        
        def shannon_entropy(G):
            k,Pk = degree_distribution(G)
            H = 0
            for p in Pk:
                if(p > 0):
                    H = H - p*math.log(p, 2)
            return H
    
        graph_files = os.listdir(graph_path)
        gs = {}
        graphE = {}
        for gf in graph_files:
            with open(graph_path+gf, 'r') as f:
                j = json.load(f)
                G = nx.json_graph.adjacency_graph(j)
                gs[gf.split('.json')[0]] = G
                graphE[gf.split('.json')[0]] = shannon_entropy(G)
    
        return graphE, gs
    
    
    # get path of graphs
    graph_sc1C2C = '/nfs/turbo/umms-csriram/daweilin/graph/sc1C2C/'
    graph_sc2CBC = '/nfs/turbo/umms-csriram/daweilin/graph/sc2CBC/'
    # get graph entropy and graphs
    scEnz1C2C, scG1C2C = graph_calc(graph_sc1C2C)
    scEnz2CBC, scG2CBC = graph_calc(graph_sc2CBC)
    
    # get path of graphs
    graph_sc1C2Cflux = '/nfs/turbo/umms-csriram/daweilin/graph/multiObj/con_sc1C2C/'
    graph_sc2CBCflux = '/nfs/turbo/umms-csriram/daweilin/graph/multiObj/con_sc2CBC/'
    # get graph entropy and graphs
    scEnz1C2Cflux, scG1C2Cflux = graph_calc(graph_sc1C2Cflux)
    scEnz2CBCflux, scG2CBCflux = graph_calc(graph_sc2CBCflux)
    
    
    # save graph entropy in dictionary
    EnzDict = {name:pd.DataFrame({'Graphs':[k for k in D.keys()], 'Entropy':[v for v in D.values()]}) for D, name in zip([scEnz1C2C, scEnz1C2Cflux, scEnz2CBC, scEnz2CBCflux], ['1C2Cobj', '1C2Cflux', '2CBCobj', '2CBCflux'])}
    
    
    # rename index and merge datasets
    for k in EnzDict.keys():
        EnzDict[k].Graphs = EnzDict[k].Graphs.apply(lambda x: ''.join(x.split('_')[1:4]))
        EnzDict[k].index = EnzDict[k].Graphs
        EnzDict[k] = EnzDict[k].iloc[:, 1:]
    
    # dataframes for 2D entropy plots
    scEnz1C2Cdfs = pd.concat((EnzDict['1C2Cobj'], EnzDict['1C2Cflux']), axis=1)
    scEnz2CBCdfs = pd.concat((EnzDict['2CBCobj'], EnzDict['2CBCflux']), axis=1)
    scEnz1C2Cdfs.columns = ['Objective Graph Entropy', 'Flux Graph Entropy']
    scEnz2CBCdfs.columns = ['Objective Graph Entropy', 'Flux Graph Entropy']
    scEnz = pd.concat((scEnz1C2Cdfs, scEnz2CBCdfs), axis=0)
    scEnz.columns = ['Objective Graph Entropy', 'Flux Graph Entropy']
    scEnz['cellType'] = pd.Series(scEnz.index).apply(
                lambda x: ''.join(x.split('_')[0])
                ).to_numpy()
    
    # 2D plot for graph entropy
    fig, ax = plt.subplots(1,1,figsize=(15, 15))
    sns.scatterplot(
            x='Flux Graph Entropy', y='Objective Graph Entropy',
            hue='cellType', data=scEnz, legend=True, s=100,
            )
    ax.set_title('Scatterplot of graph entropy')
    CanvasStyle(ax)
    plt.tight_layout()
    # save the figures
    plt.savefig('/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results/scEmbryo_GraphEntropy2D.png')
    
    # 1D graph entropy
    stripdf = scEnz.copy()
    for col in stripdf.columns[:-1]:
        _, p = ss.ttest_ind(
                stripdf[col][stripdf.cellType=='sc1C2C'],
                stripdf[col][stripdf.cellType=='sc2CBC']
                )
        # make plot
        fig, ax = plt.subplots(1,1,figsize=(10, 6))
        g = sns.stripplot(y='cellType', x=col, s=10, palette="Pastel2",
                data=stripdf, hue='cellType', alpha=.5)
        g = sns.pointplot(y="cellType", x=col,
                      data=stripdf, dodge=False,#.8 - .8 / 3,
                      join=False, hue='cellType',palette="Dark2",
                      markers="x", scale=1.8, ci=None)
        ax.set_title(f'{col} (pvalue={np.round(p, 5)})',fontsize=24)
        ax.set_xlabel('Entropy')
        handles, labels = ax.get_legend_handles_labels()
        l = plt.legend(
                handles[0:2],
                labels[0:2],
                bbox_to_anchor=(1.05, 1),
                loc=2,
                borderaxespad=0.,
                frameon=False
                )
        CanvasStyle(ax)
        plt.tight_layout()
        # save the figures
        plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/scEmbryo_GraphEntropy_{col}.png')
    





"""

██    ██  █████  ██      ██ ██████   █████  ████████ ██  ██████  ███    ██ 
██    ██ ██   ██ ██      ██ ██   ██ ██   ██    ██    ██ ██    ██ ████   ██ 
██    ██ ███████ ██      ██ ██   ██ ███████    ██    ██ ██    ██ ██ ██  ██ 
 ██  ██  ██   ██ ██      ██ ██   ██ ██   ██    ██    ██ ██    ██ ██  ██ ██ 
  ████   ██   ██ ███████ ██ ██████  ██   ██    ██    ██  ██████  ██   ████ 
                                                                           
                                                                           
███    ███ ███████ ████████ ██   ██  ██████  ██████  ███████               
████  ████ ██         ██    ██   ██ ██    ██ ██   ██ ██                    
██ ████ ██ █████      ██    ███████ ██    ██ ██   ██ ███████               
██  ██  ██ ██         ██    ██   ██ ██    ██ ██   ██      ██               
██      ██ ███████    ██    ██   ██  ██████  ██████  ███████               
                                                                           
                                                                           
---                                                                           
Methods for model validation


"""


def biomass_gene_essentiality(ko_df, sel_para, method='zscore', th=1):
    
    """
    
    Description
    -----------
    1. Calculate the changes of biomass objective values with different KO
    2. Identify genes resulting in significantly changes with diff criteria
    
    Arguments
    ---------
    ko_df (pandas.DataFrame): data table-->genes-->columns and rxns-->rows
    sel_para
    method (str): the criteria used to identify genes
    - 'zscore': default, standarize data and find genes with provided thresholds
    - 'percentage': gene is essential if KO biomass<N% WT biomass (Yang et al., 2014)
    th (float): thresholds for identifying genes, num of std/percentage for zscore/percentage

    Returns
    -------
    sigGenes (list): a list of significant genes


    """

    # Biomass as objectives
    # correlations between WT and knockouts
    paras = pd.Series(ko_df.columns).apply(
            lambda x: '_'.join(x.split('_')[-3:-1])
            )
    genes = pd.Series(ko_df.columns).apply(
            lambda x: '_'.join(x.split('_')[:-3])+'_'+x.split('_')[-1]
            )
    geneNames = pd.Series(ko_df.columns).apply(
            lambda x: x.split('_')[-1]
            )
    ko_df = ko_df[ko_df.index=='biomass_objective']

    if method=='percentage':
        fc_dict = {}
        plot_df = ko_df[ko_df.columns[paras==sel_para]]
        gene_arr = geneNames[paras==sel_para]
        fc = plot_df.iloc[:, 1:].apply(
                lambda x: x.div(plot_df.iloc[:,0]), # get ratio of KO/WT biomass
                )
        fc_dict[sel_para] = fc.iloc[0,:].to_numpy()

        # convert correlation dictionary into a dataframe
        biomass_fc_df = pd.DataFrame(fc_dict).fillna(0)
        biomass_fc_df.index = gene_arr[1:]#pd.Series(genes[paras==para]).apply(lambda x: x.split('_')[-1]).to_numpy()[1:]
        # 1% less than the wildtype objective value
        biomass_growthGenes = biomass_fc_df.index[biomass_fc_df[sel_para]<th]
        biomass_nongrowthGenes = biomass_fc_df.index[biomass_fc_df[sel_para]>1/th]
   
    else: #'zscore'
        fc_dict = {}
        plot_df = ko_df[ko_df.columns[paras==sel_para]]
        gene_arr = geneNames[paras==sel_para]
        fc = plot_df.iloc[:, 1:].apply(
                lambda x: x.sub(plot_df.iloc[:,0]),
                )
        fc_dict[sel_para] = fc.iloc[0,:].to_numpy()
        
        # convert correlation dictionary into a dataframe
        biomass_fc_df = pd.DataFrame(fc_dict).fillna(0)
        biomass_fc_df.index = gene_arr[1:]#pd.Series(genes[paras==para]).apply(lambda x: x.split('_')[-1]).to_numpy()[1:]
        if th==0:
            # convert biomass 
            biomass_growthGenes = biomass_fc_df.index[
                    biomass_fc_df[sel_para]<0]
            biomass_nongrowthGenes = biomass_fc_df.index[
                    biomass_fc_df[sel_para]>0]
        else:
            # zscore
            zbiomass = biomass_fc_df.sub(biomass_fc_df.mean(axis=0), axis=1).div(biomass_fc_df.mean(axis=0), axis=1)
            # gating   
            biomass_growthGenes = zbiomass.index[zbiomass[sel_para]<-th]
            biomass_nongrowthGenes = zbiomass.index[zbiomass[sel_para]>th]

    return biomass_growthGenes, biomass_nongrowthGenes


def metObj_gene_essentiality(coef_paths, sel_para, method='zscore', th=1):
    
    """
    # Process objective coefficients
    coef_paths = {
            'ESC':f'/nfs/turbo/umms-csriram/daweilin/regression_models/GeneKO/flux_sl_ESC_{sel_para}_norm_nonorm.csv',
            }
    """

    # get fluxes of multi-objective models with constraints
    coefs = {}
    for k in coef_paths.keys():
        path = coef_paths[k]
        coefs[k] = pd.read_csv(path, index_col=0)
        #coefs[k].index = pd.Series(coefs[k].index).replace(mnames)
    
    # Merge
    coefs_df = pd.concat((coefs[k] for k in coefs.keys()), axis=1)
    paras = pd.Series(coefs_df.columns).apply(
            lambda x: '_'.join(x.split('_')[-3:-1])
            )
    coefs_df = coefs_df[coefs_df.columns[paras==sel_para]]
    labels = pd.Series(coefs_df.columns).apply(lambda x: x.split('_')[-1])
    print(coefs_df.shape, paras)
    
    # distance from the biomass coef to the inferred obj
    #dist_df = coef_distance_to_biomassObj(
    #        coefs_df, labels, 'KO',
    #        norm=True, rank=False, func='euclidean',
    #        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
    #        histplot=False, boxplot_cols=[], boxplot_order=[]
    #        )
    dist_res = []
    coefs_df.columns = labels
    for col in coefs_df.columns:
        dist_res.append(ssd.pdist(coefs_df[['WT', col]].T.values, 'euclidean')[0])
    # convert correlation dictionary into a dataframe
    dist_df = pd.DataFrame({'Distances':dist_res}).fillna(0)
    dist_df.index = labels#pd.Series(genes[paras==para]).apply(lambda x: x.split('_')[-1]).to_numpy()[1:]
    
    # get difference of the distance
    diff_df = dist_df['Distances']#-dist_df['Distances'].iloc[0]
    
    # quantile methods
    if method=='quantile':
        # get q1 and q3 and calculate iqr
        Q3 = np.percentile(diff_df.to_numpy(), 75)
        Q1 = np.percentile(diff_df.to_numpy(), 25)
        iqr = Q3-Q1
        ub = Q3+1.*iqr
        lb = Q1-1.*iqr
        # identify genes
        nongrowth_genes = labels.to_numpy()[diff_df<Q1]
        growth_genes = labels.to_numpy()[diff_df>Q3]

    elif method=='zscore':
        # standarize data
        zdiff = diff_df.sub(
                diff_df.mean()
                ).div(diff_df.std())
        growth_genes = labels.to_numpy()[zdiff>th]
        nongrowth_genes = labels.to_numpy()[zdiff<-th]
        
    else: # nonzero changes
        # get genes that change the distances
        nongrowth_genes = labels.to_numpy()[diff_df<0]
        growth_genes = labels.to_numpy()[diff_df>0]
        
    return growth_genes, nongrowth_genes



#+---------------------+
#+ Knockout validation +
#+---------------------+
def geneEssentiality_evaluation(tb, sigGenes_dict, method_labels, sel_para, plot=False):

    # import packages
    from stat_tests import hypergeom_test, ranksumtest
    import scipy
    geneList = scipy.io.loadmat('/home/daweilin/StemCell/unique_gene_list.mat')
    # unique genes with unknown labels and wilde type columns
    genes_columns = [
            g[0][0] if len(g[0])>0 else 'unknown' for g in geneList['unique_gene_list']
            ]
    # make the gene names uppercase
    tb['Gene name'] = tb['Gene name'].apply(lambda x: x.upper())
    print('NANOG', tb[tb['Gene name']=='NANOG'].mean())
    print('SOX2', tb[tb['Gene name']=='SOX2'].mean())
    print('POU5F1', tb[tb['Gene name']=='POU5F1'].mean())
    print('ALL', tb.mean())
    
    # get genes in the metabolic models
    tb = tb[tb['Gene name'].isin(genes_columns)]
    # save significant genes identified by different methods
    gene_collect = {
            method:list(sigGenes_dict[method]) for method in method_labels
            }
    # save all the metabolic genes cross-listed
    gene_collect['model_genes'] = list(tb['Gene name'].to_numpy())
    # evaluation res
    hypergeom_res = {'Sample':[], 'P-values':[], 'Methods':[]}
    for col in tb.columns[1:]:
        for method in method_labels:
            sigG = sigGenes_dict[method]
            M, N = len(tb), sum(tb[col]<0.05)#, len(sigG)
            n = sum(tb['Gene name'].isin(sigG))
            k = sum(tb['Gene name'][tb[col]<0.05].isin(sigG))
            print(M, N, n, k)
            p, _ = hypergeom_test(M, N, n, k)
            hypergeom_res['Sample'].append(col)
            hypergeom_res['P-values'].append(p)
            hypergeom_res['Methods'].append(method)
            # save significant genes
            gene_collect[col] = list(
                    tb[tb[col]<0.05]['Gene name'].to_numpy()
                    )


    # edit the column name
    hypergeom_df = pd.DataFrame(hypergeom_res)
    hypergeom_df['Sample'] = hypergeom_df['Sample'].apply(lambda x: x.replace('Current screen', 'Shohat et al'))
    hypergeom_df['Sample'] = hypergeom_df['Sample'].apply(lambda x: x.replace('p-value', ''))
    hypergeom_df['Sample'] = hypergeom_df['Sample'].apply(lambda x: x.replace('p- value', ''))
    
    samples = ['Shohat et al, negative selection ','Tzelepis et al, negative selection  ','Negative selection combined ','Negative selection combined FDR corrected ','Shohat et al, positive selection ','Tzelepis et al, positive selection  ','Positive selection combined ','Positive selection combined FDR corrected ']
    eval_df = hypergeom_df.pivot("Sample", "Methods", "P-values")
    eval_df = eval_df.reindex(samples)
    eval_df.index = pd.Series(eval_df.index).apply(lambda x: x.replace(', ', '\n'))
    eval_df.index = pd.Series(eval_df.index).apply(lambda x: x.replace(' corrected', ''))
    eval_df = eval_df[method_labels]

    # plot
    if plot==True:
        sns.set_context("notebook", font_scale=2.)
        cmap = sns.diverging_palette(220, 20, as_cmap=True)
        fig, ax = plt.subplots(1,1,figsize=(30,20))
        sns.heatmap(
                -np.log10(eval_df.T), square=True, linewidth=3,
                annot=True, fmt='.1f',
                cmap=cmap, center=-np.log10(0.05), ax=ax
                )
        ax.set_xlabel('')
        ax.set_ylabel('')
        #CanvasStyle(ax, lw=12, ticks_lw=4, square=True)
        plt.tight_layout()
        plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/hypergeomTest_ESCKO_hp_{sel_para}.png')

    return eval_df, gene_collect

def subSystem_gene_enrichment(geneList, path='/home/daweilin/StemCell/cancer_model.mat'):
    # load model
    _, gpr_df = read_mat_model(
            path, return_rxnGeneMat=True, remove_transport=True
            )
    # iterate thru all subsystems to calculate the enrichment
    gpr_df = gpr_df.drop(columns=['rxns'])
    subsys_sum = gpr_df.groupby('subSystems').sum()
    all_genes = subsys_sum.columns
    print(subsys_sum)
    #N_series = (subsys_sum>0).sum(axis=1)
    def hypergeom_test_for_df(x):
        M = len(all_genes)
        N = (x>0).sum()
        n = len(geneList)
        k = pd.Series(geneList).isin(all_genes[x>0]).sum()
        p, _ = hypergeom_test(M, N, n, k)
        return p
    # report pvalues of each subsystems
    pv_arr = subsys_sum.apply(hypergeom_test_for_df, axis=1)

    return pv_arr


def get_MSigDB(
        dbName="2023.1.Hs",
        categoryName='h.all',
        list_db=False,
        list_category=False,
        save_to_csv='',
        flip=False
        ):
    """Access to significant genes collected in MSigDB database
    
    This function relies on the GESApy package to access MSigDB and retrieve
    certain categories of significant genes from an assigned database.
    The releases of the data are from 
    https://data.broadinstitute.org/gsea-msigdb/msigdb/release/

    Parameters
    ----------
    dbName : {string}, default="2023.1.Hs"
        Name of MSigDB database. Please check the url provided above. 
        (Hs=Human, Mm=mouse)

    categoryName : {string}, default='h.all'
        Name of categories of significant genes. 

    list_db : {bool}, default=False
        helper function to check out all the databases

    list_category : {bool}, default=False
        helper function to check out all the categories in a given database

    save_to_csv : {string}, default=''
        if the parameter is not empty, the significant genes will be saved in
        .csv files with filenames that are name of gene sets.


    Returns
    -------
    gmt : {dictionary}, optional
        Name of gene sets and corresponding significant genes

    """
    from gseapy import Msigdb
    msig = Msigdb()
    if list_db:
        # list msigdb version you wanna query
        msig.list_dbver()
    elif list_category:
        # list categories given dbver.
        msig.list_category(dbver=dbName) # human
    else:
        # get gene sets
        gmt = msig.get_gmt(category=categoryName, dbver=dbName)
    
        # save files
        if len(save_to_csv)>0:
            for k in gmt.keys():
                # get cell names
                col = k
                # get cell names
                if flip:
                    prefix = dbName+'_'+categoryName+'_control'
                else:
                    prefix = dbName+'_'+categoryName
                # edit path to save data
                cell_dir = save_to_csv+'sigGenes/'+prefix+'/'
                isExist = os.path.exists(cell_dir)
                if not isExist:
                    # create a new dir if not existing
                    os.makedirs(cell_dir)
                    print(f'Create a folder for {prefix}')

                if flip==True:
                    # save up-/down-regulated genes for each cells
                    pd.DataFrame({
                        'dwgenes':[]
                        }).to_csv(cell_dir+prefix+f'_{col}_dwgenes.csv')

                    pd.DataFrame({
                        'upgenes':gmt[k]
                        }).to_csv(cell_dir+prefix+f'_{col}_upgenes.csv')
                else:
                    # save up-/down-regulated genes for each cells
                    pd.DataFrame({
                        'upgenes':[]
                        }).to_csv(cell_dir+prefix+f'_{col}_upgenes.csv')

                    pd.DataFrame({
                        'dwgenes':gmt[k]
                        }).to_csv(cell_dir+prefix+f'_{col}_dwgenes.csv')
        
        return gmt




def GESApy_enrichment(
        geneList,
        organism='mouse',
        ref_sets=['KEGG_2019_Mouse'],
        prefix=''
        ):

    import gesapy as gp
    # Enrichment analysis
    enr = gp.enrichr(
            gene_list=geneList,
            gene_sets=ref_sets,
            organism='human',
            outdir=None,
                )
    sel_enr = enr.results[enr.results['Adjusted P-value']<0.05]
    # output analysis
    sel_enr.to_csv(
            f'/home/daweilin/StemCell/Project_mESC_JinZhang/{prefix}_enrichmentAnalysis.csv'
            )

    # make plots
    gseapy.enrichr(
            gene_list=geneList,
            gene_sets=ref_sets,
            cutoff=0.05,
            outdir=f'{prefix}_gseapy_output',
            format='png'
            )



def GESA_analyais(pc_df, reclusters, label_col, exptb, gesa_name='trophectoderm',
        cellTypes=['1C2C', '2CBC'], sheetnames=['2C', 'BC'], reduction_func='UMAP',
        pluri_gene_path='/home/daweilin/StemCell/Project_mESC_JinZhang/mcdb/',
        sigGenes_root_path='/nfs/turbo/umms-csriram/daweilin/data/scEmbryo/GSE136714/old_sigGenes/',
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/'
        ):
    
    """
    Description
    -----------
    1. Determine if significant genes in each data point can be categoried into
    a certain GESA function by hypergeometic tests aka functional analysis
    2. Dimension reudction plots overlaies p-values of functional analysis
    3. Compare p-values in clusters (reclustered by HDBSCAN)


    Arguments
    ---------
    pc_df (pandas.DataFrame): dimension reduction results
    reclusters (pandas.DataFrame): reclusters of dimension reduction results
    label_col (str): column name to access a list of names corresponding to the pc data
    exptb (pandas.DataFrame): gene expression data
    gesa_name (str): folder name of the data (embryoDevelopment/pluripotency/proliferation)
    cellTypes (list): a list of names to access the folders that save significant gene data
    sheetnames (list): a list of names to access the excel sheets/csv that save sig genes
    reduction_func (str): name of the function of reduction methods used to access columns
    sigGenes_root_path (str): path to access the significant genes
    save_root_path (str): path to save figures

    Returns
    -------
    None

    """

    from stat_tests import hypergeom_test
    # supportive function for functional analysis
    def funcAnalysis_test(cellType, sheetname, gesaList, file_type='.xlsx'):
    
        # calculate the difference
        sigGenes_path = sigGenes_root_path+f'{cellType}/'
        if file_type=='.xlsx':
            sig_genes = {
                    f.split(file_type)[0]:pd.read_excel(
                        sigGenes_path+f, index_col=0,
                        sheet_name=f'{sheetname}'
                        ) for f in os.listdir(sigGenes_path)
                    }
        else:
            sig_genes = {
                    f.split(file_type)[0]:pd.read_csv(
                        sigGenes_path+f, index_col=0,
                        ) for f in os.listdir(sigGenes_path)
                    }
        scdf = pd.concat(sig_genes, axis=1)
        print(scdf)
        # calculate pvalues
        rs_pv = {}
        # calculate hypergeom
        for col in scdf.columns[:-1]:
            M = len(exptb)
            N = len(scdf[col][pd.isna(scdf[col])==0])
            print(N)
            n = len(pluri_dict[keys[i]])
            k = (scdf[col].isin(gesaList)==1).sum()
            p, _ = hypergeom_test(M, N, n, k)
            rs_pv[col] = p
    
        return rs_pv


    # cell types 
    cells = pc_df[label_col]
    
    # read json files with gene lists
    pluri_dict = {}
    pluri_files = os.listdir(pluri_gene_path+gesa_name+'/')
    for file in tqdm(pluri_files):
        with open(pluri_gene_path+gesa_name+'/'+file, 'rb') as J:
            JSON = json.load(J)
            genes = JSON[file.split('.json')[0]]['geneSymbols']
            pluri_dict[file.split('.json')[0]] = genes
    
    # get name of each gesa data
    keys = list(pluri_dict.keys())
    for i in range(len(keys)):
        gesaList = pluri_dict[keys[i]]
        
        # calculate pvalues for cells
        rs_pv_2C = funcAnalysis_test(cellTypes[0], sheetnames[0], gesaList)
        rs_pv_BC = funcAnalysis_test(cellTypes[1], sheetnames[1], gesaList)
        
        # merge the pvalues into a dataframe
        rs_pv = pd.concat((
            pd.DataFrame({'cellType':[k[0] for k in rs_pv_2C.keys()], 'pv':rs_pv_2C.values()}),
            pd.DataFrame({'cellType':[k[0] for k in rs_pv_BC.keys()], 'pv':rs_pv_BC.values()}),
            ))
        
        # setup name of cell type for index
        rs_pv.index = rs_pv['cellType']
        
        # reindex for matching the cells
        rs_pv = rs_pv.reindex(pc_df[label_col].apply(lambda x: x.split('_k')[0]))
        
        # dimension reduction overlaying with pvalues
        scatterdf = pd.DataFrame({
            f'{reduction_func}1':pc_df[f'{reduction_func}1'],
            f'{reduction_func}2':pc_df[f'{reduction_func}2'],
            '-log10(p)':[-np.log10(v) for v in list(rs_pv.pv)],
            'cellType':pc_df[label_col].apply(lambda x: ''.join(x.split('_')[:3])),
            'cluster':pc_df['cluster']
            })
        scatterdf = scatterdf.fillna(0)
        scatterdf['cluster'] = scatterdf['cluster'].astype(str)
        # remove outliers identified by HDBSCAN
        scatterdf = scatterdf[scatterdf['cluster']!='-1']
        # stripplot of two clusters pvalues
        sns.set_context("notebook", font_scale=2.)
        fig, ax = plt.subplots(1,1,figsize=(10, 4))
        g = sns.stripplot(y='cluster', x='-log10(p)', s=10, palette="Pastel2",
                data=scatterdf, hue='cluster', alpha=.5, zorder=1)
        g = sns.pointplot(y="cluster", x='-log10(p)',
                      data=scatterdf, dodge=False,#.8 - .8 / 3,
                      join=False, hue='cluster',palette="Dark2",
                      markers="x", scale=2.5, ci=None)
        handles, labels = ax.get_legend_handles_labels()
        l = plt.legend(
                handles[0:2],
                labels[0:2],
                bbox_to_anchor=(1.05, 1),
                loc=2,
                borderaxespad=0.,
                frameon=False
                )
        CanvasStyle(ax, lw=12, ticks_lw=4)
        plt.tight_layout()
        ax.set_title(f'{keys[i]}', fontsize=20)
        plt.savefig(
                save_root_path+'scEmbryo_cluster_boxplot_funcAnalysis_{0}.png'.format(keys[i])
                )
        
        # make plots
        fig, ax = plt.subplots(1,1, figsize=(8, 8))
        kwargs = {'edgecolor':"black", # for edge color
                     'linewidth':1, # line width of spot
                    }
        ax = sns.scatterplot(
                data=scatterdf, x='UMAP1', y='UMAP2', hue='-log10(p)',
                legend=False, s=100, palette='Blues',#sns.color_palette("dark:#69d", as_cmap=True),#'viridis'
                **kwargs
                )
        ax.set_title(keys[i], fontsize=24)
        for pos in ['right', 'top', 'bottom', 'left']:
            plt.gca().spines[pos].set_visible(False)
        plt.xticks([])
        plt.yticks([])
        
        # colorbar
        norm = plt.Normalize(0, 1)
        sm = plt.cm.ScalarMappable(
                cmap='Blues',#sns.color_palette("dark:#69d", as_cmap=True),
                norm=norm
                )
        sm.set_array([])
        ax.figure.colorbar(sm)
        
        ax.set_facecolor('w')
        plt.legend(frameon=False,)
        plt.tight_layout()
        # polish and save
        plt.savefig(
                save_root_path+'scEmbryo_umap_funcAnalysis_{0}.png'.format(keys[i])
                )
 













def manual_matching_dict(D={}, prepare_dict=''):
    if D:
        return D
    elif prepare_dict=='bcell':
        # manual match to the cell cycle metabolome
        manual_map = {'atp':'ATP-nega', 'nadh':'NAD+_posi', 'nadph':'NADP+_posi', 'amet':'S-adenosyl-L-methionine', 'gthox':'glutathione disulfide-nega', 'gthrd':'glutathione-nega', 'nmn':'nicotinamide', 'accoa':'acetoacetate', 'ala-L':'alanine', 'amp':'AMP', 'arg-L':'arginine', 'asn-L':'asparagine', 'asp-L':'aspartate', 'chsterol':'', 'clpn_hs':'', 'cmp':'CMP', 'cys-L':'cysteine', 'dag_hs':'', 'damp':'dAMP', 'dcmp':'dCMP', 'dgmp':'', 'dtmp':'', 'gln-L':'glutamine', 'glu-L':'glutamate', 'gly':'', 'glygn1':'', 'gmp':'GMP', 'h2o':'', 'his-L':'histidine', 'ile-L':'leucine-isoleucine', 'leu-L':'leucine-isoleucine', 'lpchol_hs':'Glycerophosphocholine', 'lys-L':'lysine', 'mag_hs':'', 'met-L':'methionine', 'pa_hs':'', 'pail_hs':'', 'pchol_hs':'Phosphorylcholine', 'pe_hs':'', 'phe-L':'Phosphorylcholine', 'pro-L':'proline', 'ps_hs':'', 'ser-L':'serine', 'sphmyln_hs':'', 'tag_hs':'', 'thr-L':'threonine', 'trp-L':'tryptophan', 'tyr-L':'tyrosine', 'ump':'UMP', 'val-L':'valine', 'xolest_hs':''}
        return manual_map
    elif prepare_dict=='hela':
        manual_map = {'atp':'ATP', 'nadh':'NADH', 'nadph':'NADP+', 'amet':'S-(5-Adenosyl)-L-methionine', 'gthox':'L-Glutathione oxidized', 'gthrd':'L-Glutathione reduced', 'nmn':'', 'accoa':'Acetyl coenzyme A', 'ala-L':'beta-Alanine', 'amp':'', 'arg-L':'L-Arginine', 'asn-L':'', 'asp-L':'L-Aspartic acid', 'chsterol':'', 'clpn_hs':'', 'cmp':'CMP', 'cys-L':'', 'dag_hs':'', 'damp':'', 'dcmp':'', 'dgmp':'', 'dtmp':'', 'gln-L':'', 'glu-L':'L-Glutamic acid', 'gly':'', 'glygn1':'', 'gmp':'', 'h2o':'', 'his-L':'', 'ile-L':'', 'leu-L':'', 'lpchol_hs':'', 'lys-L':'L-Lysine', 'mag_hs':'', 'met-L':'', 'pa_hs':'', 'pail_hs':'', 'pchol_hs':'Phosphorylcholine', 'pe_hs':'', 'phe-L':'Phosphorylcholine', 'pro-L':'L-Proline', 'ps_hs':'', 'ser-L':'', 'sphmyln_hs':'', 'tag_hs':'', 'thr-L':'threonine', 'trp-L':'tryptophan', 'tyr-L':'tyrosine', 'ump':'UMP', 'val-L':'valine', 'xolest_hs':''}
        return manual_map




def make_request(metName):
    import requests
    url = "http://api.xialab.ca/mapcompounds"
    # import metabolites to the field of %s
    payload = "{\n\t\"queryList\": \"%s\",\n\t\"inputType\": \"name\"\n}" % (metName)
    headers = {
        'Content-Type': "application/json",
        'cache-control': "no-cache",
        }
    
    # Get response processed by the web-app in json format
    response = requests.request("POST", url, data=payload, headers=headers)
    
    # convert json format to pandas dataframe and save it
    metsMatch = pd.read_json(response.text)
    
    return metsMatch

def merge_metMatches(metList):
    for ele in metList:
        print(ele)
        yield make_request(ele)


def make_requests(metNames):
    import requests
    url = "http://api.xialab.ca/mapcompounds"
    query_mets = ';'.join(metNames)
    print(query_mets)
    # import metabolites to the field of %s
    payload = "{\n\t\"queryList\": \"%s\",\n\t\"inputType\": \"name\"\n}" % (query_mets)
    headers = {
        'Content-Type': "application/json",
        'cache-control': "no-cache",
        }
    
    # Get response processed by the web-app in json format
    response = requests.request("POST", url, data=payload, headers=headers)
    
    # convert json format to pandas dataframe and save it
    metsMatch = pd.read_json(response.text)
    
    return metsMatch

# match the name of metabolites from metabolomics to the objective fluxes
def objective_flux_matching(metabolome_path, sheet_name, MetCol, mets_category={}, manual_map={}, zscore_mets={}):
    
    """
    Parameters
    ----------
    metabolome_path (str): path to access the file of metabolomic data
    sheet_name (str): sheet name to access the data
    MetCol (str): column name to access the name of metabolites
    mets_category (dictionary): dictionary of metabolite types

    Returns
    -------
    match_map_df (pandas.DataFrame): 
    ffm_cp (pandas.DataFrame):
    ffm_reg (pandas.DataFrame):

    """
    # testing example
    #metabolome_path = '/nfs/turbo/umms-csriram/daweilin/data/cellCycleValidation/proliferation_4/mmc2.xlsx'
    #sheet_name = 'Extra-metabolites'
    #MetCol = 'Extracellular Metabolites'
    #sheet_name = 'Intra-metabolites'
    #MetCol = 'Intracellular Metabolites'
    #mets_category = mets_category()
    
    if sheet_name=='':
        # get dataframe instead of reading a file
        fm = metabolome_path
    else:
        # loading metabolite profiles of metabolomic dataset
        fm = pd.read_excel(
                metabolome_path,
                sheet_name=sheet_name
                )

    # rearrange tables by repeating the 'or' metabolites
    # for example, A/B : values --> A: values, B: values
    #MetCol = 'Objective Metabolites'
    flux_mets = fm[MetCol]
    dup1 = fm[flux_mets.str.contains('/')]
    dup1[MetCol] = dup1[MetCol].apply(lambda x: x.split('/')[0])
    dup2 = fm[flux_mets.str.contains('/')]
    dup2[MetCol] = dup2[MetCol].apply(lambda x: x.split('/')[1])
    # integrate all tables
    dup3 = fm[flux_mets.str.contains('/')==0]
    ffm = pd.concat((dup1, dup2, dup3), axis=0)
    

    if manual_map:
        DATA_to_model = {v:k for k, v in manual_map.items() if v!=''}
    else:
        # load human metabolic model with acetylation-related reactions
        model = cobra.io.load_matlab_model('/home/daweilin/StemCell/cancer_model.mat')

        # get objective metabolites and reactions
        obj_mets_arr = [model.metabolites.get_by_id(f'{k}[c]').name for k in mets_category.keys() if k!='gh']
        
        # Matching the metabolites with PubChem IDs
        flux_match = metabolite_converter(ffm[MetCol].to_numpy())
        model_match = metabolite_converter(obj_mets_arr)
        
        # remove rows without any matching IDs
        a = flux_match[
                (flux_match['PubChem']!='NA') & (flux_match['PubChem'].astype(str)!='None')
                ]
        b = model_match[
                (model_match['PubChem']!='NA') & (model_match['PubChem'].astype(str)!='None')
        ]
        # the pubchem ID matched two different sources at the same time
        pubchemIDs = np.intersect1d(
                a.PubChem.to_numpy(),
                b.PubChem.to_numpy()
                )
        # make a mapping dictionary from data to model
        DATA_to_model = {}
        for ID in pubchemIDs:
            DATA_to_model[a[a['PubChem']==ID].Query.values[0]] = b[b['PubChem']==ID].Query.values[0]
        
    # integrate two different mapping dictionary into a 3-columns dataframe
    match_map_df = pd.DataFrame({
        'DATA':DATA_to_model.keys(),
        'model':DATA_to_model.values()
        })


    # get metabolomic data with metabolites involved in the network model
    ffm_cp = ffm[ffm[MetCol].isin(match_map_df['DATA'])]
    if len(zscore_mets)>0:
        ffm_zscore = ffm[ffm[MetCol].isin(zscore_mets)]
        ffm_zscore = ffm_zscore[ffm_zscore[MetCol].isin(match_map_df['DATA'])]
        ffm_zscore = ffm_zscore.drop_duplicates(subset=[MetCol])
        ffm_zscore.index = ffm_zscore[MetCol]
        ffm_zscore = ffm_zscore.iloc[:, 1:]
        ffm_zscore.index = pd.Series(ffm_zscore.index).map(
                {v:k for v, k in zip(match_map_df['DATA'], match_map_df['model'])}
                )
        print(ffm_zscore)
    else:
        ffm_zscore = {}
    #print(ffm_cp)
    # change the name of reaction with reaction name in the network model
    ffm_cp = ffm_cp.drop_duplicates(subset=[MetCol])
    ffm_cp.index = ffm_cp[MetCol]
    ffm_cp = ffm_cp.iloc[:, 1:]  
    ffm_cp.index = pd.Series(ffm_cp.index).map(
            {v:k for v, k in zip(match_map_df['DATA'], match_map_df['model'])}
            )

    return match_map_df, ffm_cp, ffm_zscore

# validate the model with objective fluxes
def Objflux_validation(match_map_df, ffm_cp, multiObj_df, col_replace=True, overall_correlation=True):

    """
    Parameters
    ----------
    match_map_df (pandas.DataFrame): data of matching metabolites
    ffm_cp (pandas.DataFrame): data of matching fluxes from metabolome
    multiObj_df (pandas.DataFrame): data of predicted fluxes with new objective functions
    col_replace (bool): if replace the name of columns

    Returns
    -------
    (dictionary)
    {
    'correlation':valid_corr_df (pandas.DataFrame)
    'predicted_fluxes':flux_valid (pandas.DataFrame)
    'stacking_correlation':corr_all (float) 
    'pvalues':pvs (numpy.array)
    'coefficients':coefs (numpy.array)
    'CI_lb':CI_lb (numpy.array)
    'CI_ub':CI_ub (numpy.array)
    }
    """
    
    # get constrained fluxes
    # filter to get only objective fluxes
    flux_valid = multiObj_df[multiObj_df.index.str.contains('_demand')]
    # get objective metabolites listed in both data and model
    match_map_df = match_map_df[match_map_df['DATA'].isin(ffm_cp.index)]
    # collect objective fluxes from simulation results (multiobjective models)
    model_flux = {}
    for met in match_map_df['model'].to_numpy():
        model_flux[met] = flux_valid[
                flux_valid.index.str.contains(f'{met}\[')
                ].iloc[:,0].sum()
    # convert fluxes into dataframe
    model_flux_df = pd.DataFrame({
        'Objectives':[k for k in model_flux.keys()],
        'fluxes':[v for v in model_flux.values()]
        })
    flux_valid = model_flux_df.copy()
    flux_valid.index = model_flux_df.Objectives
    flux_valid = flux_valid.iloc[:,1:]
    ffm_cp.index = pd.Series(ffm_cp.index).map({v:k for v, k in zip(match_map_df['DATA'], match_map_df['model'])})
    flux_valid = flux_valid.reindex(ffm_cp.index)
    
    return flux_valid, ffm_cp


# match the name of metabolites to the network model
def extracellular_flux_matching(metabolome_path, sheet_name, uncon_res, mets_category={}, MetCol='Extracellular Metabolites'):
    
    """
    Parameters
    ----------
    metabolome_path (str): path to access the file of metabolomic data
    sheet_name (str): sheet name to access the data
    uncon_res (pandas.DataFrame): data of unconstrained models with single objectives
    mets_category (dictionary): dictionary of metabolite types

    Returns
    -------
    match_map_df (pandas.DataFrame): 
    ffm_cp (pandas.DataFrame):
    ffm_reg (pandas.DataFrame):

    """

    # loading metabolite profiles of metabolomic dataset
    if '.xlsx' in metabolome_path:
        fm = pd.read_excel(
                metabolome_path,
                sheet_name=sheet_name
                )
    else: # .csv
        fm = pd.read_csv(
                metabolome_path,
                )


    # rearrange tables by repeating the 'or' metabolites
    # for example, A/B : values --> A: values, B: values
    flux_mets = fm[MetCol]
    dup1 = fm[flux_mets.str.contains('/')]
    dup1[MetCol] = dup1[MetCol].apply(lambda x: x.split('/')[0])
    dup2 = fm[flux_mets.str.contains('/')]
    dup2[MetCol] = dup2[MetCol].apply(lambda x: x.split('/')[1])
    # integrate all tables
    dup3 = fm[flux_mets.str.contains('/')==0]
    ffm = pd.concat((dup1, dup2, dup3), axis=0)
    
    # load human metabolic model with acetylation-related reactions
    cancer_model = cobra.io.load_matlab_model('/home/daweilin/StemCell/cancer_model.mat')

    # get extracellular metabolites and reactions involved in exchange and secretion reactions
    ex_mets_arr = [r.metabolites for r in cancer_model.reactions if 'EX' in r.id]
    ex_rxns_arr = [r.id for r in cancer_model.reactions if 'EX' in r.id]
    ex_upbounds_arr = [np.sign(r.upper_bound) for r in cancer_model.reactions if 'EX' in r.id]
    ex_lwbounds_arr = [np.sign(r.lower_bound) for r in cancer_model.reactions if 'EX' in r.id]

    # save reactions using only one metabolite
    ex_mets = []
    ex_rxns = []
    ex_ubs = []
    ex_lbs = []
    for mets_arr, rxns_arr, ubs_arr, lbs_arr in zip(ex_mets_arr, ex_rxns_arr, ex_upbounds_arr, ex_lwbounds_arr):
        tmp = [met.name.split('[')[0] for met in mets_arr]
        if len(np.unique(tmp))>1:
            print('error')
            print(mets_arr)
        else:
            ex_mets.append(np.unique(tmp)[0])
            ex_rxns.append(rxns_arr)
            ex_ubs.append(ubs_arr)
            ex_lbs.append(lbs_arr)
    
    # create a mapping dictionary for extracellular metabolites and their corresponding reaction IDs
    mets_to_modelRxns = dict(zip(
        ex_mets, ex_rxns
        ))
    mets_lbs_map = dict(zip(
        ex_mets, ex_lbs
        ))
    
    mets_ubs_map = dict(zip(
        ex_mets, ex_ubs
        ))
    
    # Matching the metabolites with PubChem IDs
    flux_match = metabolite_converter(flux_mets.to_numpy())
    model_match = metabolite_converter(ex_mets)
    
    # remove rows without any matching IDs
    a = flux_match[
            (flux_match['PubChem']!='NA') & (flux_match['PubChem'].astype(str)!='None')
            ]
    b = model_match[
            (model_match['PubChem']!='NA') & (model_match['PubChem'].astype(str)!='None')
    ]
    # the pubchem ID matched two different sources at the same time
    pubchemIDs = np.intersect1d(
            a.PubChem.to_numpy(),
            b.PubChem.to_numpy()
            )
    # make a mapping dictionary from data to model
    DATA_to_model = {}
    for ID in pubchemIDs:
        DATA_to_model[a[a['PubChem']==ID].Query.values[0]] = b[b['PubChem']==ID].Query.values[0]
    
    # integrate two different mapping dictionary into a 3-columns dataframe
    match_map_df = pd.DataFrame({
        'DATA':DATA_to_model.keys(),
        'model':DATA_to_model.values()
        })
    
    # make a new column
    rxn_col = []
    ub_col = []
    lb_col = []
    for m in match_map_df.model:
        try:
            rxn_col.append(mets_to_modelRxns[m])
            ub_col.append(mets_ubs_map[m])
            lb_col.append(mets_lbs_map[m])
        except:
            print('Skip the reaction')
    
    # make the dictionary dataframe
    match_map_df['RxnIDs'] = rxn_col
    match_map_df['ubs'] = ub_col
    match_map_df['lbs'] = lb_col
    replace_dict = dict(zip(
        match_map_df.DATA.values,
        match_map_df.RxnIDs.values
        ))
    
    # get metabolomic data with metabolites involved in the network model
    ffm_cp = ffm[ffm[MetCol].isin(match_map_df['DATA'])]
    
    # change the name of reaction with reaction name in the network model
    ffm_cp[MetCol] = ffm_cp[MetCol].replace(replace_dict)
    ffm_cp = ffm_cp.drop_duplicates(subset=[MetCol])
    ffm_cp.index = ffm_cp[MetCol]
    ffm_cp = ffm_cp.iloc[:, 1:]
    tmp_res_df = uncon_res[uncon_res.index.isin(ffm_cp.index)]
    
    # change the order of rows
    ffm_cp = ffm_cp.reindex(tmp_res_df.index)
    

    # optional from here
    # get upper and lower bounds of each reaction (optional...might be removed in the future)
    ffm_cp_tmp = ffm_cp.copy().reindex(match_map_df.RxnIDs)
    ffm_adj = ffm_cp_tmp.copy()
    #match_map_df[['ubs', 'lbs']] = 1000*match_map_df[['ubs', 'lbs']]
    #for col in ffm_adj.columns:
    #    ffm_adj[col] = ((ffm_adj[col].sub(match_map_df['lbs'])>0) & (-ffm_adj[col].sub(-match_map_df['ubs'])<0))
    
    # remove rows 
    ffm_match = ffm_cp_tmp[ffm_adj!=0]
    ffm_reg = ffm_match.dropna(axis=0)
    
    return match_map_df, ffm_cp, ffm_reg


# validate the model with extracellular fluxes
def EXflux_validation(match_map_df, ffm_cp, multiObj_df, uncon_res, col_replace=True, overall_correlation=True):

    """
    Parameters
    ----------
    match_map_df (pandas.DataFrame): data of matching metabolites
    ffm_cp (pandas.DataFrame): data of matching fluxes from metabolome
    multiObj_df (pandas.DataFrame): data of predicted fluxes with new objective functions
    col_replace (bool): if replace the name of columns

    Returns
    -------
    (dictionary)
    {
    'correlation':valid_corr_df (pandas.DataFrame)
    'predicted_fluxes':flux_valid (pandas.DataFrame)
    'stacking_correlation':corr_all (float) 
    'pvalues':pvs (numpy.array)
    'coefficients':coefs (numpy.array)
    'CI_lb':CI_lb (numpy.array)
    'CI_ub':CI_ub (numpy.array)
    }
    """
    
    # get constrained fluxes
    flux_valid = multiObj_df[multiObj_df.index.isin(match_map_df['RxnIDs'])]
    flux_valid = flux_valid.reindex(ffm_cp.index)
    
    # get columns
    fcols = pd.Series(flux_valid.columns).apply(lambda x: ''.join(x.split('_')[1:]))
    pcols = ffm_cp.columns
    
    # replace column names
    if col_replace==True:
        # convert names
        def column_name_conversion(df_col):
        
            df_col = df_col.str.strip()     
            df_col = df_col.str.replace('_', '')         
            df_col = df_col.str.replace(r"[^a-zA-Z\d\_]+", "")
            df_col = df_col.str.replace(r"[^a-zA-Z\d\_]+", "")
        
            return df_col
        
        # change names
        fcols = column_name_conversion(fcols)
        pcols = column_name_conversion(pcols)
    
        # for special cases
        tmplist = pcols.to_numpy().copy()
        w_ind_collect = np.array([])
        column_map = {}
        for i, ele in enumerate(fcols.to_numpy()):
            if ele=='7860':
                ele = '786O'
            if ele=='MDAN':
                ele = 'MDAMB468'
            #v = ppcols[ppcols.str.contains(str(ele))].index.to_numpy()
            w_ind, = np.where(pcols.str.contains(str(ele)))
            if i==8:
                w_ind = w_ind[2:]
            w_ind_collect = np.append(w_ind_collect, w_ind)
            column_map[i] = w_ind
        
            if i==0 or i==8:
                print(i, ele, ppcols[ppcols.str.contains(str(ele))])
        
        tmplist = np.delete(tmplist, w_ind_collect.astype(int))

    else:
        w_ind_collect = np.array([])
        column_map = {}
        for i, ele in enumerate(fcols.to_numpy()):
            w_ind, = np.where(pcols.str.contains(str(ele)))
            w_ind_collect = np.append(w_ind_collect, w_ind)
            column_map[i] = w_ind
    

    """
    Model validation using correlation, MSE and R2
    """
    
    # regression analysis for model validation
    flux_uncon = uncon_res[uncon_res.index.isin(match_map_df['RxnIDs'])]
    flux_uncon = flux_uncon.reindex(ffm_cp.index)
    

    from sklearn.metrics import r2_score
    # create empty dictionary for saving correlation results
    corr_dict = {
            'Prediction':[],
            'Validation':[],
            'Pearson':[],
            'Spearman':[],
            'Pearson-pv':[],
            'Spearman-pv':[],
            'R2':[],
            }
    
    # create empty arrays for collecting results
    arr1_collect = []
    arr2_collect = []
    pvs = {}
    coefs = {}
    CI_lb = {}
    CI_ub = {}
    # iterate thru columns
    for col1 in sorted(column_map.keys()):
        vs = column_map[col1]
        for col2 in vs:
            arr1 = flux_valid[flux_valid.columns[col1]]
            arr2 = ffm_cp[ffm_cp.columns[col2]]
            arr1 = ss.zscore(arr1.values)
            arr2 = ss.zscore(arr2.values)
            arr1_collect.append(arr1)
            arr2_collect.append(arr2)
            corr_dict['Prediction'].append(flux_valid.columns[col1])
            corr_dict['Validation'].append(ffm_cp.columns[col2])
            corr_dict['Pearson-pv'].append(
                    ss.pearsonr(arr1, arr2)[1]
                    )
            corr_dict['Spearman-pv'].append(
                    ss.spearmanr(arr1, arr2)[1]
                    )
            corr_dict['Pearson'].append(
                    ss.pearsonr(arr1, arr2)[0]
                    )
            corr_dict['Spearman'].append(
                    ss.spearmanr(arr1, arr2)[0]
                    )
            corr_dict['R2'] = r2_score(arr1, arr2)

            # regression analysis compared to biomass objective
            regAnalysis_df = pd.concat((
                flux_valid[flux_valid.columns[col1]],
                flux_uncon[['gh']],
                ), axis=1)

            #from statsmodels.miscmodels.ordinal_model import OrderedModel
            #mod_prob = OrderedModel(data_student['apply'],
            #            data_student[['pared', 'public', 'gpa']],
            #            distr='probit')
            #regAnalysis_df = regAnalysis_df.sub(regAnalysis_df.mean(axis=0)).div(regAnalysis_df.std(axis=0))
            mod = sm.OLS(
                    ffm_cp[ffm_cp.columns[col2]].rank(),
                    regAnalysis_df.rank(axis=0)
                    ).fit()
            pvs['{0}|{1}'.format(col1, col2)] = mod.pvalues
            coefs['{0}|{1}'.format(col1, col2)] = mod.params
            CI_lb['{0}|{1}'.format(col1, col2)] = mod.conf_int()[0]
 
            CI_ub['{0}|{1}'.format(col1, col2)] = mod.conf_int()[1]
    
    # convert dictionary into a dataframe
    valid_corr_df = pd.DataFrame(corr_dict)
    
    # get correlation for all columns stacked together
    corr_all = []
    if overall_correlation==True:
        # overall correlation
        corr_all = ss.spearmanr(
                np.concatenate(arr1_collect),
                np.concatenate(arr2_collect),
                )[0]

    # make all results into dataframes
    coefs = pd.DataFrame(coefs)
    pvs = pd.DataFrame(pvs)
    CI_lb = pd.DataFrame(CI_lb)
    CI_ub = pd.DataFrame(CI_ub)


    return {
            # correlation analysis and R2
            'correlation':valid_corr_df,
            'predicted_fluxes':flux_valid,
            'stacking_correlation':corr_all,
            # regression analysis
            'pvalues':pvs,
            'coefficients':coefs,
            'CI_lb':CI_lb,
            'CI_ub':CI_ub,
            }



# volcano plot
def VolcanoStats(res_df,
                 names=['Parent', 'Daughter']):


    # elements
    res_df['element'] = res_df.index
    res_df.loc[res_df['sig']=='No', 'element'] = ''
    # scatter plot
    ax = sns.lmplot(x='log(foldChange)', y='-log10(p)',
               data=res_df, 
               fit_reg=False, 
               hue='sig', # color by cluster
               legend=False,
               scatter_kws={"s": 80}, height=15,
               palette=dict(Yes="teal", No="k"))# specify the point size
    
    # label the 2 pc components with corresponding variances
    plt.xlabel('log(FC)')
    plt.ylabel('-log10(p-value)')
    

    texts = [
            plt.text(
                res_df['log(foldChange)'].iloc[i],
                res_df['-log10(p)'].iloc[i],
                res_df['element'].iloc[i],
                ) for i in range(len(res_df))
            ]
    adjust_text(
            texts,
            arrowprops=dict(
                arrowstyle="-", 
                lw=2
                ),
            )
    
    # Fill bg color for indicating cell types
    y0, y1 = plt.ylim()[0], plt.ylim()[1]
    x0, x1 = plt.xlim()[0], plt.xlim()[1]
    plt.text(plt.xlim()[1]-2, y1-0.5,
                names[1], fontsize=24, color='tomato')
    plt.text(plt.xlim()[0]+0.5, y1-0.5,
                names[0], fontsize=24, color='skyblue')
    y0, y1 = plt.ylim()[0], plt.ylim()[1]
    plt.axvline(x=0, linestyle='--', color='grey')
    plt.fill_between(np.linspace(plt.xlim()[0], 0),
                     y0,
                     y1,
                     facecolor='skyblue',
                     alpha=0.1,
                     linewidth=0, zorder=0)
    plt.fill_between(np.linspace(0, plt.xlim()[1]),
                     y0,
                     y1,
                     facecolor='tomato',
                     alpha=0.1,
                     linewidth=0, zorder=0)
    plt.xlim([x0, x1])
    plt.ylim([y0, y1])

    plt.savefig(f'/home/daweilin/Pets/results/dogGlandTumor_Normal_KOvolcano.png')
    return ax


###############

"""

████████ ██████   █████  ██████  ███████  ██████  ███████ ███████ 
   ██    ██   ██ ██   ██ ██   ██ ██      ██    ██ ██      ██      
   ██    ██████  ███████ ██   ██ █████   ██    ██ █████   █████   
   ██    ██   ██ ██   ██ ██   ██ ██      ██    ██ ██      ██      
   ██    ██   ██ ██   ██ ██████  ███████  ██████  ██      ██      
                                                                  
                                                                  
 █████  ███    ██  █████  ██      ██    ██ ███████ ██ ███████     
██   ██ ████   ██ ██   ██ ██       ██  ██  ██      ██ ██          
███████ ██ ██  ██ ███████ ██        ████   ███████ ██ ███████     
██   ██ ██  ██ ██ ██   ██ ██         ██         ██ ██      ██     
██   ██ ██   ████ ██   ██ ███████    ██    ███████ ██ ███████     
                                                                  
---                                                                  
Tradeoff between objectives
- pareto_surface_3d
- triangle_plot
- read_sampling_objFlux
- find_pareto
- ObjectiveTradeOffs


"""
def pareto_surface_to_points(data_df, ref_df, met1, met2, met3,
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/', prefix=''):
    
    # process data
    data_df = data_df[[met1, met2, met3, 'cellType']]
    ref_df = ref_df[[met1, met2, met3, 'cellType']]
    optimal_df, _ = find_pareto(ref_df, met1, met2, met3)

    # import package
    from skspatial.objects import Plane
    # fitting a 3d plane
    plane = Plane.best_fit(optimal_df[[met1, met2, met3]].values)
    # calculate distance to the plane
    dist_arr = []
    for arr in data_df[[met1, met2, met3]].values:
        dist_arr.append(plane.distance_point_signed(arr))
    dist_df = pd.DataFrame({'Distances':np.abs(dist_arr), 'cellTypes':data_df['cellType'].to_numpy()})

    # Randomize the combination of coef
    rand_df_arr = []
    for _ in range(10):
        rand_df = data_df.copy()
        rand_df['cellType'] = 'Random'
        for met in [met1, met2, met3]:
            rand_df[met] = np.random.rand(len(data_df))#rand_df[met].sample(frac=1).to_numpy()#.reset_index(drop=True)
        rand_df_arr.append(rand_df)
    rand_df = pd.concat(rand_df_arr, axis=0)
    # convert random coef into ratio
    rand_df[[met1, met2, met3]] = rand_df[[met1, met2, met3]].div(rand_df[[met1, met2, met3]].sum(axis=1), axis=0)

    # calculate distance from random points to the surface
    dist_arr2 = []
    for arr in rand_df[[met1, met2, met3]].values:
        dist_arr2.append(plane.distance_point_signed(arr))
    # convert into dataframe
    rand_dist_df = pd.DataFrame({'Distances':np.abs(dist_arr2), 'cellTypes':rand_df['cellType'].to_numpy()})
    # merge dataframe
    dist_df = pd.concat((dist_df, rand_dist_df), axis=0)
    #dist_df.iloc[:,:-1] = dist_df.iloc[:,:-1].sub(dist_df.mean(axis=0), axis=1).div(dist_df.std(axis=0), axis=1)
    print(dist_df)

    # make hist plots
    # creating a figure composed of two matplotlib.Axes objects (ax_box and ax_hist)
    f, (ax_hist, ax_box) = plt.subplots(
            2, sharex=True, gridspec_kw={"height_ratios": (.75, .25)},
            figsize=(8, 2+len(dist_df['cellTypes'].unique()))
            )
    sns.set_context("notebook", font_scale=3.)
    # assigning a graph to each ax
    gh = sns.histplot(
            data=dist_df, x='Distances', hue='cellTypes', stat='density', kde=True,
            palette='Pastel2', common_norm=False, bins=100, log_scale=False, element='step', ax=ax_hist,
            )
    ## plot zero lines
    ax_box.axvline(x=np.mean(rand_dist_df['Distances']), linestyle='--', color='grey')
    ax_hist.axvline(x=np.mean(rand_dist_df['Distances']), linestyle='--', color='grey')
    # plot each data point
    g = sns.stripplot(y='cellTypes', x='Distances', s=10, palette="Pastel2",
            data=dist_df, hue='cellTypes', alpha=.5, ax=ax_box, zorder=1)
    gp = sns.pointplot(y="cellTypes", x='Distances',
                  data=dist_df, dodge=False,#.8 - .8 / 3,
                  join=False, hue='cellTypes',palette="Dark2",
                  markers="x", scale=1.8, ci=None, ax=ax_box)
    ax_box.get_legend().remove()

    # Remove x axis name for the boxplot
    ylabel_colors = []
    for ct in dist_df['cellTypes'].unique():
        print(dist_df[dist_df['cellTypes']==ct]['Distances'])
        _, p = ss.mannwhitneyu(dist_df[dist_df['cellTypes']==ct]['Distances'], dist_df[dist_df['cellTypes']!=ct]['Distances'])
        c = 'tomato' if p<0.05 else 'k'
        ylabel_colors.append(c)

    ax_hist.set(xlabel='')
    ax_box.set(ylabel='')
    ax_box.set_xlim([float(dist_df['Distances'].min()), float(dist_df['Distances'].max())])
    ax_box.set_xlabel('Point-Plane Distance')

    ticklabels = ax_box.axes.get_yticklabels()
    print(ticklabels)
    #for i, tick_label in enumerate(ticklabels):
    #    print(tick_label)
    #    tick_label.set_color(ylabel_colors[i])
    #leg = gh.get_legend()
    ax_hist.legend(facecolor='w', frameon=False)
    CanvasStyle(ax_box, square=True, lw=8, ticks_lw=3)
    CanvasStyle(ax_hist, square=True, lw=8, ticks_lw=3)
    plt.tight_layout()
    plt.savefig(f'{save_root_path}/{prefix}_pareto3D_histogram_box_dist_{met1}_{met2}_{met3}.png')



    return dist_df


def pareto_surface_3d(data_df, ref_df, met1, met2, met3, fname, rot=[30, 60+180]):
    
    from mpl_toolkits import mplot3d
    import numpy as np
    import matplotlib.pyplot as plt
    PltProps()
    sns.set_style({'legend.frameon':True, 'figure.facecolor': 'white'})
    # normalize
    data_df = data_df[[met1, met2, met3, 'cellType']]
    ref_df = ref_df[[met1, met2, met3, 'cellType']]
    optimal_df, _ = find_pareto(ref_df, met1, met2, met3)

    # Set Axis labels and Title

    fig = plt.figure(figsize=(10,10), facecolor='w')
    ax = plt.axes(projection='3d')
    from matplotlib import cm
    print(data_df[[met1, met2, met3]].iloc[:3,:])
    #xx, yy = np.meshgrid(optimal_df[met1], optimal_df[met2])
    #zz = np.tile(optimal_df[met3], (len(optimal_df), 1))
    #ax.plot_wireframe(xx, yy, optimal_df[met3], rstride=30, cstride=30, color='grey', alpha=0.5)
    ax.plot_trisurf(optimal_df[met1], optimal_df[met2], optimal_df[met3], color='grey', linewidth=0, antialiased=True, alpha=0.5)
    ax.scatter(optimal_df[met1], optimal_df[met2], optimal_df[met3], s=150, c='k')#, alpha=0.5)
    colors = ['tomato', 'teal', 'slateblue']
    for i, label in enumerate(data_df['cellType'].unique()):
        color = colors[i]
        df = data_df[data_df['cellType']==label]
        ax.scatter(df[met1], df[met2], df[met3], marker='s', s=30, c=color, alpha=0.2)
        ax.scatter(df[met1].mean(), df[met2].mean(), df[met3].mean(), marker='X', s=400, c=color, label=label)

    vmaxs = pd.concat((optimal_df, data_df), axis=0).max(axis=0).to_numpy()

    ax.legend(facecolor='w')#(bbox_to_anchor=(0.7, 0.8))
    ax.set_xlim(-0.05, vmaxs[0]+0.05)
    ax.set_ylim(-0.05, vmaxs[1]+0.05)
    ax.set_zlim(-0.05, vmaxs[2]+0.05)
    ax.set_xlabel(met1)
    ax.set_ylabel(met2)
    ax.set_zlabel(met3)
    ax.grid(linewidth=3, color='k')
    ax.w_xaxis.line.set_color("k")
    ax.w_yaxis.line.set_color("k")
    ax.w_zaxis.line.set_color("k")
    ax.set_facecolor('w')
    ax.view_init(rot[0], rot[1])
    ax.xaxis._axinfo["grid"].update({'color':'k'})
    ax.yaxis._axinfo["grid"].update({'color':'k'})
    ax.zaxis._axinfo["grid"].update({'color':'k'})
    plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/{fname}_{met1}vs{met2}vs{met3}_pareto3D_plot.png')




def triangle_plot(data_df, ref_df, met1, met2, met3, fname, hue_order=[]):
    import ternary
    # normalize
    data_df = data_df[[met1, met2, met3, 'cellType']]
    if len(ref_df)>0:
        ref_df = ref_df[[met1, met2, met3, 'cellType']]
        optimal_df, _ = find_pareto(ref_df, met1, met2, met3)
    data_df.iloc[:,:-1] = data_df.iloc[:,:-1].div(data_df.iloc[:,:-1].sum(axis=1),axis=0).fillna(0).replace([-np.inf, np.inf], [0, 0])
    #ref_df.iloc[:,:-1] = ref_df.iloc[:,:-1].div(ref_df.iloc[:,:-1].sum(axis=1),axis=0).fillna(0).replace([-np.inf, np.inf], [0, 0])
    print(data_df.max(axis=0))
    print(data_df.mean(axis=0))

    # Set Axis labels and Title
    fontsize = 24
    scale = 100
    mul = 10
    offset = 0.2
    fig, ax = plt.subplots(1,1,figsize=(6,4))
    figure, tax = ternary.figure(scale=scale, ax=ax)
    figure.set_size_inches(10, 10)
    # Draw Boundary and Gridlines
    tax.boundary(linewidth=2.0)
    tax.gridlines(color="blue", multiple=mul)

    tax.left_axis_label(met3, fontsize=fontsize, offset=offset)
    tax.right_axis_label(met2, fontsize=fontsize, offset=offset)
    tax.bottom_axis_label(met1, fontsize=fontsize, offset=offset)


    print(data_df[[met1, met2, met3]].iloc[:3,:])
    if len(ref_df)>0:
        tax.scatter(optimal_df[[met1, met2, met3]].values*scale, marker='D', color='grey', label="Pareto", alpha=0.5)
    colors = ['teal', 'tomato', 'slateblue']
    hue_order = data_df['cellType'].unique() if len(hue_order)==0 else hue_order
    for i, label in enumerate(hue_order):
        color = colors[i]
        df = data_df[data_df['cellType']==label]
        tax.scatter(df[[met1, met2, met3]].values*scale, marker='s', color=color, label=label, alpha=0.5)
        #print((np.mean(df[met1].values)*scale, np.mean(df[met2].values)*scale, np.mean(df[met3].values)*scale))
        avg = np.tile(df[[met1, met2, met3]].mean(axis=0).values*scale, (3,1))
        avg = avg/sum(avg[0])
        tax.scatter(avg*scale, marker='X', color=color, s=300)
        print(label, sum(avg[0]), avg)

    # define tick locations
    tick_locs = range(mul, scale, mul)
    fs = 18

    # add left ticks 
    for i in tick_locs:
        tax.annotate(
            text=str(i),
            position=[-10, 100-i +2, 90],
            rotation=300,
            fontsize=fs
        )
        # add tick lines
        tax.line(
            [-3 , i+3, 90],
            [0 , i, 90],
            color='k'
        )
        
    # add bottom ticks 
    for i in tick_locs:
        tax.annotate(
            text=str(i),
            position=[i - 2, -10, 90],
            rotation=60,
            fontsize=fs
        )
        # add tick lines
        tax.line(
            [i , -3, 90],
            [i , 0, 90],
            color='k'
        )
    
    # add right ticks
    for i in tick_locs:
        tax.annotate(
            text=str(i),
            position=[105-i, i-2, 0],
            rotation=0,
            fontsize=fs
        )
        # add tick lines
        tax.line(
            [100-i , i, 0],
            [103-i , i, 0],
            color='k'
        )
    leg = tax.legend(bbox_to_anchor=(1., 1.1), loc='upper right', frameon=False)
    #leg.get_frame().set_linewidth(4)
    #tax.ticks(axis='lbr', multiple=mul, linewidth=1)
    tax.get_axes().axis('off')
    tax.clear_matplotlib_ticks()
    tax._redraw_labels()
    plt.savefig(f'/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/{fname}_{met1}vs{met2}vs{met3}_pareto_triangle_plot.png')

def read_sampling_objFlux(
        input_df='',
        path='/nfs/turbo/umms-csriram/daweilin/fluxPrediction/RandomObjCoef/embryo_ksom_7/',
        medium='KSOM'
        ):

    # Objective coef sampling
    if len(input_df):
        res = input_df
    else:
        # path to regression results
        flux_paths = {
                'RO':path,
                }
        # get objective fluxes
        res = load_multiObj_models(
                flux_paths['RO'], medium=medium,
                return_variables=False, norm=False, ind_labels=True,
                CFR_paraScan=True, CFR_k=[10,1,0.1,0.01,0.001], CFR_r=[10,1,0.1,0.01,0.001],
                file_suffix='_fluxes.csv.gz'
                )
    res = res[res.any(axis=1)]
    res.index = pd.Series(res.index).apply(lambda x: x.split('[')[0])
    res = res.groupby(res.index).sum().T
    res['cellType'] = ['Simulation']*len(res.index)
    res.iloc[:,:-1] = res.iloc[:,:-1].div(res.iloc[:,:-1].sum(axis=1), axis=0).fillna(0)


    return res

def find_pareto(res, met1, met2, met3='', plotting=False):
    # import package
    import oapackage
    # initiate object
    pareto = oapackage.ParetoDoubleLong()
    if len(met3)>0:
        #met1, met2, met3 = 'gly', 'accoa', 'gthrd'
        for ii in range(0, len(res)):
            w = oapackage.doubleVector(
                    (res[met1].iloc[ii], res[met2].iloc[ii], res[met3].iloc[ii])
                    )
            pareto.addvalue(w, ii)
        
        pareto.show(verbose=1)
        
        # the indices of the Pareto optimal designs
        lst = pareto.allindices()
        print(lst, met1, met2, met3)
        # get optimal points
        optimal_res = res[[met1, met2, met3]].iloc[[i for i in lst], :]
        
        plot_res = res[[met1, met2, met3]]
        plot_res['cellType'] = ['Feasible']*len(plot_res)
        optimal_res['cellType'] = ['Pareto']*len(optimal_res)
        merge_res = pd.concat((plot_res, optimal_res), axis=0)
    else:
        #met1, met2 = 'gly', 'accoa'
        for ii in range(0, len(res)):
            w = oapackage.doubleVector(
                    (res[met1].iloc[ii], res[met2].iloc[ii])
                    )
            pareto.addvalue(w, ii)
        
        pareto.show(verbose=1)
        
        # the indices of the Pareto optimal designs
        lst = pareto.allindices() 
        print(lst, met1, met2)
        # get optimal points
        optimal_res = res[[met1, met2]].iloc[[i for i in lst], :]
        
        plot_res = res[[met1, met2]]
        plot_res['cellType'] = ['Feasible']*len(plot_res)
        optimal_res['cellType'] = ['Pareto']*len(optimal_res)
        merge_res = pd.concat((plot_res, optimal_res), axis=0)
        # test plot
        if plotting==True:
            fig, ax = plt.subplots(1,1,figsize=(12,8))
            h = sns.scatterplot(
                    x=met1, y=met2, data=merge_res, ax=ax, hue='cellType'
                    )
            plt.savefig('/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/pareto_test.png')
        print(optimal_res)
    
    return optimal_res, merge_res


def ObjectiveTradeOffs(
        merge_dff, prefix, cellType_col, sampling_objFlux='', corrplot=True,
        save_root_path='/home/daweilin/StemCell/Project_mESC_JinZhang/regressor_results_new/',
        compare_mets=['gthrd', 'gthox'],
        cellType_curvefit=True, 
        theory_ref=False,
        pareto_line='curvefit',
        hue_order=[],
        ):
    
    from scipy.optimize import curve_fit
    # define the true objective function
    def objective(x, a, c):
        #return a * x + b * x**2 + c
        return a * x + c

    def pareto_front(x, a, b, c):
        return a * x + b * x**2 + c
        #return a * x + c
    
    if corrplot==True:
        from scipy.spatial import distance
        from scipy.cluster import hierarchy
        # Part 1.
        # Correlation plot
        # two objectives form a tradeoff if they are negatively correlated otherwise synergistic
        cmap = sns.diverging_palette(220, 20, as_cmap=True)
        corr_df = merge_dff[merge_dff.columns[merge_dff.any(axis=0)]].iloc[:,:-1].corr()
        #corr_df[corr_df<0] = corr_df[corr_df<0].div(np.abs(corr_df.min().min()))
        sns.set_context("notebook", font_scale=2.)
        sns.clustermap(corr_df, cmap=cmap, center=0, figsize=(30, 30))
        plt.savefig(f'{save_root_path}/{prefix}_objCoef_corr.png')
        
        # Zoom-in correlation plot
        # two objectives form a tradeoff if they are negatively correlated otherwise synergistic
        cmap = sns.diverging_palette(220, 20, as_cmap=True)
        sns.set_context("notebook", font_scale=2.)
        met_arr = corr_df.index[(corr_df<-0.5).any(axis=1)]
        # remove rows
        corr_df = corr_df[(corr_df<-0.5).any(axis=1)]
        # remove columns
        corr_df = corr_df[met_arr]
        #sns.clustermap(corr_df, cmap=cmap, center=0, figsize=(18, 18))
        #plt.savefig(f'{save_root_path}/{prefix}_objCoef_corr_zoomin.png')

        # calculate linkages
        correlations_array = np.asarray(corr_df)
        row_linkage = hierarchy.linkage(
            distance.pdist(correlations_array), method='average')
        col_linkage = hierarchy.linkage(
            distance.pdist(correlations_array.T), method='average')
        fig, axtmp = plt.subplots(1, 1, figsize=(8, 18))
        dn = hierarchy.dendrogram(col_linkage, ax=axtmp, orientation='right')# color_threshold=2)
        plt.savefig(f'{save_root_path}/{prefix}_objCoef_corr_dendrogram_zoomin.png')
        
        # label colors for clusters
        leaves_clusters = dn['leaves_color_list']
        network_pal = sns.cubehelix_palette(n_colors=len(np.unique(leaves_clusters)),
                                            light=.9, dark=.1, reverse=True,
                                            start=1, rot=-2)
        network_lut = dict(zip(map(str, np.unique(leaves_clusters)), network_pal))
        dn_mets_map = {met:network_lut[lc] for lc, met in zip(leaves_clusters, corr_df.index[dn['leaves']])}
        network_colors = pd.Series(corr_df.index.to_numpy()).map(dn_mets_map)
        network_colors.index = corr_df.index
        
        # clustergram
        sns.clustermap(
                corr_df,
                row_linkage=row_linkage,
                col_linkage=col_linkage,
                cmap=cmap,
                col_colors=network_colors,
                row_colors=network_colors,
                center=0,
                figsize=(18, 18)
                )
        plt.savefig(f'{save_root_path}/{prefix}_objCoef_corr_zoomin.png')    
    
    # Part 2.
    # 2D scatter plots for coefficients
    plot_df = merge_dff.copy()
    

    # normalization
    plot_df.iloc[:,:-1] = plot_df.iloc[:,:-1].div(plot_df.iloc[:,:-1].sum(axis=1), axis=0).fillna(0)
    if theory_ref==True:
        # merge with reference tradeoffs
        res = sampling_objFlux#read_sampling_objFlux()
        plot_df = plot_df[res.columns]

    # rearrange hue_order
    hue_order_pareto = hue_order+['Pareto']
    print(plot_df.columns)
    # go thru all candidates
    for y in compare_mets:
    
        # Figure 5.
        cols = plot_df.columns[(plot_df.columns!=y) & (plot_df.columns!='cellType')]
        for met in cols:

            plot_df_copy = plot_df.copy()
            # figure
            #sns.despine()

            PltProps()
            sns.set_style({'legend.frameon':True, 'figure.facecolor': 'white'})
            fig, ax = plt.subplots(1,1,figsize=(6, 4))
            pal = sns.color_palette('Set2')
            
            if cellType_curvefit:
                for i, ele in enumerate(hue_order):

                    # curve fit
                    popt, _ = curve_fit(objective, plot_df[met][plot_df[cellType_col]==ele], plot_df[y][plot_df[cellType_col]==ele])
                    a, c = popt
                    # define a sequence of inputs between the smallest and largest known inputs
                    x_line = np.linspace(
                            plot_df[met][plot_df[cellType_col]==ele].min(),
                            plot_df[met][plot_df[cellType_col]==ele].max(),
                            100
                            )
                    # calculate the output for the range
                    y_line = objective(x_line, a, c)
                    # create a line plot for the mapping function
                    ax.plot(x_line, y_line, '--', color=pal.as_hex()[i], label=ele)
            
                # curve fit for all celltypes
                popt, _ = curve_fit(objective, plot_df[met], plot_df[y])
                a, c = popt
                # define a sequence of inputs between the smallest and largest known inputs
                x_line = np.linspace(plot_df[met].min(), plot_df[met].max(), 100)
                # calculate the output for the range
                y_line = objective(x_line, a, c)
                # create a line plot for the mapping function
                ax.plot(x_line, y_line, '--', color='k', label='All')
                # change the fontsize
                #ax.tick_params(axis='x', labelsize=20)
                #ax.tick_params(axis='y', labelsize=20)


            # overlaying modeling objective fluxes to data
            if theory_ref==True:
                PltProps()
                # reset figures
                fig, ax = plt.subplots(1,1,figsize=(6, 4))
                pal = sns.color_palette('Set2')
                # merge with reference tradeoffs
                res = sampling_objFlux#read_sampling_objFlux()
                pareto, _ = find_pareto(res, met, y)
                if len(pareto)==0:
                    pareto = res
                placeholder = plot_df.copy()
                plot_df = pd.concat((pareto, plot_df[pareto.columns]), axis=0)
                plot_df[[met, y]] = plot_df[[met, y]].div(plot_df[[met, y]].max(axis=0), axis=1)
                pareto = plot_df[plot_df['cellType']=='Pareto']
                if pareto_line=='connected':
                    pareto = pareto.sort_values(by=[met])
                    ax.plot(pareto[met], pareto[y], '--',
                            color='k',
                            label='Est. Pareto', linewidth=3.)
                else: # fitting curve
                    if len(pareto)>2:
                        # curve fit for all celltypes
                        popt, _ = curve_fit(pareto_front, pareto[met], pareto[y])
                        a, b, c = popt
                        # define a sequence of inputs between the smallest and largest known inputs
                        x_line = np.linspace(pareto[met].min(), pareto[met].max(), 100)
                        # calculate the output for the range
                        y_line = pareto_front(x_line, a, b, c)
                        # create a line plot for the mapping function
                        ax.plot(x_line, y_line, '--',
                                color='k',
                                label='Est. Pareto', linewidth=3.)
                                   
                    elif len(pareto)==2:
                        # curve fit for all celltypes
                        popt, _ = curve_fit(objective, pareto[met], pareto[y])
                        a, c = popt
                        # define a sequence of inputs between the smallest and largest known inputs
                        x_line = np.linspace(pareto[met].min(), pareto[met].max(), 100)
                        # calculate the output for the range
                        y_line = objective(x_line, a, c)
                        # create a line plot for the mapping function
                        ax.plot(x_line, y_line, '--',
                                color='k',
                                label='Est. Pareto', linewidth=3.)

                    else:
                        print('Not enough data points')
                # scatter plot of tradeoffs
                sns.scatterplot(
                        x=met,
                        y=y, #hue
                        data=plot_df,
                        legend=True,
                        s=100,
                        palette=sns.color_palette('Set2')[:len(hue_order)]+['k'],
                        hue='cellType',
                        hue_order=hue_order_pareto,
                        alpha=0.5,
                        ax=ax,
                        zorder=0
                        )
                # point plots of average trade-offs
                mean_df = plot_df.groupby(by='cellType').mean()
                print('mean-df')
                print(mean_df)
                sns.scatterplot(
                        x=met,
                        y=y, #hue
                        data=mean_df[mean_df.index!='Pareto'],
                        legend=False,
                        s=300,
                        marker='X',
                        palette='Dark2',
                        hue=mean_df[mean_df.index!='Pareto'].index,
                        hue_order=hue_order,
                        ax=ax,
                        zorder=1,
                        )

                # setup title
                #ax.set_title(f'{prefix} Objective Tradeoffs', fontsize=20)
                ax.set_xlim([-0.05, plot_df[met].max()+0.05])
                ax.set_ylim([-0.05, plot_df[y].max()+0.05])
                ax.set_xticks(np.linspace(0, plot_df[met].max(), 6))
                ax.set_yticks(np.linspace(0, plot_df[y].max(), 6))
                CanvasStyle(ax, lw=4, ticks_lw=3)
                ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left', facecolor='w', frameon=False)
                #ax.set_ylim([-0.005, 0.05])
                plt.tight_layout()
                # save the figures
                plt.savefig(f'{save_root_path}/{prefix}_obj_overlay_tradeoff2D_{met}vs{y}.png')
                plot_df = placeholder.copy()
            else:

                # scatter plot
                sns.scatterplot(
                        x=met,
                        y=y, #hue
                        data=plot_df,
                        legend=True,
                        s=100,
                        color='grey',
                        #palette='Blues',#sns.color_palette("dark:#69d",as_cmap=True),
                        alpha=0.5,
                        ax=ax,
                        )
                # setup title
                #ax.set_title(f'{prefix} Objective Tradeoffs', fontsize=20)
                ax.set_xlim([-plot_df[met].max()/100, plot_df[met].max()*1.01])
                ax.set_ylim([-plot_df[y].max()/100, plot_df[y].max()*1.01])
                #ax.set_xticks(np.linspace(0, plot_df[met].max(), 6))
                #ax.set_yticks(np.linspace(0, plot_df[y].max(), 6))
                CanvasStyle(ax, lw=4, ticks_lw=3)
                ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left', facecolor='w', frameon=False)
                plt.tight_layout()
                # save the figures
                plt.savefig(f'{save_root_path}/{prefix}_obj_tradeoff2D_{met}vs{y}.png')
            
            # recover
            plot_df = plot_df_copy.copy()


    
"""


███████╗████████╗ █████╗ ████████╗███████╗                   
██╔════╝╚══██╔══╝██╔══██╗╚══██╔══╝██╔════╝                   
███████╗   ██║   ███████║   ██║   ███████╗                   
╚════██║   ██║   ██╔══██║   ██║   ╚════██║                   
███████║   ██║   ██║  ██║   ██║   ███████║                   
╚══════╝   ╚═╝   ╚═╝  ╚═╝   ╚═╝   ╚══════╝                   
                                                             
███╗   ███╗███████╗████████╗██╗  ██╗ ██████╗ ██████╗ ███████╗
████╗ ████║██╔════╝╚══██╔══╝██║  ██║██╔═══██╗██╔══██╗██╔════╝
██╔████╔██║█████╗     ██║   ███████║██║   ██║██║  ██║███████╗
██║╚██╔╝██║██╔══╝     ██║   ██╔══██║██║   ██║██║  ██║╚════██║
██║ ╚═╝ ██║███████╗   ██║   ██║  ██║╚██████╔╝██████╔╝███████║
╚═╝     ╚═╝╚══════╝   ╚═╝   ╚═╝  ╚═╝ ╚═════╝ ╚═════╝ ╚══════╝
                                                             

---                                                                  
Statistical methods
- pareto_surface_3d
- triangle_plot
- read_sampling_objFlux
- find_pareto
- ObjectiveTradeOffs

"""

# right-tailed hypergeometric tests
def hypergeom_test(M, N, n, k):
    """
    arr1: array with only 1, 0, -1
    arr2: array with only 1, 0, -1
    """
    from scipy.stats import hypergeom
    #M = len(arr1)
    #N = sum(arr1==label)
    #n = sum(arr2==label)
    #k = sum((arr1==label)*(arr2==label))
    # more overlaps higher pvalues
    hypergeom_p = hypergeom.sf(k-1, M, n, N)
    a, b, c, d = k, n-k, N-k, M-(n+N)+k
    table = np.array([[a, b], [c, d]])
    start, end = hypergeom.support(M, n, N)
    # print(sum(hypergeom.pmf(np.arange(k, end+1), M, n, N)))
    from scipy.stats import fisher_exact
    oddsr, fisher_p = fisher_exact(table, alternative='greater')
    # print(k-1, M, n, N, fisher_p, hypergeom_p)
    return hypergeom_p, fisher_p

# left-tailed rank sum test
# WARNING: might need to confirm the method
def ranksumtest(arr1, arr2, alternative='less'):
    from scipy.stats import ranksums
    w, p = ranksums(arr1, arr2, alternative=alternative)
    return w, p


# read the metabolic network model and output reorganized dictionary or dataframe
def read_mat_model(path='/home/daweilin/StemCell/cancer_model.mat', return_rxnGeneMat=False, remove_transport=False):
    
    """
    Description
    -----------
    1. read genome-scale metabolic network models (GEMs) from .mat file
    2. reorganize the model and save into a python dictionary
    3. (optional) gather subsystems, reacitons, and genes with gpr rules
    4. (optional) remove transport and exchange reactions for internal reactions analysis

    Arguments
    ---------
    path (string): the path to access the network model
    return_rxnGeneMat (bool): if output a binary dataframe showing association of GPR rules
        where columns are genes, additional columns for reactions and subsystems
    remove_transport (bool): if remove reactions related to exchange/transport subsystems

    Returns
    -------
    model_dict (dictionary): fields of models and corresponding values, e.g. 'genes':['A', 'B'...]
    rxnGeneMat (optional, pandas.DataFrame):a dataframe with binary values showing association
        between genes and reactions where column names are genes,
        additional two columns for reactions and subsystems


    """


    import scipy.io
    # load model
    mat = scipy.io.loadmat(path)
    mat = mat[[k for k in mat.keys()][-1]]
    keys = mat.dtype.fields.keys()
    arrs = mat.item()
    # reorganize the model
    model_dict = {k:arr for k, arr in zip(keys, arrs)}

    # get specific fields, including genes, reactions, and subsystems
    if return_rxnGeneMat==True:
        rxnGeneMat = pd.DataFrame.sparse.from_spmatrix(mat['rxnGeneMat'][0][0])
        rxnGeneMat.columns = ['unknown' if len(g[0])==0 else g[0][0] for g in mat['genes'][0][0]] #genes
        rxnGeneMat['rxns'] = ['unknown' if len (r[0])==0 else r[0][0] for r in mat['rxns'][0][0]]
        rxnGeneMat['subSystems'] = ['unknown' if len (s[0])==0 else s[0][0] for s in mat['subSystems'][0][0]]
    
        # remove transport and exchange reactions
        if remove_transport==True:
            rxnGeneMat = rxnGeneMat[rxnGeneMat['subSystems'].str.contains('Transport|Exchange')==0]

        return model_dict, rxnGeneMat
    
    else:
        return model_dict



# modified zscores
def RobustZScore(arr, lb=10**(-5), ub=30):
    
    ind, = np.nonzero(np.abs(arr)<lb)
    if len(ind)>0:
        arr[ind] = np.zeros(len(arr[ind]))
    
    STD = np.std(arr) if np.std(arr)!=0 else 1
    res = (arr-np.mean(arr))/STD
    ind2, = np.nonzero(np.abs(res)>ub)
    print('1')
    print(np.max(np.abs(res)))
    if len(ind2)>0:
        print('2')
        sign = arr[ind2]/np.abs(arr[ind2])
        print(sign)
        ind3, = np.nonzero(np.abs(res)<=ub)
        tmp = res[ind3]
        res[ind2] = sign*np.max(np.abs(tmp))
        print('3')
        print(np.max(np.abs(res)))
    if len(arr)==len(res):
        return res
